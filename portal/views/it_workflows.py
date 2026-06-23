from __future__ import annotations

from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from academics.models import AcademicClass, EC, Semester, UE
from academics.services.semester import compute_semester_result
from accounts.access import get_user_position
from accounts.dashboards.helpers import get_user_branch
from portal.selectors import (
    get_it_academic_classes,
    get_it_semesters_for_class,
    get_it_students_for_class,
)
from portal.services.academic_structure_service import (
    archive_academic_class,
    assign_student_to_class,
    build_academic_structure_context,
    delete_ec,
    save_academic_class,
    save_ec,
    save_ue,
)
from portal.services.archive_service import (
    archive_batches_for_branch,
    archive_candidates_for_branch,
    archive_class,
    archive_detail,
    archive_year,
    preview_archive,
    restore_archive_batch,
)
from portal.services.notes_workflow import (
    apply_notes_workflow_action,
    get_available_actions,
    get_notes_state,
    get_retake_candidates,
)
from portal.services.it_support_service import get_account_support_state, get_scoped_staff_queryset
from portal.services.it_support_service import create_temp_password, log_support_action
from portal.services.informaticien_workflows import (
    build_audit_context,
    build_catalog_context,
    build_home_context,
    build_import_context,
    build_supervision_context,
    build_support_context,
    create_branch_ticket,
    create_catalog_item,
    get_branch_settings,
    import_notes_file,
    resolve_branch_ticket,
    take_branch_ticket,
    update_branch_settings,
)
from portal.selectors.informaticien import grade_entries_for_class, support_tickets_for_branch
from portal.models import SupportAuditLog, SupportTicket
from students.models import Student
from portal.views.admin_grades import _build_notes_grid_context
from communication.models.notifications import CommunicationNotification


def _require_it_support(request):
    if get_user_position(request.user) != "it_support":
        return False
    if get_user_branch(request.user) is None:
        return False
    return True


def _render_it_section(request, module_key, template_name, context):
    """
    Rend le fragment d'une section du dashboard informaticien.

    Si la requete vient de HTMX (navigation interne), seul le fragment est
    renvoye. Sinon (navigation directe, F5, retour/avance navigateur), la
    coquille complete du dashboard est reconstruite avec ce meme fragment
    deja insere dans #it-workspace, pour eviter une page cassee.
    """
    if getattr(request, "htmx", False):
        return render(request, template_name, context)
    fragment = render(request, template_name, context)
    from portal.views.views import _render_it_dashboard
    return _render_it_dashboard(
        request,
        initial_module=module_key,
        initial_workspace_html=fragment.content.decode(fragment.charset or "utf-8"),
    )


def _same_branch_or_forbidden(*, request, target_user):
    request_branch = get_user_branch(request.user)
    target_branch = getattr(getattr(target_user, "profile", None), "branch", None)
    if request_branch is None or target_branch != request_branch:
        return False
    return True


def _resolve_workflow_selection(request):
    branch = get_user_branch(request.user)
    classes_qs = get_it_academic_classes(branch=branch)
    selected_class = None
    selected_semester = None
    semesters = []

    class_id = (request.GET.get("class_id") or request.GET.get("classe") or request.POST.get("class_id") or request.POST.get("classe") or "").strip()
    semester_id = (request.GET.get("semester_id") or request.GET.get("semester") or request.POST.get("semester_id") or request.POST.get("semester") or "").strip()
    if class_id.isdigit():
        selected_class = get_object_or_404(classes_qs, pk=int(class_id))
        semesters = list(get_it_semesters_for_class(academic_class=selected_class))

    if selected_class and semester_id.isdigit():
        selected_semester = (
            Semester.objects.select_related("academic_class")
            .filter(pk=int(semester_id), academic_class=selected_class)
            .first()
        )

    return {
        "branch": branch,
        "classes": list(classes_qs[:200]),
        "selected_class": selected_class,
        "semesters": semesters,
        "selected_semester": selected_semester,
    }


def _build_notes_workflow_context(request, *, toast=None):
    context = _resolve_workflow_selection(request)
    state = get_notes_state(
        academic_class=context["selected_class"],
        semester=context["selected_semester"],
    )
    context.update(
        {
            "workflow_module": "notes",
            "state": state,
            "actions": get_available_actions(state=state),
            "toast": toast,
        }
    )
    return context


@login_required
def it_notes_kpi(request):
    if get_user_position(request.user) != "it_support":
        return HttpResponseForbidden("Acces refuse.")

    from academics.models import AcademicEnrollment, ECGrade, AcademicDebt

    from portal.services.notes_workflow import get_notes_state

    branch = get_user_branch(request.user)
    classes_qs = get_it_academic_classes(branch=branch)

    class_id = (request.GET.get("class_id") or "").strip()
    semester_id = (request.GET.get("semester_id") or "").strip()

    academic_class = get_object_or_404(classes_qs, pk=int(class_id)) if class_id.isdigit() else None
    semester = get_object_or_404(Semester.objects.select_related("academic_class"), pk=int(semester_id)) if semester_id.isdigit() and academic_class else None

    if not academic_class or not semester:
        return HttpResponse("Selection invalide", status=400)

    enrollments = AcademicEnrollment.objects.filter(
        academic_class=academic_class,
        academic_year=academic_class.academic_year,
        is_active=True,
    )
    enrollment_count = enrollments.count()

    ec_count = semester.ues.aggregate(total=Count("ecs", distinct=True))["total"] or 0
    expected_grades = enrollment_count * ec_count

    grade_stats = ECGrade.objects.filter(
        enrollment__in=enrollments,
        ec__ue__semester=semester,
    ).aggregate(
        entered_grades=Count("id", filter=Q(normal_score__isnull=False)),
        validated_grades=Count("id", filter=Q(is_validated=True)),
        failed_grades=Count("id", filter=Q(final_score__isnull=False, is_validated=False)),
    )
    entered_grades = grade_stats["entered_grades"]
    validated_grades = grade_stats["validated_grades"]
    failed_grades = grade_stats["failed_grades"]

    pending_debts = AcademicDebt.objects.filter(
        academic_class=academic_class,
        semester=semester,
        status="pending",
    ).count()

    progress = int((entered_grades / expected_grades) * 100) if expected_grades else 0

    state = get_notes_state(academic_class=academic_class, semester=semester)

    context = {
        "enrollment_count": enrollment_count,
        "entered_grades": entered_grades,
        "expected_grades": expected_grades,
        "validated_grades": validated_grades,
        "failed_grades": failed_grades,
        "pending_debts": pending_debts,
        "progress": min(progress, 100),
        "state_code": state.code if state else "empty",
        "state_label": state.label if state else "Non commence",
        "ec_count": ec_count,
        "class_id": class_id,
        "semester_id": semester_id,
    }

    return render(request, "portal/informaticien/partials/notes_kpi.html", context)


@login_required
def it_workflow_section(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    context = _build_notes_workflow_context(request)
    academic_class = context.get("selected_class")
    semester = context.get("selected_semester")
    state = context.get("state")
    if not academic_class or not semester or state is None:
        return HttpResponse("Selection invalide", status=400)
    from ui.components.notes.notes_workflow_bar import build_workflow_bar_data
    has_candidates = bool(state.retake_candidates_count) if state else None
    workflow_bar = build_workflow_bar_data(state.code, has_candidates=has_candidates)
    return render(request, "notes/notes_workflow_section.html", {
        "academic_class": academic_class,
        "semester": semester,
        "state": state,
        "actions": context.get("actions", []),
        "workflow_bar": workflow_bar,
        "drawer_mode": bool(request.GET.get("drawer")),
    })


@login_required
def it_notes_decisions(request):
    if get_user_position(request.user) != "it_support":
        return HttpResponseForbidden("Acces refuse.")

    from portal.selectors import get_it_academic_classes
    from academics.models import AcademicEnrollment, AcademicDebt
    from academics.services.year import (
        DECISION_ADMISSIBLE,
        DECISION_NON_ADMIS,
        DECISION_VALIDE,
        compute_annual_decision,
    )

    branch = get_user_branch(request.user)
    classes_qs = get_it_academic_classes(branch=branch)

    class_id = (request.GET.get("class_id") or "").strip()

    academic_class = get_object_or_404(classes_qs, pk=int(class_id)) if class_id.isdigit() else None

    if not academic_class:
        return HttpResponse("", status=200)

    enrollments = AcademicEnrollment.objects.filter(
        academic_class=academic_class,
        academic_year=academic_class.academic_year,
        is_active=True,
    ).select_related("student__student_profile__inscription__candidature")

    decision_labels = {
        DECISION_VALIDE: "VALIDÉ",
        DECISION_ADMISSIBLE: "ADMISSIBLE",
        DECISION_NON_ADMIS: "NON ADMIS",
    }

    decisions = []
    for enrollment in enrollments:
        result = compute_annual_decision(enrollment)
        semester_validated = {
            sr["semester"]: sr.get("is_validated")
            for sr in result.get("semester_results", [])
        }

        def _semester_label(key):
            validated = semester_validated.get(key)
            if validated is None:
                return "—"
            return "VALIDÉ" if validated else "NON ADMIS"

        active_debts = AcademicDebt.objects.filter(
            enrollment=enrollment,
            status="pending",
        ).select_related("ec")

        debt_names = [f"{d.ec.title} (S{d.semester.number})" for d in active_debts]

        decisions.append({
            "student_name": enrollment.student.student_profile.full_name,
            "s1": _semester_label("S1"),
            "s2": _semester_label("S2"),
            "year": decision_labels.get(result["decision"], result["decision"]),
            "debts": debt_names,
        })

    active_debts_count = AcademicDebt.objects.filter(
        academic_class=academic_class,
        status="pending",
    ).count()

    return render(request, "portal/informaticien/partials/notes_decisions.html", {
        "decisions": decisions,
        "debts_count": active_debts_count,
    })


@login_required
def it_notes_flow_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    context = _build_notes_workflow_context(request)
    if request.GET.get("drawer") == "1":
        return render(
            request,
            "portal/informaticien/drawers/notes_drawer.html",
            context,
        )
    return _render_it_section(
        request,
        "notes",
        "portal/informaticien/workflows/notes_workspace.html",
        context,
    )


@login_required
def it_home_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    return _render_it_section(
        request,
        "home",
        "portal/informaticien/workflows/home_workspace.html",
        build_home_context(branch=get_user_branch(request.user)),
    )


@login_required
def it_notes_workflow_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")

    context = _resolve_workflow_selection(request)
    toast = None
    successful = False
    try:
        action = (request.POST.get("action") or "").strip()
        apply_notes_workflow_action(
            actor=request.user,
            academic_class=context["selected_class"],
            semester=context["selected_semester"],
            action=action,
        )
        successful = True
        messages = {
            "verifier_notes": "Controle termine: aucune note manquante bloquante.",
            "publier_session_normale": "Session normale publiee. Le rattrapage peut maintenant etre prepare.",
            "activer_rattrapage": "Rattrapage active. Seules les notes de rattrapage restent modifiables.",
            "publier_resultats_finaux": "Resultats finaux publies. Les releves et exports sont deblocables.",
            "generer_decisions_annuelles": "Decisions annuelles generees avec bulletins.",
            "generer_bulletins": "Bulletins semestriels generes.",
        }
        toast = {"level": "success", "message": messages.get(action, "Workflow notes mis a jour.")}
    except ValidationError as exc:
        if request.POST.get("from_modal") == "retake":
            modal_context = _build_retake_modal_context(request)
            modal_context["form_error"] = " ".join(exc.messages)
            response = render(request, "portal/informaticien/workflows/retake_modal.html", modal_context)
            response["HX-Retarget"] = "#it-modal-root"
            return response
        toast = {"level": "error", "message": " ".join(exc.messages)}

    template_name = (
        "portal/informaticien/drawers/notes_drawer.html"
        if request.POST.get("drawer") == "1"
        else "portal/informaticien/workflows/notes_workspace.html"
    )
    response = render(
        request,
        template_name,
        _build_notes_workflow_context(request, toast=toast),
    )
    if successful:
        response["HX-Trigger"] = '{"it-modal-close": "", "kpi-update": "", "workflow-update": ""}'
    return response


def _build_retake_modal_context(request):
    context = _resolve_workflow_selection(request)
    academic_class = context["selected_class"]
    semester = context["selected_semester"]
    state = get_notes_state(academic_class=academic_class, semester=semester)
    candidates = get_retake_candidates(academic_class=academic_class, semester=semester)
    return {
        "academic_class": academic_class,
        "semester": semester,
        "state": state,
        "candidates": candidates,
        "subject_count": sum(len(candidate.failed_subjects) for candidate in candidates),
    }


@login_required
def load_notes_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    context = _resolve_workflow_selection(request)
    academic_class = context["selected_class"]
    semester = context["selected_semester"]
    if academic_class is None or semester is None:
        return HttpResponse(
            '<div class="workflow-grid-placeholder">Selectionne une classe et un semestre pour afficher la maquette.</div>'
        )
    if semester.status == Semester.STATUS_DRAFT:
        semester.status = Semester.STATUS_NORMAL_ENTRY
        semester.save(update_fields=["status"])

    grid_context = _build_notes_grid_context(
        academic_class=academic_class,
        semester=semester,
        requested_session_type=request.GET.get("session", "normal"),
    )
    return render(
        request,
        "portal/admin/grades/partials/notes_maquette.html",
        {
            **grid_context,
            "embedded_in_dashboard": True,
        },
    )


@login_required
def it_cards_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    selection = _resolve_workflow_selection(request)
    students = []
    if selection["selected_class"]:
        students = list(get_it_students_for_class(academic_class=selection["selected_class"])[:80])
    selection.update({"students": students, "workflow_module": "cards"})
    return _render_it_section(request, "cards", "portal/informaticien/workflows/cards_workspace.html", selection)


@login_required
def it_support_flow_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    status = (request.GET.get("status") or "").strip()
    return _render_it_section(
        request,
        "support",
        "portal/informaticien/workflows/support_workspace.html",
        build_support_context(branch=branch, status=status, page=request.GET.get("page")),
    )


@login_required
def it_support_flow_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")

    branch = get_user_branch(request.user)
    action = (request.POST.get("action") or "").strip()
    toast = None
    try:
        if action == "create":
            create_branch_ticket(
                actor=request.user,
                branch=branch,
                title=request.POST.get("title"),
                description=request.POST.get("description"),
            )
            toast = {"level": "success", "message": "Ticket cree."}
        else:
            ticket = get_object_or_404(support_tickets_for_branch(branch=branch), pk=request.POST.get("ticket_id"))
            if action == "take":
                take_branch_ticket(actor=request.user, branch=branch, ticket=ticket)
                toast = {"level": "success", "message": "Ticket pris en charge."}
            elif action == "resolve":
                resolve_branch_ticket(
                    actor=request.user,
                    branch=branch,
                    ticket=ticket,
                    resolution=request.POST.get("resolution"),
                )
                toast = {"level": "success", "message": "Ticket resolu."}
            else:
                toast = {"level": "error", "message": "Action ticket inconnue."}
    except (ValidationError, ValueError) as exc:
        message = " ".join(exc.messages) if hasattr(exc, "messages") else str(exc)
        toast = {"level": "error", "message": message}

    return render(
        request,
        "portal/informaticien/workflows/support_workspace.html",
        build_support_context(branch=branch, status=request.POST.get("status", ""), toast=toast),
    )


@login_required
def it_audit_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    return _render_it_section(
        request,
        "audit",
        "portal/informaticien/workflows/audit_workspace.html",
        build_audit_context(branch=get_user_branch(request.user), page=request.GET.get("page")),
    )


def _resolve_import_selection(request):
    selection = _resolve_workflow_selection(request)
    return selection["classes"], selection["selected_class"], selection["selected_semester"]


def _paginate_items(items, page_number, *, per_page):
    paginator = Paginator(items, per_page)
    try:
        return paginator.page(page_number or 1)
    except (EmptyPage, PageNotAnInteger):
        return paginator.page(1)


def _build_structure_drawer_context(*, branch, class_id="", semester_id="", ue_id="", ec_id="", section="maquettes", ue_page=1, ec_page=1, ec_ue_id=""):
    actual_class = (
        AcademicClass.objects.select_related("programme", "academic_year", "branch")
        .filter(pk=class_id, branch=branch, is_archived=False)
        .first()
        if class_id
        else None
    )
    context = build_academic_structure_context(
        branch=branch,
        selected_class_id=actual_class.id if actual_class else None,
        section=section,
    )
    selected_class = actual_class
    context["selected_class"] = selected_class
    selected_semester = None
    selected_ue = None
    selected_ec = None

    if selected_class is not None and semester_id:
        selected_semester = Semester.objects.filter(
            pk=semester_id,
            academic_class__branch=branch,
            academic_class=selected_class,
        ).first()
    elif selected_class is not None:
        selected_semester = selected_class.semesters.order_by("number").first()

    if ue_id:
        selected_ue = UE.objects.filter(
            pk=ue_id,
            semester__academic_class__branch=branch,
        ).select_related("semester", "semester__academic_class").first()

    if ec_id:
        selected_ec = EC.objects.filter(
            pk=ec_id,
            ue__semester__academic_class__branch=branch,
        ).select_related("ue", "ue__semester", "ue__semester__academic_class").first()

    if selected_ec is not None and selected_ue is None:
        selected_ue = selected_ec.ue
    if selected_ue is not None and selected_semester is None:
        selected_semester = selected_ue.semester

    semesters = list(selected_class.semesters.prefetch_related("ues__ecs").order_by("number")) if selected_class else []
    ue_pairs = []
    ue_rows = []
    ec_count = 0
    for semester in semesters:
        semester_ues = list(semester.ues.all().order_by("code", "id"))
        ue_pairs.extend((semester, ue) for ue in semester_ues)
        ec_count += sum(len(ue.ecs.all()) for ue in semester_ues)
    drawer_ues_page = _paginate_items(ue_pairs, ue_page, per_page=8) if ue_pairs else None
    row_map = {}
    for semester, ue in (drawer_ues_page.object_list if drawer_ues_page else []):
        ec_list = list(ue.ecs.all().order_by("title", "id"))
        active_ec_page = str(ec_ue_id or "") == str(ue.id)
        ecs_page = _paginate_items(ec_list, ec_page if active_ec_page else 1, per_page=10) if ec_list else None
        row_map.setdefault(semester.id, {"semester": semester, "ues": []})["ues"].append(
            {
                "ue": ue,
                "ecs_page": ecs_page,
                "ec_count": len(ec_list),
                "active_ec_page": active_ec_page,
            }
        )
    ue_rows = list(row_map.values())

    context.update(
        {
            "drawer_kind": "class",
            "drawer_class": selected_class,
            "drawer_semesters": semesters,
            "drawer_ue_rows": ue_rows,
            "drawer_ues_page": drawer_ues_page,
            "drawer_ec_page": ec_page,
            "drawer_ec_ue_id": ec_ue_id,
            "drawer_selected_semester": selected_semester,
            "drawer_selected_ue": selected_ue,
            "drawer_selected_ec": selected_ec,
            "drawer_ec_count": ec_count,
            "drawer_url": (
                f"{reverse('accounts_portal:it_structure_drawer')}?class_id={selected_class.id}&section={section}"
                + (f"&semester_id={selected_semester.id}" if selected_semester else "")
                + (f"&ue_id={selected_ue.id}" if selected_ue else "")
                + (f"&ec_id={selected_ec.id}" if selected_ec else "")
            ) if selected_class else reverse("accounts_portal:it_structure_drawer"),
        }
    )
    return context


@login_required
def it_import_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    classes, selected_class, selected_semester = _resolve_import_selection(request)
    return _render_it_section(
        request,
        "import",
        "portal/informaticien/workflows/import_workspace.html",
        build_import_context(
            branch=branch,
            classes=classes,
            selected_class=selected_class,
            selected_semester=selected_semester,
        ),
    )


@login_required
def it_import_upload(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")
    branch = get_user_branch(request.user)
    classes, selected_class, selected_semester = _resolve_import_selection(request)
    feedback = None
    upload = request.FILES.get("file")
    try:
        if selected_class is None or selected_semester is None:
            raise ValidationError("Selectionne une classe et un semestre.")
        if upload is None:
            raise ValidationError("Fichier Excel obligatoire.")
        feedback = import_notes_file(
            actor=request.user,
            branch=branch,
            academic_class=selected_class,
            semester=selected_semester,
            file=upload,
        )
    except (ValidationError, ValueError) as exc:
        message = " ".join(exc.messages) if hasattr(exc, "messages") else str(exc)
        feedback = type("Feedback", (), {
            "level": "error",
            "message": message,
            "invalid_lines": [],
            "updated": 0,
            "skipped_empty": 0,
            "skipped_unknown_columns": 0,
            "skipped_unknown_students": 0,
            "skipped_invalid_scores": 0,
            "unknown_columns": [],
        })()
    return render(
        request,
        "portal/informaticien/workflows/import_workspace.html",
        build_import_context(
            branch=branch,
            classes=classes,
            selected_class=selected_class,
            selected_semester=selected_semester,
            feedback=feedback,
        ),
    )


@login_required
def it_export_notes_excel(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl doit etre installe pour exporter en Excel.", status=500)
    selection = _resolve_workflow_selection(request)
    academic_class = selection["selected_class"]
    selected_semester = selection["selected_semester"]
    if academic_class is None:
        return HttpResponseForbidden("Classe obligatoire.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes classe"
    sheet.append(["Classe", "Semestre", "Etudiant", "EC", "Coefficient", "Credits", "Session normale", "Rattrapage", "Note finale", "Valide"])
    grades = grade_entries_for_class(academic_class=academic_class)
    if selected_semester:
        grades = grades.filter(ec__ue__semester=selected_semester)
    for grade in grades.order_by("enrollment__student__username", "ec__ue__semester__number", "ec__title"):
        sheet.append([
            academic_class.display_name,
            f"S{grade.ec.ue.semester.number}",
            grade.enrollment.student.get_full_name() or grade.enrollment.student.username,
            grade.ec.title,
            grade.ec.coefficient,
            grade.ec.credit_required,
            grade.normal_score,
            grade.retake_score,
            grade.final_score,
            "oui" if grade.is_validated else "non",
        ])

    anomalies = workbook.create_sheet("Anomalies")
    anomalies.append(["Type", "Detail", "Action attendue"])
    state = get_notes_state(academic_class=academic_class, semester=selected_semester) if selected_semester else None
    if state:
        for alert in state.technical_alerts:
            anomalies.append(["Notes", alert, "Completer la grille puis verifier"])
    if anomalies.max_row == 1:
        anomalies.append(["Aucune", "Aucune anomalie bloquante detectee dans la selection.", "-"])

    results = workbook.create_sheet("Resultats")
    results.append(["Etudiant", "Semestre", "Moyenne", "Pourcentage", "Credits obtenus", "Credits requis", "Statut"])
    if selected_semester:
        enrollments = academic_class.enrollments.filter(is_active=True, academic_year=academic_class.academic_year).select_related("student")
        for enrollment in enrollments:
            summary = compute_semester_result(selected_semester, enrollment)
            results.append([
                enrollment.student.get_full_name() or enrollment.student.username,
                f"S{selected_semester.number}",
                summary["average"],
                summary["percentage"],
                summary["credit_obtained"],
                summary["credit_required"],
                state.label if state else "",
            ])

    for worksheet in workbook.worksheets:
        for column in ("A", "B", "C", "D"):
            worksheet.column_dimensions[column].width = 28
    for column in ("A", "B", "C"):
        sheet.column_dimensions[column].width = 28
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="notes-classe-{academic_class.id}.xlsx"'
    log_support_action(
        actor=request.user,
        branch=get_user_branch(request.user),
        action_type=SupportAuditLog.ACTION_EXCEL_EXPORTED,
        target_label=f"Export notes {academic_class.display_name}",
        details=f"Export Excel {'S' + str(selected_semester.number) if selected_semester else 'classe complete'}.",
    )
    return response


@login_required
def it_structure_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    selected_class_id = (request.GET.get("class_id") or request.GET.get("classe") or "").strip()
    student_query = (request.GET.get("student_q") or "").strip()
    section = (request.GET.get("section") or "classes").strip()
    return _render_it_section(
        request,
        "structure",
        "portal/informaticien/workflows/structure_workspace.html",
        build_academic_structure_context(
            branch=get_user_branch(request.user),
            selected_class_id=selected_class_id,
            student_query=student_query,
            section=section,
            class_page=request.GET.get("class_page") or 1,
            student_page=request.GET.get("student_page") or 1,
        ),
    )


@login_required
def it_structure_drawer(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    branch = get_user_branch(request.user)
    context = _build_structure_drawer_context(
        branch=branch,
        class_id=(request.GET.get("class_id") or request.GET.get("id") or request.GET.get("class") or "").strip(),
        semester_id=(request.GET.get("semester_id") or "").strip(),
        ue_id=(request.GET.get("ue_id") or "").strip(),
        ec_id=(request.GET.get("ec_id") or "").strip(),
        section=(request.GET.get("section") or "maquettes").strip(),
        ue_page=request.GET.get("ue_page") or 1,
        ec_page=request.GET.get("ec_page") or 1,
        ec_ue_id=(request.GET.get("ec_ue_id") or "").strip(),
    )
    return render(request, "portal/informaticien/drawers/structure_drawer.html", context)


@login_required
def it_structure_modal(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    kind = (request.GET.get("kind") or "").strip()
    object_id = (request.GET.get("id") or "").strip()
    class_id = (request.GET.get("class_id") or "").strip()
    section = (request.GET.get("section") or "classes").strip()
    semester_id = (request.GET.get("semester_id") or "").strip()
    ue_id = (request.GET.get("ue_id") or "").strip()
    student_id = (request.GET.get("student_id") or "").strip()

    context = build_academic_structure_context(branch=branch, selected_class_id=class_id, section=section)
    context.update(
        {
            "kind": kind,
            "target_id": object_id,
            "target_class": AcademicClass.objects.filter(pk=object_id, branch=branch).first() if kind == "class" and object_id else None,
            "target_ue": UE.objects.filter(pk=object_id, semester__academic_class__branch=branch).first() if kind == "ue" and object_id else None,
            "target_ec": EC.objects.filter(pk=object_id, ue__semester__academic_class__branch=branch).first() if kind == "ec" and object_id else None,
            "target_student": Student.objects.select_related("inscription__candidature").filter(
                pk=student_id,
                inscription__candidature__branch=branch,
            ).first() if kind == "assign" and student_id else None,
            "selected_semester": Semester.objects.filter(pk=semester_id, academic_class__branch=branch).first() if semester_id else None,
            "selected_ue": UE.objects.filter(pk=ue_id, semester__academic_class__branch=branch).first() if ue_id else None,
        }
    )
    return render(request, "portal/informaticien/workflows/structure_modal.html", context)


def _render_structure_modal_from_post(request, *, branch, message):
    action = (request.POST.get("action") or "").strip()
    section = (request.POST.get("section") or "classes").strip()
    kind = {
        "save_class": "class",
        "save_ue": "ue",
        "save_ec": "ec",
        "assign_student": "assign",
    }.get(action, "class")
    object_id = {
        "save_class": request.POST.get("class_id"),
        "save_ue": request.POST.get("ue_id"),
        "save_ec": request.POST.get("ec_id"),
    }.get(action)
    class_id = request.POST.get("selected_class_id") or request.POST.get("class_id")
    semester_id = request.POST.get("semester_id")
    ue_id = request.POST.get("ue_id")
    student_id = request.POST.get("student_id")

    context = build_academic_structure_context(branch=branch, selected_class_id=class_id, section=section)
    context.update(
        {
            "kind": kind,
            "target_id": object_id,
            "target_class": AcademicClass.objects.filter(pk=object_id, branch=branch).first() if kind == "class" and object_id else None,
            "target_ue": UE.objects.filter(pk=object_id, semester__academic_class__branch=branch).first() if kind == "ue" and object_id else None,
            "target_ec": EC.objects.filter(pk=object_id, ue__semester__academic_class__branch=branch).first() if kind == "ec" and object_id else None,
            "target_student": Student.objects.select_related("inscription__candidature").filter(
                pk=student_id,
                inscription__candidature__branch=branch,
            ).first() if kind == "assign" and student_id else None,
            "selected_semester": Semester.objects.filter(pk=semester_id, academic_class__branch=branch).first() if semester_id else None,
            "selected_ue": UE.objects.filter(pk=ue_id, semester__academic_class__branch=branch).first() if ue_id else None,
            "form_error": message,
            "form_values": request.POST,
        }
    )
    response = render(request, "portal/informaticien/workflows/structure_modal.html", context)
    response["HX-Retarget"] = "#it-modal-root"
    return response


@login_required
def it_structure_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")

    branch = get_user_branch(request.user)
    selected_class_id = (request.POST.get("selected_class_id") or "").strip()
    action = (request.POST.get("action") or "").strip()
    section = (request.POST.get("section") or "classes").strip()
    toast = None
    try:
        if action == "save_class":
            academic_class = save_academic_class(
                branch=branch,
                class_id=(request.POST.get("class_id") or "").strip() or None,
                programme_id=request.POST.get("programme_id"),
                academic_year_id=request.POST.get("academic_year_id"),
                level=request.POST.get("level"),
                threshold=request.POST.get("validation_threshold"),
                actor=request.user,
            )
            selected_class_id = str(academic_class.id)
            section = "classes"
            toast = {"level": "success", "message": "Classe academique enregistree."}
        elif action == "archive_class":
            archive_academic_class(branch=branch, class_id=request.POST.get("class_id"))
            section = "classes"
            toast = {"level": "success", "message": "Classe archivee."}
        elif action == "save_ue":
            ue = save_ue(
                branch=branch,
                ue_id=(request.POST.get("ue_id") or "").strip() or None,
                semester_id=request.POST.get("semester_id"),
                code=request.POST.get("code"),
                title=request.POST.get("title"),
            )
            selected_class_id = str(ue.semester.academic_class_id)
            section = "maquettes"
            toast = {"level": "success", "message": "UE enregistree."}
        elif action == "save_ec":
            ec = save_ec(
                branch=branch,
                ec_id=(request.POST.get("ec_id") or "").strip() or None,
                ue_id=request.POST.get("ue_id"),
                title=request.POST.get("title"),
                coefficient=request.POST.get("coefficient"),
                credit_required=request.POST.get("credit_required"),
            )
            selected_class_id = str(ec.ue.semester.academic_class_id)
            section = "maquettes"
            toast = {"level": "success", "message": "EC enregistre."}
        elif action == "delete_ec":
            delete_ec(branch=branch, ec_id=request.POST.get("ec_id"))
            section = "maquettes"
            toast = {"level": "success", "message": "EC supprime."}
        elif action == "assign_student":
            enrollment = assign_student_to_class(
                branch=branch,
                student_id=request.POST.get("student_id"),
                class_id=request.POST.get("class_id"),
            )
            selected_class_id = str(enrollment.academic_class_id)
            section = "affectations"
            toast = {"level": "success", "message": "Etudiant affecte a la classe."}
        else:
            raise ValidationError("Action de parametrage inconnue.")
    except ValidationError as exc:
        return _render_structure_modal_from_post(
            request,
            branch=branch,
            message=" ".join(exc.messages),
        )

    context = build_academic_structure_context(
        branch=branch,
        selected_class_id=selected_class_id,
        student_query=(request.POST.get("student_q") or "").strip(),
        section=section,
    )
    context["toast"] = toast
    response = render(request, "portal/informaticien/workflows/structure_workspace.html", context)
    response["HX-Trigger"] = '{"it-modal-close": true, "it-structure-updated": true}'
    return response


@login_required
def it_archives_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    academic_year_id = (request.GET.get("academic_year_id") or "").strip()
    class_id = (request.GET.get("class_id") or "").strip()
    candidates = archive_candidates_for_branch(branch=branch)
    preview = None
    if academic_year_id or class_id:
        preview = preview_archive(
            branch=branch,
            academic_year_id=academic_year_id or None,
            class_id=class_id or None,
        )
    return _render_it_section(
        request,
        "archives",
        "portal/informaticien/workflows/archive_workspace.html",
        {
            "branch": branch,
            "archive_batches": archive_batches_for_branch(branch=branch)[:40],
            "academic_years": candidates["academic_years"],
            "classes": candidates["classes"],
            "selected_academic_year_id": academic_year_id,
            "selected_class_id": class_id,
            "preview": preview,
            "toast": None,
        },
    )


@login_required
def it_archives_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")
    branch = get_user_branch(request.user)
    action = (request.POST.get("action") or "").strip()
    toast = {"level": "success", "message": "Archive mise a jour."}
    try:
        if action == "archive_class":
            archive_class(
                branch=branch,
                class_id=request.POST.get("class_id"),
                actor=request.user,
                reason=request.POST.get("reason"),
            )
            toast = {"level": "success", "message": "Classe archivee avec succes."}
        elif action == "archive_year":
            archive_year(
                branch=branch,
                academic_year_id=request.POST.get("academic_year_id"),
                actor=request.user,
                reason=request.POST.get("reason"),
            )
            toast = {"level": "success", "message": "Annee academique archivee avec succes."}
        elif action == "restore":
            restore_archive_batch(
                branch=branch,
                batch_id=request.POST.get("batch_id"),
                actor=request.user,
            )
            toast = {"level": "success", "message": "Archive restauree."}
        else:
            raise ValidationError("Action d'archivage inconnue.")
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages)}

    candidates = archive_candidates_for_branch(branch=branch)
    return render(
        request,
        "portal/informaticien/workflows/archive_workspace.html",
        {
            "branch": branch,
            "archive_batches": archive_batches_for_branch(branch=branch)[:40],
            "academic_years": candidates["academic_years"],
            "classes": candidates["classes"],
            "selected_academic_year_id": "",
            "selected_class_id": "",
            "preview": None,
            "toast": toast,
        },
    )


@login_required
def it_archive_detail(request, batch_id):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    try:
        context = archive_detail(branch=branch, batch_id=batch_id)
    except ValidationError as exc:
        return HttpResponse(" ".join(exc.messages), status=404)
    return render(request, "portal/informaticien/workflows/archive_detail.html", context)


@login_required
def it_supervision_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    return _render_it_section(
        request,
        "supervision",
        "portal/informaticien/workflows/supervision_workspace.html",
        build_supervision_context(branch=get_user_branch(request.user), page=request.GET.get("page")),
    )


@login_required
def it_catalog_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    return _render_it_section(request, "catalog", "portal/informaticien/workflows/catalog_workspace.html", build_catalog_context())


@login_required
def it_catalog_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")
    toast = None
    try:
        create_catalog_item(
            kind=request.POST.get("kind"),
            name=request.POST.get("name"),
            code=request.POST.get("code"),
            description=request.POST.get("description"),
        )
        toast = {"level": "success", "message": "Element ajoute."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages)}
    return render(request, "portal/informaticien/workflows/catalog_workspace.html", build_catalog_context(toast=toast))


def _build_accounts_flow_context(request, *, toast=None):
    branch = get_user_branch(request.user)
    query = (request.GET.get("q") or "").strip()
    users_qs = get_scoped_staff_queryset(branch=branch).filter(profile__branch=branch) if branch else get_scoped_staff_queryset(branch=branch).none()
    if query:
        users_qs = users_qs.filter(
            Q(username__icontains=query)
            | Q(email__icontains=query)
            | Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
        )
    rows = []
    for user in users_qs.select_related("profile")[:40]:
        state = get_account_support_state(user)
        if state.is_suspended:
            account_state = "suspendu"
        elif state.is_blocked:
            account_state = "bloque"
        elif user.is_active:
            account_state = "actif"
        else:
            account_state = "inactif"
        rows.append({"user": user, "support_state": state, "account_state": account_state})

    return {
        "branch": branch,
        "query": query,
        "rows": rows,
        "toast": toast,
    }


@login_required
def it_accounts_flow_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    return _render_it_section(
        request,
        "accounts",
        "portal/informaticien/workflows/accounts_workspace.html",
        _build_accounts_flow_context(request),
    )


@login_required
def it_accounts_flow_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")

    branch = get_user_branch(request.user)
    target_user = get_object_or_404(
        get_scoped_staff_queryset(branch=branch).filter(profile__branch=branch),
        pk=request.POST.get("target_user_id"),
    )
    if not _same_branch_or_forbidden(request=request, target_user=target_user):
        return HttpResponseForbidden("Action hors annexe refusee.")
    action = (request.POST.get("action") or "").strip()
    toast = None

    if action == "toggle_active":
        target_user.is_active = not target_user.is_active
        target_user.save(update_fields=["is_active"])
        log_support_action(
            actor=request.user,
            branch=branch,
            action_type="account_activated" if target_user.is_active else "account_deactivated",
            target_user=target_user,
            target_label=target_user.get_full_name() or target_user.username,
            details="Action depuis workflow comptes informaticien.",
        )
        toast = {"level": "success", "message": "Etat du compte mis a jour."}
    elif action == "reset_password":
        temp_password = create_temp_password()
        target_user.set_password(temp_password)
        target_user.save(update_fields=["password"])
        log_support_action(
            actor=request.user,
            branch=branch,
            action_type="password_reset",
            target_user=target_user,
            target_label=target_user.get_full_name() or target_user.username,
            details="Reset mot de passe depuis workflow comptes informaticien.",
        )
        toast = {"level": "success", "message": f"Mot de passe temporaire: {temp_password}"}
    else:
        toast = {"level": "error", "message": "Action compte inconnue."}

    return render(
        request,
        "portal/informaticien/workflows/accounts_workspace.html",
        _build_accounts_flow_context(request, toast=toast),
    )


@login_required
def it_user_modal(request, user_id):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    branch = get_user_branch(request.user)
    target_user = get_object_or_404(
        get_scoped_staff_queryset(branch=branch).filter(profile__branch=branch).select_related("profile"),
        pk=user_id,
    )
    if not _same_branch_or_forbidden(request=request, target_user=target_user):
        return HttpResponseForbidden("Action hors annexe refusee.")
    return render(
        request,
        "portal/informaticien/workflows/user_modal.html",
        {
            "target_user": target_user,
            "profile": target_user.profile,
        },
    )


@login_required
def it_user_modal_save(request, user_id):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")

    branch = get_user_branch(request.user)
    target_user = get_object_or_404(
        get_scoped_staff_queryset(branch=branch).filter(profile__branch=branch).select_related("profile"),
        pk=user_id,
    )
    if not _same_branch_or_forbidden(request=request, target_user=target_user):
        return HttpResponseForbidden("Action hors annexe refusee.")
    profile = target_user.profile

    target_user.first_name = (request.POST.get("first_name") or "").strip()
    target_user.last_name = (request.POST.get("last_name") or "").strip()
    target_user.email = (request.POST.get("email") or "").strip().lower()
    target_user.is_active = bool(request.POST.get("is_active"))
    target_user.save(update_fields=["first_name", "last_name", "email", "is_active"])

    log_support_action(
        actor=request.user,
        branch=branch,
        action_type="email_updated",
        target_user=target_user,
        target_label=target_user.get_full_name() or target_user.username,
        details="Modification utilisateur depuis modal informaticien.",
    )

    return render(
        request,
        "portal/informaticien/workflows/user_modal.html",
        {
            "target_user": target_user,
            "profile": profile,
            "toast": {"level": "success", "message": "Utilisateur mis a jour."},
        },
    )


@login_required
def it_branch_settings_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    settings = get_branch_settings(branch=branch)
    return _render_it_section(
        request,
        "settings",
        "portal/informaticien/workflows/branch_settings_workspace.html",
        {
            "branch": branch,
            "settings": settings,
            "profile": getattr(request.user, "profile", None),
        },
    )


@login_required
def it_branch_settings_save(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")
    branch = get_user_branch(request.user)
    toast = None
    try:
        settings = update_branch_settings(
            actor=request.user,
            branch=branch,
            validation_threshold=request.POST.get("validation_threshold"),
            active_academic_year=request.POST.get("active_academic_year"),
            local_config=request.POST.get("local_config"),
        )
        toast = {"level": "success", "message": "Parametres mis a jour."}
    except ValidationError as exc:
        settings = get_branch_settings(branch=branch)
        toast = {"level": "error", "message": " ".join(exc.messages)}
    return render(
        request,
        "portal/informaticien/workflows/branch_settings_workspace.html",
        {
            "branch": branch,
            "settings": settings,
            "toast": toast,
            "profile": getattr(request.user, "profile", None),
        },
    )


@login_required
def it_notes_retake_modal(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    return render(
        request,
        "portal/informaticien/workflows/retake_modal.html",
        _build_retake_modal_context(request),
    )


@login_required
def it_my_account_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    profile = getattr(request.user, "profile", None)
    return _render_it_section(
        request,
        "settings",
        "portal/informaticien/workflows/my_account_workspace.html",
        {
            "profile": profile,
        },
    )


@login_required
def it_my_account_save(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    if request.method != "POST":
        return HttpResponseForbidden("Methode non autorisee.")
    request.user.first_name = (request.POST.get("first_name") or "").strip()
    request.user.last_name = (request.POST.get("last_name") or "").strip()
    request.user.email = (request.POST.get("email") or "").strip().lower()
    request.user.save(update_fields=["first_name", "last_name", "email"])
    if request.POST.get("return_settings"):
        branch = get_user_branch(request.user)
        return render(
            request,
            "portal/informaticien/workflows/branch_settings_workspace.html",
            {
                "branch": branch,
                "settings": get_branch_settings(branch=branch),
                "profile": getattr(request.user, "profile", None),
                "toast": {"level": "success", "message": "Profil mis a jour."},
            },
        )
    return render(
        request,
        "portal/informaticien/workflows/my_account_workspace.html",
        {
            "profile": getattr(request.user, "profile", None),
            "toast": {"level": "success", "message": "Profil mis a jour."},
        },
    )


@login_required
def it_student_card_pdf(request, student_id):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    branch = get_user_branch(request.user)
    student_qs = Student.objects.select_related(
        "inscription__candidature__branch",
        "inscription__candidature__programme",
    )
    if branch:
        student_qs = student_qs.filter(inscription__candidature__branch=branch)
    student = get_object_or_404(student_qs, pk=student_id)

    pdf_bytes = _render_student_card_pdf(request, student, branch)
    buffer = BytesIO(pdf_bytes)
    log_support_action(
        actor=request.user,
        branch=branch,
        action_type=SupportAuditLog.ACTION_STUDENT_CARD_GENERATED,
        target_user=student.user,
        target_label=f"Carte {student.matricule or student.id}",
        details="Carte individuelle generee depuis le dashboard informaticien.",
    )
    return FileResponse(
        buffer,
        as_attachment=request.GET.get("preview") != "1",
        filename=f"carte-{student.matricule or student.id}.pdf",
        content_type="application/pdf",
    )


@login_required
def it_class_cards_pdf(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")
    branch = get_user_branch(request.user)
    selection = _resolve_workflow_selection(request)
    academic_class = selection["selected_class"]
    if academic_class is None or academic_class.branch_id != branch.id:
        return HttpResponseForbidden("Classe hors annexe refusee.")

    students = get_it_students_for_class(academic_class=academic_class)
    pdf_bytes = _render_class_cards_pdf(request, list(students), academic_class, branch)
    buffer = BytesIO(pdf_bytes)
    log_support_action(
        actor=request.user,
        branch=branch,
        action_type=SupportAuditLog.ACTION_STUDENT_CARD_GENERATED,
        target_label=f"Cartes classe {academic_class.display_name}",
        details=f"{students.count()} carte(s) generee(s).",
    )
    return FileResponse(
        buffer,
        as_attachment=request.GET.get("preview") != "1",
        filename=f"cartes-classe-{academic_class.id}.pdf",
        content_type="application/pdf",
    )


def _get_or_create_carte(student, branch):
    from datetime import date
    from students.models import CarteEtudiant

    enrollment = (
        student.user.academic_enrollments.select_related("academic_class__academic_year")
        .filter(is_active=True)
        .order_by("-created_at")
        .first()
    )
    academic_year = None
    if enrollment and enrollment.academic_class:
        academic_year = getattr(enrollment.academic_class, "academic_year", None)
    if academic_year is None and enrollment:
        academic_year = getattr(enrollment, "academic_year", None)

    today = date.today()
    annee = str(academic_year)[:9] if academic_year else f"{today.year}-{today.year + 1}"
    code_annexe = (getattr(branch, "code", None) or (branch.name[:20] if branch else "ESFE"))
    date_expiration = date(today.year + 1, 9, 30)

    carte, _ = CarteEtudiant.objects.get_or_create(
        etudiant=student,
        annee=annee,
        defaults={
            "code_annexe": code_annexe,
            "date_expiration": date_expiration,
            "statut": "active",
        },
    )
    return carte


def _render_student_card_pdf(request, student, branch=None):
    from weasyprint import HTML
    from django.template.loader import render_to_string
    from students.services.card_security import signer_carte, generer_code_lisible, generer_qr_png, generer_qr_svg
    from students.views_carte import _logo_data_uri, _get_classe

    if branch is None:
        branch = get_user_branch(request.user)

    carte = _get_or_create_carte(student, branch)
    token = signer_carte(carte.etudiant.matricule, carte.annee, carte.code_annexe)
    verify_url = request.build_absolute_uri(f"/carte/v/{token}/")

    ctx = {
        "carte": carte,
        "etudiant": student,
        "classe": _get_classe(carte),
        "qr_png": generer_qr_png(verify_url),
        "qr_svg": generer_qr_svg(verify_url),
        "code_verification": generer_code_lisible(token),
        "logo_data_uri": _logo_data_uri(),
    }
    html_str = render_to_string("students/carte_etudiant.html", ctx, request=request)
    return HTML(string=html_str, base_url=request.build_absolute_uri("/")).write_pdf()


def _render_class_cards_pdf(request, students_list, academic_class, branch=None):
    from weasyprint import HTML
    from django.template.loader import render_to_string
    from students.services.card_security import signer_carte, generer_code_lisible, generer_qr_png, generer_qr_svg
    from students.views_carte import _logo_data_uri

    if branch is None:
        branch = get_user_branch(request.user)

    logo = _logo_data_uri()
    base_url = request.build_absolute_uri("/")
    docs = []

    for student in students_list:
        carte = _get_or_create_carte(student, branch)
        token = signer_carte(carte.etudiant.matricule, carte.annee, carte.code_annexe)
        verify_url = request.build_absolute_uri(f"/carte/v/{token}/")
        ctx = {
            "carte": carte,
            "etudiant": student,
            "classe": str(academic_class) if academic_class else "",
            "qr_png": generer_qr_png(verify_url),
            "qr_svg": generer_qr_svg(verify_url),
            "code_verification": generer_code_lisible(token),
            "logo_data_uri": logo,
        }
        html_str = render_to_string("students/carte_etudiant.html", ctx, request=request)
        docs.append(HTML(string=html_str, base_url=base_url).render())

    if not docs:
        return b""

    all_pages = []
    for doc in docs:
        all_pages.extend(doc.pages)
    docs[0].pages = all_pages
    return docs[0].write_pdf()


@login_required
def it_notifications_workspace(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    user = request.user

    # Return just the unread count badge (used by the bell icon)
    if request.GET.get("count_only"):
        unread = CommunicationNotification.objects.filter(
            recipient=user, read_at__isnull=True,
        ).count()
        badge = (
            f'<span id="it-notif-badge" class="absolute -right-1 -top-1 grid min-h-[18px] min-w-[18px] '
            f'place-items-center rounded-full bg-red-500 px-1 text-[10px] font-black leading-none text-white">'
            f'{unread}</span>'
        )
        return HttpResponse(badge)

    # Dropdown partial for the bell icon (7 latest notifications)
    if request.GET.get("partial") == "dropdown":
        from communication.selectors import get_user_notifications, get_user_unread_count
        notifs = get_user_notifications(user, limit=7)
        unread = get_user_unread_count(user)
        return render(request, "portal/informaticien/partials/notifications_dropdown_it.html", {
            "notifications": notifs,
            "unread_count": unread,
        })

    q = request.GET.get("q", "")
    channel = request.GET.get("channel", "all")
    status_filter = request.GET.get("status", "")
    priority = request.GET.get("priority", "")
    source = request.GET.get("source", "")
    page = request.GET.get("page", 1)
    notification_id = request.GET.get("notification_id")

    qs = CommunicationNotification.objects.filter(
        Q(recipient=user) | Q(legacy_source__icontains="system"),
    )

    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(body__icontains=q))
    if channel and channel != "all":
        qs = qs.filter(channel=channel)
    if status_filter == "unread":
        qs = qs.filter(read_at__isnull=True)
    elif status_filter == "read":
        qs = qs.filter(read_at__isnull=False)
    elif status_filter == "failed":
        qs = qs.filter(status=CommunicationNotification.STATUS_FAILED)
    if priority:
        qs = qs.filter(priority=priority)
    if source:
        qs = qs.filter(Q(event__source_app__icontains=source) | Q(legacy_source__icontains=source))

    qs = qs.select_related("event", "recipient").prefetch_related("deliveries").order_by("-created_at")

    paginator = Paginator(qs, 12)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    notifications = list(page_obj)
    selected_notification = None
    if notification_id:
        try:
            selected_notification = CommunicationNotification.objects.filter(
                id=notification_id,
            ).select_related("event").prefetch_related("deliveries").first()
        except ValueError:
            pass

    all_user_notifs = CommunicationNotification.objects.filter(recipient=user)
    stats = {
        "total": all_user_notifs.count(),
        "unread": all_user_notifs.filter(read_at__isnull=True).count(),
        "critical": all_user_notifs.filter(priority=CommunicationNotification.PRIORITY_CRITICAL).count(),
        "email": all_user_notifs.filter(channel=CommunicationNotification.CHANNEL_EMAIL_TRANSACTIONAL).count(),
        "failed": all_user_notifs.filter(status=CommunicationNotification.STATUS_FAILED).count(),
        "by_source": list(
            all_user_notifs.values("event__source_app").annotate(total=Count("id")).order_by("-total")[:5]
        ),
    }
    unread_count = stats["unread"]

    filters = {
        "q": q,
        "channel": channel,
        "status": status_filter,
        "priority": priority,
        "source": source,
    }

    return _render_it_section(request, "notifications", "portal/informaticien/workflows/notifications_workspace.html", {
        "notifications": notifications,
        "page_obj": page_obj,
        "stats": stats,
        "unread_count": unread_count,
        "channels": CommunicationNotification.CHANNEL_CHOICES,
        "priorities": CommunicationNotification.PRIORITY_CHOICES,
        "sources": list(
            CommunicationNotification.objects.filter(recipient=user)
            .values_list("event__source_app", flat=True)
            .distinct()[:10]
        ),
        "filters": filters,
        "selected_notification": selected_notification,
        "toast": None,
    })


@login_required
def it_notifications_action(request):
    if not _require_it_support(request):
        return HttpResponseForbidden("Acces refuse.")

    if request.method != "POST":
        return HttpResponse("Methode non autorisee.", status=405)

    user = request.user
    action = request.POST.get("action", "")
    notification_id = request.POST.get("notification_id")

    toast = None

    if action == "mark_read" and notification_id:
        updated = CommunicationNotification.objects.filter(
            id=notification_id, recipient=user, read_at__isnull=True,
        ).update(read_at=timezone.now(), status=CommunicationNotification.STATUS_READ)
        if updated:
            toast = {"level": "success", "message": "Notification marquee comme lue."}
        else:
            toast = {"level": "error", "message": "Notification deja lue ou introuvable."}

    elif action == "mark_unread" and notification_id:
        CommunicationNotification.objects.filter(
            id=notification_id, recipient=user,
        ).update(read_at=None, status=CommunicationNotification.STATUS_DELIVERED)
        toast = {"level": "success", "message": "Notification marquee comme non lue."}

    elif action == "mark_all_read":
        count = CommunicationNotification.objects.filter(
            recipient=user, read_at__isnull=True,
        ).update(read_at=timezone.now(), status=CommunicationNotification.STATUS_READ)
        toast = {"level": "success", "message": f"{count} notification(s) marquee(s) comme lue(s)."}

    else:
        toast = {"level": "error", "message": f"Action inconnue: {action}"}

    # Re-render workspace with updated state
    from django.http import QueryDict
    get_params = QueryDict(mutable=True)
    for key in ("channel", "status", "priority", "source", "q", "page"):
        val = request.POST.get(key, "")
        if val:
            get_params[key] = val

    from django.test.client import RequestFactory
    factory = RequestFactory()
    fake_get = factory.get(f"/?{get_params.urlencode()}")
    fake_get.user = user
    fake_get.GET = get_params
    fake_get.META = request.META

    response = it_notifications_workspace(fake_get)
    if hasattr(response, "context_data"):
        response.context_data["toast"] = toast
    elif hasattr(response, "context"):
        response.context["toast"] = toast
    return response
