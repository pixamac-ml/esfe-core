from datetime import date
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

THIN_BORDER = Border(
    left=Side(style="thin", color="dfe7f3"),
    right=Side(style="thin", color="dfe7f3"),
    top=Side(style="thin", color="dfe7f3"),
    bottom=Side(style="thin", color="dfe7f3"),
)
HEADER_FILL = PatternFill(start_color="1e4f6f", end_color="1e4f6f", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="ffffff", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1e4f6f")
SUB_FONT = Font(name="Calibri", bold=True, size=12, color="1e4f6f")
MONEY_FONT = Font(name="Calibri", bold=True, size=11, color="16a34a")
LOSS_FONT = Font(name="Calibri", bold=True, size=11, color="ef4444")
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)


def _style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            length = max(length, min(len(val) + 4, max_width))
        ws.column_dimensions[col_letter].width = length


def _write_cell(ws, row, col, value, font=None, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or BODY_FONT
    cell.border = THIN_BORDER
    if align:
        cell.alignment = Alignment(horizontal=align)
    if fmt:
        cell.number_format = fmt
    return cell


def build_branch_xlsx_report(
    *,
    branch,
    report_period,
    report_rows,
    report_movements,
    period_summary,
    annual_revenue_rows,
    payroll_stats,
    honorarium_stats,
    expense_stats,
    cash_stats,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"
    ws.sheet_properties.tabColor = "1e4f6f"
    row = 1

    # Title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"Rapport financier — {branch.name}")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left")
    row = 3

    # Period
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value=f"Periode: {report_period['label']} ({report_period['start']} — {report_period['end']})").font = SUB_FONT
    row += 2

    # Summary
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Situation de la periode").font = SUB_FONT
    row += 1
    summary_headers = ["Indicateur", "Montant (FCFA)"]
    for col_i, h in enumerate(summary_headers, 1):
        _write_cell(ws, row, col_i, h, font=HEADER_FONT)
    _style_header_row(ws, row, len(summary_headers))
    row += 1
    summary_data = [
        ("Recettes etudiants", period_summary["student_revenue"], MONEY_FONT),
        ("Recettes boutique", period_summary["shop_revenue"], MONEY_FONT),
        ("Dons / Donations", period_summary["donation_revenue"], MONEY_FONT),
        ("Total recettes", period_summary["total_revenue"], MONEY_FONT),
        ("Depenses payees", period_summary["expenses_paid"], LOSS_FONT),
        ("Salaires payes", period_summary["salary_paid"], LOSS_FONT),
        ("Honoraires payes", period_summary["honorarium_paid"], LOSS_FONT),
        ("Charges payees", period_summary["charges_paid"], LOSS_FONT),
        ("Resultat net", period_summary["net_result"], MONEY_FONT if period_summary["net_result"] >= 0 else LOSS_FONT),
        ("Caisse estimee", period_summary["estimated_cash"], BOLD_FONT),
    ]
    for label, val, fnt in summary_data:
        _write_cell(ws, row, 1, label, font=BOLD_FONT)
        _write_cell(ws, row, 2, val, font=fnt, fmt="#,##0")
        row += 1
    row += 1

    # Report detail
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Detail des mouvements").font = SUB_FONT
    row += 1
    detail_headers = ["Indicateur", "Montant (FCFA)"]
    for col_i, h in enumerate(detail_headers, 1):
        _write_cell(ws, row, col_i, h, font=HEADER_FONT)
    _style_header_row(ws, row, len(detail_headers))
    row += 1
    for item in report_rows:
        _write_cell(ws, row, 1, item["label"], font=BOLD_FONT)
        _write_cell(ws, row, 2, item["amount"], font=BODY_FONT, fmt="#,##0")
        row += 1
    row += 1

    # Recent movements
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Derniers mouvements de caisse").font = SUB_FONT
    row += 1
    mov_headers = ["Date", "Type", "Source", "Montant", "Libelle"]
    for col_i, h in enumerate(mov_headers, 1):
        _write_cell(ws, row, col_i, h, font=HEADER_FONT)
    _style_header_row(ws, row, len(mov_headers))
    row += 1
    for m in report_movements:
        _write_cell(ws, row, 1, m.movement_date.isoformat() if hasattr(m, "movement_date") else "")
        _write_cell(ws, row, 2, "Entree" if m.movement_type == "in" else "Sortie")
        _write_cell(ws, row, 3, m.get_source_display() if hasattr(m, "get_source_display") else m.source)
        _write_cell(ws, row, 4, m.amount, font=MONEY_FONT if m.movement_type == "in" else LOSS_FONT, fmt="#,##0")
        _write_cell(ws, row, 5, m.label)
        row += 1
    row += 1

    # Stats
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Statistiques du mois").font = SUB_FONT
    row += 1
    stats = [
        ("Employes staff", payroll_stats.get("employees", 0)),
        ("Fiches de paie preparees", payroll_stats.get("prepared", 0)),
        ("Salaires dus", payroll_stats.get("due_total", 0)),
        ("Salaires restants", payroll_stats.get("remaining_total", 0)),
        ("Enseignants", honorarium_stats.get("teachers", 0)),
        ("Honoraires dus", honorarium_stats.get("due_total", 0)),
        ("Honoraires restants", honorarium_stats.get("remaining_total", 0)),
        ("Depenses du mois", expense_stats.get("month_amount", 0)),
        ("Caisse disponible", cash_stats.get("available_balance", 0)),
    ]
    for label, val in stats:
        _write_cell(ws, row, 1, label, font=BOLD_FONT)
        _write_cell(ws, row, 2, val, font=BODY_FONT, fmt="#,##0")
        row += 1

    _auto_width(ws)
    for r in range(1, row):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER

    return wb


def export_branch_report_xlsx(*, branch, report_period, branch_staff_profiles, branch_teacher_profiles):
    from datetime import date, timedelta
    from django.db.models import Q, Sum
    from accounts.models import BranchCashMovement, BranchExpense, PayrollEntry, TeacherHonorariumEntry
    from accounts.dashboards.manager_dashboard import _resolve_report_period
    from accounts.services.manager_intelligence import get_branch_cash_balance

    today = date.today()
    start_of_month = today.replace(day=1)

    report_movements = BranchCashMovement.objects.filter(
        branch=branch,
        movement_date__gte=report_period["start"],
        movement_date__lte=report_period["end"],
    ).order_by("-movement_date")[:200]

    report_total_entries = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_IN
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_total_exits = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_OUT
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_student_payments = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_IN, source=BranchCashMovement.SOURCE_STUDENT_PAYMENT
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_shop_sales = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_IN, source=BranchCashMovement.SOURCE_SHOP
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_donations = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_IN, source=BranchCashMovement.SOURCE_DONATION
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_expenses = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_OUT, source=BranchCashMovement.SOURCE_EXPENSE
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_salaries = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_OUT, source=BranchCashMovement.SOURCE_PAYROLL
    ).aggregate(total=Sum("amount"))["total"] or 0
    report_honorarium = report_movements.filter(
        movement_type=BranchCashMovement.TYPE_OUT, source=BranchCashMovement.SOURCE_HONORARIUM
    ).aggregate(total=Sum("amount"))["total"] or 0

    report_rows = [
        {"label": "Total entrees", "amount": report_total_entries},
        {"label": "Total sorties", "amount": report_total_exits},
        {"label": "Paiements scolaires", "amount": report_student_payments},
        {"label": "Ventes boutique", "amount": report_shop_sales},
        {"label": "Dons / Donations", "amount": report_donations},
        {"label": "Depenses", "amount": report_expenses},
        {"label": "Salaires", "amount": report_salaries},
        {"label": "Honoraires enseignants", "amount": report_honorarium},
        {"label": "Solde net", "amount": report_total_entries - report_total_exits},
    ]

    payroll_month = today.replace(day=1)
    branch_staff_user_ids = list(branch_staff_profiles.values_list("user_id", flat=True)) if branch_staff_profiles else []
    branch_teacher_user_ids = list(branch_teacher_profiles.values_list("user_id", flat=True)) if branch_teacher_profiles else []

    payroll_entries_qs = PayrollEntry.objects.filter(
        branch=branch, employee_id__in=branch_staff_user_ids, period_month=payroll_month
    )
    honorarium_entries_qs = TeacherHonorariumEntry.objects.filter(
        branch=branch, teacher_id__in=branch_teacher_user_ids, period_month=payroll_month
    )

    cash_in_month = BranchCashMovement.objects.filter(
        branch=branch, movement_date__gte=start_of_month, movement_type=BranchCashMovement.TYPE_IN
    ).aggregate(total=Sum("amount"))["total"] or 0
    cash_out_month = BranchCashMovement.objects.filter(
        branch=branch, movement_date__gte=start_of_month, movement_type=BranchCashMovement.TYPE_OUT
    ).aggregate(total=Sum("amount"))["total"] or 0
    salary_paid_month = payroll_entries_qs.aggregate(total=Sum("paid_amount"))["total"] or 0
    honorarium_paid_month = honorarium_entries_qs.aggregate(total=Sum("paid_amount"))["total"] or 0

    expenses_paid = BranchExpense.objects.filter(branch=branch, expense_date__gte=start_of_month, status=BranchExpense.STATUS_PAID).aggregate(total=Sum("amount"))["total"] or 0
    total_revenue = report_student_payments + report_shop_sales + report_donations
    period_summary = {
        "student_revenue": report_student_payments,
        "shop_revenue": report_shop_sales,
        "donation_revenue": report_donations,
        "total_revenue": total_revenue,
        "expenses_paid": expenses_paid,
        "salary_paid": salary_paid_month,
        "honorarium_paid": honorarium_paid_month,
        "charges_paid": expenses_paid + salary_paid_month + honorarium_paid_month,
        "net_result": total_revenue - expenses_paid - salary_paid_month - honorarium_paid_month,
        "estimated_cash": cash_in_month - cash_out_month,
    }

    payroll_stats = {
        "employees": len(branch_staff_user_ids),
        "prepared": payroll_entries_qs.count(),
        "due_total": sum(e.net_salary for e in payroll_entries_qs),
        "remaining_total": sum(max(e.net_salary - e.paid_amount, 0) for e in payroll_entries_qs),
    }
    honorarium_stats = {
        "teachers": len(branch_teacher_user_ids),
        "prepared": honorarium_entries_qs.count(),
        "due_total": sum(e.net_amount for e in honorarium_entries_qs),
        "remaining_total": sum(max(e.net_amount - e.paid_amount, 0) for e in honorarium_entries_qs),
    }
    expense_stats = {
        "month_amount": BranchExpense.objects.filter(branch=branch, expense_date__gte=start_of_month).exclude(status=BranchExpense.STATUS_REJECTED).aggregate(total=Sum("amount"))["total"] or 0,
    }
    cash_stats = {
        "available_balance": get_branch_cash_balance(branch),
    }

    wb = build_branch_xlsx_report(
        branch=branch,
        report_period=report_period,
        report_rows=report_rows,
        report_movements=report_movements,
        period_summary=period_summary,
        annual_revenue_rows=[],
        payroll_stats=payroll_stats,
        honorarium_stats=honorarium_stats,
        expense_stats=expense_stats,
        cash_stats=cash_stats,
    )
    return wb


def xlsx_response(wb, filename):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
