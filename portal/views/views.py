from collections import defaultdict
import csv
from datetime import datetime, timedelta
from decimal import Decimal
from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.urls.exceptions import NoReverseMatch
from django.utils import timezone
from django.utils.safestring import mark_safe

from academics.models import (
    AcademicBulletin,
    AcademicClass,
    AcademicDiplomaAward,
    AcademicEnrollment,
    AcademicScheduleEvent,
    EC,
    ECChapter,
    ECContent,
    ECGrade,
    LessonLog,
    Semester,
    WeeklyScheduleSlot,
)
from academics.permissions import BULLETIN_MANAGEMENT_POSITIONS, can_manage_bulletins, can_manage_diplomas
from academic_cycle.services.academic_excel_reports import build_academic_report_xlsx, xlsx_response
from academics.services.documents import (
    generate_annual_bulletins_for_class,
    generate_semester_bulletins_for_class,
    prepare_diploma_awards_for_class,
)
from admissions.models import Candidature
from academic_cycle.services.audit_service import log_action
from accounts.models import BranchExpense, Profile, SensitiveActionRequest
from accounts.services.sensitive_actions import SensitiveActionError, confirm_sensitive_action, request_sensitive_action
from academics.services.lesson_log_service import create_lesson_log, update_lesson_log
from academics.services.schedule_service import (
    create_weekly_schedule_slot,
    deactivate_weekly_schedule_slot,
    get_class_week_schedule,
    get_teacher_week_schedule,
    get_director_schedule_overview,
    list_weekly_slots_for_class,
    serialize_weekly_slot_for_ui,
    update_weekly_schedule_slot,
)
from academics.services.semester import compute_semester_result
from academics.services.grading import resolve_threshold
from academics.services.workflow import can_publish_semester
from accounts.access import can_access, get_user_position, get_user_scope
from accounts.dashboards.helpers import get_user_branch, paginate_queryset
from branches.models import Branch
from inscriptions.models import Inscription
from payments.models import Payment
from notification_center.selectors import get_user_unread_count
from portal.permissions import get_post_login_portal_url
from portal.services import (
    build_it_dashboard_context,
    build_teacher_class_detail_context,
    build_teacher_dashboard_context,
    build_teacher_overview_context,
    build_teacher_classes_context,
    build_teacher_logs_context,
    build_teacher_notifications_context,
    build_teacher_salary_context,
    build_teacher_schedule_context,
    build_teacher_supports_context,
    build_teacher_lesson_log_context,
)
from portal.services.teacher_dashboard_service import (
    TEACHER_EXCLUDED_NOTIFICATION_SOURCES,
    _validate_content_file_extension,
    build_teacher_support_workspace_context,
    build_teacher_settings_context,
    delete_teacher_content,
    get_teacher_content_for_edit,
    update_teacher_content,
    update_teacher_dashboard_preference,
)
from students.models import TeacherAttendance
from portal.services.director import (
    build_director_classroom_ops_context,
    build_director_document_context,
    build_director_planning_assignment_context,
    build_director_tasks_center,
    build_director_teacher_assignment_context,
    build_director_transfer_context,
    create_transfer_request,
    create_teacher_with_account,
    generate_teacher_contract_pdf,
    review_teacher_document,
    review_transfer_request,
    upload_teacher_document,
)
from portal.services.supervisor_service import build_class_detail_context
from portal.services.academic_structure_service import (
    delete_ec,
    save_ec,
    save_ue,
)
from portal.services.it_support_service import (
    add_support_ticket_comment,
    assign_support_ticket,
    build_diagnostic_payload,
    can_manage_user_in_branch,
    create_support_ticket,
    create_temp_password,
    get_support_ticket_metrics,
    get_support_ticket_queryset,
    reactivate_account,
    get_scoped_staff_queryset,
    log_support_action,
    suspend_account,
    unblock_account,
    update_account_email,
    update_support_ticket_status,
)
from portal.models import AdministrativeDocument, DirectorTeacherAssignment, SupportAuditLog, SupportTicket
from portal.views.it_grades_import import build_it_grade_selection_context
from portal.dg.services import (
    build_dg_dashboard_context,
    build_dg_drawer_context,
    build_dg_section_context,
)
from portal.dg.forms import DgRecruitmentForm
from portal.dg.rh_service import create_staff_from_recruitment
from portal.dg.actions_service import (
    create_finance_followup,
    escalate_student_case,
    resolve_attendance_alert,
    resolve_student_case,
)
from portal.dg.executive_actions import (
    arbitrate_decision,
    deliver_diploma,
    nominate_branch_manager,
    publish_class_diplomas,
    transition_branch_cycle,
    validate_closure,
)
from students.models import AttendanceAlert, Student, StudentAttendance, StudentCase, StudentYearDecision
from secretary.permissions import is_secretary


def _build_portal_context(request, *, page_title, module_cards):
    scope = get_user_scope(request.user)
    user_display_name = request.user.get_full_name() or request.user.username
    try:
        secretary_url = reverse("accounts_portal:portal_secretary")
    except NoReverseMatch:
        secretary_url = ""

    return {
        "page_title": page_title,
        "user_display_name": user_display_name,
        "detected_role": scope.get("role") or "public",
        "scope": scope,
        "module_cards": module_cards,
        "secretary_available": is_secretary(request.user),
        "secretary_url": secretary_url,
        "welcome_message": f"Bienvenue {user_display_name}, vous etes connecte en tant que {scope.get('role') or 'public'}",
    }


def _resolve_academic_branch(request):
    return get_user_branch(request.user)


def _build_academic_dashboard_context(request, *, page_title, page_kicker, sidebar_links, highlight):
    branch = _resolve_academic_branch(request)
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=7)
    overview = get_director_schedule_overview(branch, week_start) if branch else {"stats": {}, "quality": {"score": 0, "warnings": []}, "alerts": [], "timetable": []}

    classes_qs = AcademicClass.objects.select_related("programme", "academic_year", "branch").filter(is_active=True)
    if branch:
        classes_qs = classes_qs.filter(branch=branch)
    classes_qs = classes_qs.annotate(student_count=Count("enrollments")).order_by("level", "programme__title")

    classes = list(classes_qs[:8])
    current_week_events_qs = AcademicScheduleEvent.objects.select_related("academic_class", "teacher", "ec", "branch").filter(
        start_datetime__date__gte=week_start,
        start_datetime__date__lt=week_end,
    )
    if branch:
        current_week_events_qs = current_week_events_qs.filter(branch=branch)
    current_week_events = list(current_week_events_qs.order_by("start_datetime", "id")[:10])

    total_students = AcademicEnrollment.objects.filter(branch=branch).count() if branch else AcademicEnrollment.objects.count()
    total_classes = classes_qs.count()
    total_teachers = (
        AcademicScheduleEvent.objects.filter(branch=branch).values("teacher_id").distinct().count()
        if branch
        else AcademicScheduleEvent.objects.values("teacher_id").distinct().count()
    )

    schedule_stats = overview.get("stats", {})
    quality = overview.get("quality", {})
    alerts = overview.get("alerts", [])
    timetable = overview.get("timetable", [])
    class_load_items = sorted(
        (schedule_stats.get("class_load") or {}).items(),
        key=lambda item: (-item[1]["hours"], item[0]),
    )[:5]
    teacher_load_items = sorted(
        (schedule_stats.get("teacher_load") or {}).items(),
        key=lambda item: (-item[1]["hours"], item[0]),
    )[:5]

    return {
        **_build_portal_context(
            request,
            page_title=page_title,
            module_cards=highlight,
        ),
        "dashboard_kind": page_kicker,
        "branch": branch,
        "week_start": week_start,
        "week_end": week_end,
        "current_week_events": current_week_events,
        "classes": classes,
        "total_students": total_students,
        "total_classes": total_classes,
        "total_teachers": total_teachers,
        "schedule_stats": schedule_stats,
        "quality": quality,
        "alerts": alerts[:8],
        "timetable": timetable,
        "class_load_items": class_load_items,
        "teacher_load_items": teacher_load_items,
        "sidebar_links": sidebar_links,
    }


def _build_it_dashboard_context(request):
    return build_it_dashboard_context(
        request,
        branch=_resolve_academic_branch(request),
        base_context_builder=_build_portal_context,
    )


def _deny_portal_access(request):
    return HttpResponseForbidden("Acces portail refuse.")


def _redirect_supervisor_dashboard(anchor="overview"):
    base_url = reverse("accounts_portal:portal_dashboard")
    return redirect(f"{base_url}#{anchor}")


def _redirect_it_dashboard(request, anchor="diagnostics"):
    params = []
    query = (request.POST.get("q") or request.GET.get("q") or "").strip()
    kind = (request.POST.get("kind") or request.GET.get("kind") or "").strip()
    object_id = (request.POST.get("id") or request.GET.get("id") or "").strip()
    if query:
        params.append(f"q={query}")
    if kind:
        params.append(f"kind={kind}")
    if object_id:
        params.append(f"id={object_id}")
    url = reverse("accounts_portal:portal_dashboard")
    if params:
        url = f"{url}?{'&'.join(params)}"
    return redirect(f"{url}#{anchor}")


def _store_it_support_feedback(request, *, level, title, message, password=None):
    request.session["it_support_feedback"] = {
        "level": level,
        "title": title,
        "message": message,
        "password": password,
    }


def _parse_optional_time(raw_value):
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError as exc:
        raise ValidationError("Le format d'heure attendu est HH:MM.") from exc


def _position_required(expected_positions):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            position = get_user_position(request.user)
            if position not in expected_positions:
                return _deny_portal_access(request)
            return view_func(request, *args, **kwargs)
        return login_required(wrapper)
    return decorator


def _render_director_dashboard(request):
    workspace_context = _build_director_workspace_context(request)
    branch = workspace_context.get("branch")
    context = {
        **_build_portal_context(
            request,
            page_title="Dashboard Direction des Etudes",
            module_cards=["Pilotage academique", "Resultats", "Enseignants"],
        ),
        "dashboard_kind": "Direction des etudes",
        **workspace_context,
        "director_active_section": workspace_context.get("section", "home"),
        "quality_score": (workspace_context.get("quality") or {}).get("score", 0),
        "home_alerts_count": len(workspace_context.get("alerts") or []),
        "lesson_logs_count": len(workspace_context.get("recent_lesson_logs") or []),
    }
    return render(request, "portal/staff/director_dashboard.html", context)


def _parse_director_section(request, default="home"):
    section = (request.GET.get("section") or request.POST.get("section") or default).strip().lower()
    aliases = {
        # anciens noms → nouveaux noms
        "operations": "planification",
        "assignments": "planification",
        "edt": "planification",
        "schedule": "planification",
        "planning": "planification",
        "academic": "programme",
        "classes": "programme",
        "teachers": "enseignants",
        "documents": "enseignants",
        "transfers": "enseignants",
        "results": "evaluations",
        "publications": "evaluations",
        "settings": "enseignants",
        "stats": "evaluations",
        "notifications": "evaluations",
        "logs": "evaluations",
        "students": "programme",
        "anomalies": "evaluations",
    }
    section = aliases.get(section, section)
    allowed = {
        "home",
        "planification",
        "programme",
        "correspondances",
        "enseignants",
        "evaluations",
    }
    return section if section in allowed else default


def _parse_director_week_start(request):
    raw = (request.GET.get("week_start") or request.POST.get("week_start") or "").strip()
    if raw:
        try:
            week_start = datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    else:
        week_start = timezone.localdate()
    return week_start - timedelta(days=week_start.weekday())


def _director_classes_queryset(branch):
    qs = AcademicClass.objects.select_related("programme", "branch", "academic_year").filter(is_active=True)
    if branch:
        qs = qs.filter(branch=branch)
    return qs.annotate(student_count=Count("enrollments", filter=Q(enrollments__is_active=True))).order_by("level", "programme__title", "id")


def _director_semester_rows(branch):
    semesters = (
        Semester.objects.select_related("academic_class", "academic_class__programme", "academic_class__branch")
        .prefetch_related("ues__ecs")
        .filter(academic_class__is_active=True)
    )
    if branch:
        semesters = semesters.filter(academic_class__branch=branch)

    enrollments_by_class = defaultdict(list)
    if branch:
        enrollments_qs = AcademicEnrollment.objects.select_related("academic_class", "academic_year").filter(branch=branch, is_active=True)
    else:
        enrollments_qs = AcademicEnrollment.objects.select_related("academic_class", "academic_year").filter(is_active=True)
    for enrollment in enrollments_qs:
        enrollments_by_class[enrollment.academic_class_id].append(enrollment)

    rows = []
    for semester in semesters.order_by("academic_class__level", "academic_class__programme__title", "number", "id"):
        ec_ids = []
        for ue in semester.ues.all():
            for ec in ue.ecs.all():
                ec_ids.append(ec.id)
        ec_ids = list(dict.fromkeys(ec_ids))
        enrollments = enrollments_by_class.get(semester.academic_class_id, [])
        expected = len(ec_ids) * len(enrollments)
        entered = (
            ECGrade.objects.filter(enrollment__in=enrollments, ec_id__in=ec_ids, final_score__isnull=False).count()
            if expected
            else 0
        )
        progress = int((entered / expected) * 100) if expected else 0
        ready = expected > 0 and entered >= expected
        if semester.status == Semester.STATUS_PUBLISHED:
            state = "publie"
            tone = "emerald"
        elif semester.status == Semester.STATUS_FINALIZED:
            state = "pret_publication"
            tone = "blue"
        elif ready:
            state = "pret_validation"
            tone = "emerald"
        elif entered:
            state = "en_cours"
            tone = "amber"
        else:
            state = "pas_de_notes"
            tone = "rose"
        rows.append({
            "semester": semester,
            "class": semester.academic_class,
            "student_count": len(enrollments),
            "expected": expected,
            "entered": entered,
            "progress": progress,
            "ready": ready,
            "state": state,
            "tone": tone,
            "can_validate": ready and semester.status not in {Semester.STATUS_FINALIZED, Semester.STATUS_PUBLISHED},
            "can_publish": semester.status == Semester.STATUS_FINALIZED,
            "can_generate": semester.status == Semester.STATUS_PUBLISHED,
            "bulletins_count": AcademicBulletin.objects.filter(
                academic_class=semester.academic_class,
                semester=semester,
                bulletin_type=AcademicBulletin.TYPE_SEMESTER,
            ).exclude(status=AcademicBulletin.STATUS_CANCELLED).count(),
        })
    return rows


def _format_director_decimal(value):
    try:
        return f"{Decimal(str(value)):.2f}".replace(".", ",")
    except Exception:
        return "-"


def _build_director_workspace_context(request, *, toast=None):
    branch = _resolve_academic_branch(request)
    section = _parse_director_section(request)
    week_start = _parse_director_week_start(request)
    week_end = week_start + timedelta(days=7)
    director_overview = get_director_schedule_overview(branch, week_start) if branch else {"stats": {}, "quality": {"score": 0, "warnings": []}, "alerts": [], "timetable": {"events": [], "summary": {}, "day_event_counts": []}}
    schedule_stats = director_overview.get("stats") or {}
    quality = director_overview.get("quality") or {"score": 0, "warnings": []}
    alerts = director_overview.get("alerts") or []
    timetable = director_overview.get("timetable") or {"events": [], "summary": {}, "day_event_counts": [], "empty_days": []}
    class_rows = list(_director_classes_queryset(branch))
    semester_rows = _director_semester_rows(branch)
    semester_rows_by_class = defaultdict(list)
    for row in semester_rows:
        semester_rows_by_class[row["class"].id].append(row)

    class_cards = []
    rejected_semester_ids = {
        int(semester_id)
        for semester_id in (request.session.get("director_rejected_semester_ids") or [])
        if str(semester_id).isdigit()
    }
    for academic_class in class_rows:
        rows = semester_rows_by_class.get(academic_class.id, [])
        rejected_count = sum(1 for row in rows if row["semester"].id in rejected_semester_ids)
        ready_count = sum(1 for row in rows if row["can_validate"] or row["can_publish"])
        if rejected_count:
            workflow_bucket = "rejected"
        elif ready_count:
            workflow_bucket = "ready"
        else:
            workflow_bucket = "in_progress"
        class_cards.append({
            "class": academic_class,
            "semester_count": len(rows),
            "ready_to_validate_count": sum(1 for row in rows if row["can_validate"]),
            "ready_to_publish_count": sum(1 for row in rows if row["can_publish"]),
            "published_count": sum(1 for row in rows if row["can_generate"]),
            "rejected_count": rejected_count,
            "progress": int(sum(row["progress"] for row in rows) / len(rows)) if rows else 0,
            "student_count": academic_class.student_count,
            "workflow_bucket": workflow_bucket,
            "rows": rows,
        })

    selected_class = None
    raw_class = (request.GET.get("class_id") or request.POST.get("class_id") or "").strip()
    if raw_class.isdigit():
        class_id = int(raw_class)
        selected_class = next((item["class"] for item in class_cards if item["class"].id == class_id), None)
    if selected_class is None and section in {"academic", "students", "operations"} and class_cards:
        selected_class = class_cards[0]["class"]

    selected_class_rows = semester_rows_by_class.get(getattr(selected_class, "id", None), []) if selected_class else []

    student_q = (request.GET.get("student_q") or request.POST.get("student_q") or "").strip()
    selected_class_students = []
    selected_class_student_count = 0
    selected_student_entry = None
    selected_class_schedule = None
    if selected_class is not None:
        student_enrollments = (
            AcademicEnrollment.objects.select_related(
                "academic_class",
                "academic_year",
                "student",
                "student__profile",
                "student__student_profile",
                "student__student_profile__inscription__candidature",
            )
            .filter(
                academic_class=selected_class,
                academic_year=selected_class.academic_year,
                is_active=True,
            )
        )
        if student_q:
            student_enrollments = student_enrollments.filter(
                Q(student__first_name__icontains=student_q)
                | Q(student__last_name__icontains=student_q)
                | Q(student__username__icontains=student_q)
                | Q(student__profile__employee_code__icontains=student_q)
                | Q(student__student_profile__matricule__icontains=student_q)
                | Q(student__student_profile__inscription__candidature__first_name__icontains=student_q)
                | Q(student__student_profile__inscription__candidature__last_name__icontains=student_q)
            )
        selected_class_student_count = student_enrollments.count()
        for enrollment in student_enrollments.order_by(
            "student__student_profile__inscription__candidature__last_name",
            "student__student_profile__inscription__candidature__first_name",
            "student__username",
        )[:40]:
            student_profile = getattr(enrollment.student, "student_profile", None)
            selected_class_students.append({
                "enrollment": enrollment,
                "student": enrollment.student,
                "full_name": getattr(student_profile, "full_name", "") or enrollment.student.get_full_name() or enrollment.student.username,
                "matricule": getattr(student_profile, "matricule", "") or "Matricule absent",
                "email": getattr(student_profile, "email", "") or enrollment.student.email or "Email absent",
                "is_enrolled": bool(getattr(student_profile, "is_enrolled", False)),
            })
        selected_class_schedule = get_class_week_schedule(selected_class, week_start)
        raw_student_id = (request.GET.get("student_id") or request.POST.get("student_id") or "").strip()
        if raw_student_id.isdigit():
            selected_student_entry = next(
                (
                    item
                    for item in selected_class_students
                    if item["student"].id == int(raw_student_id) or item["enrollment"].id == int(raw_student_id)
                ),
                None,
            )

    classroom_ops_context = build_director_classroom_ops_context(
        class_cards=class_cards,
        selected_class=selected_class,
        selected_class_rows=selected_class_rows,
        selected_class_student_count=selected_class_student_count,
        selected_class_schedule=selected_class_schedule,
    )

    teacher_load_map = schedule_stats.get("teacher_load") or {}
    class_load_map = schedule_stats.get("class_load") or {}
    teacher_q = (request.GET.get("teacher_q") or request.POST.get("teacher_q") or "").strip()
    teacher_context = build_director_teacher_assignment_context(
        branch=branch,
        schedule_stats=schedule_stats,
        teacher_q=teacher_q,
    )
    teacher_rows = teacher_context["teacher_rows"]
    teacher_rows_page = paginate_queryset(request, teacher_rows, per_page=20, page_param="teachers_page")
    class_cards_page = paginate_queryset(request, class_cards, per_page=15, page_param="classes_page")
    teachers_query_suffix = f"section=teachers&teacher_q={quote(teacher_q)}"
    classes_query_suffix = "section=results"

    class_load_items = sorted(
        (class_load_map or {}).items(),
        key=lambda item: (-item[1]["hours"], item[0]),
    )[:5]
    teacher_load_items = sorted(
        (teacher_load_map or {}).items(),
        key=lambda item: (-item[1]["hours"], item[0]),
    )[:5]

    selected_semester = None
    selected_semester_row = None
    raw_semester = (request.GET.get("semester_id") or request.POST.get("semester_id") or "").strip()
    if raw_semester.isdigit():
        selected_semester_row = next((row for row in semester_rows if row["semester"].id == int(raw_semester)), None)
        if selected_semester_row:
            selected_semester = selected_semester_row["semester"]
    if selected_semester_row and selected_class and selected_semester_row["class"].id != selected_class.id:
        selected_semester_row = None
        selected_semester = None
    if selected_semester is None and section == "results" and selected_class_rows:
        preferred = next((row for row in selected_class_rows if row["can_validate"] or row["can_publish"]), selected_class_rows[0])
        selected_semester_row = preferred
        selected_semester = preferred["semester"]

    results_class_queues = {
        "ready": [item for item in class_cards if item["workflow_bucket"] == "ready"],
        "in_progress": [item for item in class_cards if item["workflow_bucket"] == "in_progress"],
        "rejected": [item for item in class_cards if item["workflow_bucket"] == "rejected"],
    }
    results_query_suffix = (
        f"section=results&class_id={selected_class.id}&semester_id={selected_semester.id}"
        if selected_class is not None and selected_semester is not None
        else "section=results"
    )
    result_table_rows = []
    result_anomalies = []
    result_summary = None
    if selected_semester is not None:
        semester_enrollments = list(
            AcademicEnrollment.objects.select_related(
                "student",
                "student__student_profile",
                "student__student_profile__inscription__candidature",
                "academic_class",
                "academic_year",
            ).filter(
                academic_class=selected_semester.academic_class,
                academic_year=selected_semester.academic_class.academic_year,
                is_active=True,
            )
        )
        ec_ids = list(selected_semester.ues.values_list("ecs__id", flat=True).distinct())
        ec_ids = [ec_id for ec_id in ec_ids if ec_id is not None]
        grades_by_enrollment = defaultdict(dict)
        for grade in ECGrade.objects.filter(enrollment__in=semester_enrollments, ec_id__in=ec_ids).select_related("ec"):
            grades_by_enrollment[grade.enrollment_id][grade.ec_id] = grade
        ranking_rows = []
        total_average = Decimal("0.00")
        average_count = 0
        completed_students = 0
        for enrollment in semester_enrollments:
            threshold = resolve_threshold(enrollment)
            student_profile = getattr(enrollment.student, "student_profile", None)
            student_name = getattr(student_profile, "full_name", "") or enrollment.student.get_full_name() or enrollment.student.username
            grade_map = grades_by_enrollment.get(enrollment.id, {})
            missing_count = sum(1 for ec_id in ec_ids if grade_map.get(ec_id) is None or grade_map[ec_id].final_score is None)
            semester_result = compute_semester_result(selected_semester, enrollment)
            average = semester_result["average"]
            if average is not None:
                total_average += Decimal(str(average))
                average_count += 1
            if missing_count == 0:
                completed_students += 1
            if missing_count:
                result_anomalies.append({
                    "level": "blocking",
                    "student": student_name,
                    "message": f"{missing_count} note(s) manquante(s) sur {len(ec_ids)}.",
                })
            elif average < threshold:
                result_anomalies.append({
                    "level": "attention",
                    "student": student_name,
                    "message": f"Moyenne sous le seuil ({_format_director_decimal(average)}/{_format_director_decimal(threshold)}).",
                })
            ranking_rows.append({
                "enrollment": enrollment,
                "result": semester_result,
                "missing_count": missing_count,
                "threshold": threshold,
                "student_name": student_name,
            })

        ranking_rows.sort(
            key=lambda item: (
                item["missing_count"] > 0,
                -(Decimal(str(item["result"]["average"])) if item["result"]["average"] is not None else Decimal("0.00")),
                item["student_name"],
            )
        )
        for index, item in enumerate(ranking_rows, start=1):
            enrollment = item["enrollment"]
            semester_result = item["result"]
            student_profile = getattr(enrollment.student, "student_profile", None)
            average = semester_result["average"]
            result_table_rows.append({
                "rank": index,
                "student_name": getattr(student_profile, "full_name", "") or enrollment.student.get_full_name() or enrollment.student.username,
                "matricule": getattr(student_profile, "matricule", "") or "-",
                "average": _format_director_decimal(average),
                "credits": f"{_format_director_decimal(semester_result['credit_obtained'])}/{_format_director_decimal(semester_result['credit_required'])}",
                "completion": f"{len(ec_ids) - item['missing_count']}/{len(ec_ids)}",
                "status": "Bloque" if item["missing_count"] else ("Valide" if semester_result["is_validated"] else "A surveiller"),
                "status_tone": "rose" if item["missing_count"] else ("emerald" if semester_result["is_validated"] else "amber"),
            })
        result_summary = {
            "student_count": len(semester_enrollments),
            "completed_students": completed_students,
            "blocked_students": len(semester_enrollments) - completed_students,
            "class_average": _format_director_decimal(total_average / average_count) if average_count else "-",
            "anomalies_count": len(result_anomalies),
        }

    result_anomalies_page = paginate_queryset(request, result_anomalies, per_page=12, page_param="anomalies_page")
    result_table_rows_page = paginate_queryset(request, result_table_rows, per_page=25, page_param="results_page")

    ready_to_publish = [row for row in semester_rows if row["can_publish"]]
    ready_to_validate = [row for row in semester_rows if row["can_validate"]]
    in_progress = [row for row in semester_rows if row["state"] == "en_cours"]
    without_notes = [row for row in semester_rows if row["state"] == "pas_de_notes"]
    published = [row for row in semester_rows if row["can_generate"]]
    diploma_candidate_classes = []
    for item in class_cards:
        academic_class = item["class"]
        is_terminal = str(academic_class.level or "").upper().strip() in {"L3", "M2"}
        if not is_terminal:
            continue
        if item["semester_count"] and item["published_count"] >= item["semester_count"]:
            diploma_candidate_classes.append({
                "class": academic_class,
                "student_count": item["student_count"],
                "awards_count": AcademicDiplomaAward.objects.filter(
                    academic_class=academic_class,
                ).exclude(status=AcademicDiplomaAward.STATUS_CANCELLED).count(),
            })
    document_workflows = [
        {
            "title": "Contrats enseignants",
            "summary": "Generer et suivre les contrats pedagogiques des enseignants.",
            "status": "backend_a_implanter",
            "tone": "amber",
        },
        {
            "title": "Dossiers enseignants",
            "summary": "Uploader, verifier et centraliser les pieces administratives.",
            "status": "backend_a_implanter",
            "tone": "blue",
        },
        {
            "title": "Transferts classe / ecole",
            "summary": "Constituer, valider et transmettre les dossiers de transfert.",
            "status": "backend_a_implanter",
            "tone": "rose",
        },
    ]
    assignments_alerts = [
        alert
        for alert in alerts
        if alert.get("type") in {
            "missing_teacher",
            "missing_location",
            "teacher_overload",
            "high_cancellation_rate",
            "unresolved_conflict",
            "class_without_events",
        }
    ]
    upcoming_events = list((timetable.get("events") or [])[:8])
    recent_lesson_logs = []
    if branch:
        recent_lesson_logs = list(
            LessonLog.objects.select_related(
                "academic_class",
                "ec",
                "teacher",
                "branch",
            )
            .filter(branch=branch)
            .order_by("-date", "-start_time", "-id")[:8]
        )

    raw_document_teacher = (request.GET.get("teacher_id") or request.POST.get("teacher_id") or "").strip()
    document_teacher_id = int(raw_document_teacher) if raw_document_teacher.isdigit() else None
    document_context = build_director_document_context(
        branch=branch,
        teacher_id=document_teacher_id,
    )
    transfer_context = build_director_transfer_context(branch=branch)

    teacher_form_ecs = list(
        EC.objects.select_related("ue", "ue__semester", "ue__semester__academic_class")
        .filter(ue__semester__academic_class__branch=branch, ue__semester__academic_class__is_active=True)
        .order_by("ue__semester__academic_class__level", "ue__semester__academic_class__programme__title", "title", "id")[:200]
    ) if branch else []

    planning_assignment_context = build_director_planning_assignment_context(
        assignments_alerts=assignments_alerts,
        class_load_items=class_load_items,
        teacher_load_items=teacher_load_items,
        upcoming_events=upcoming_events,
        recent_lesson_logs=recent_lesson_logs,
    )
    bulletin_scope_class = selected_class or (class_rows[0] if class_rows else None)

    # UE / EC pour la section Programme
    programme_ue_rows = []
    if selected_class is not None:
        semesters_qs = list(selected_class.semesters.prefetch_related("ues__ecs").order_by("number"))
        for sem in semesters_qs:
            programme_ue_rows.append({
                "semester": sem,
                "ues": list(sem.ues.prefetch_related("ecs").order_by("code", "id")),
            })

    admin_documents = list(
        AdministrativeDocument.objects.filter(branch=branch).order_by("-created_at")[:50]
    ) if branch else []
    admin_doc_type_choices = AdministrativeDocument.TYPE_CHOICES

    context = {
        "branch": branch,
        "section": section,
        "week_start": week_start,
        "week_end": week_end,
        "prev_week_start": week_start - timedelta(days=7),
        "next_week_start": week_start + timedelta(days=7),
        "director_overview": director_overview,
        "schedule_stats": schedule_stats,
        "quality": quality,
        "alerts": alerts,
        "timetable": timetable,
        "class_load_items": class_load_items,
        "teacher_load_items": teacher_load_items,
        "today": timezone.localdate(),
        "classes": class_rows,
        "class_cards": class_cards,
        "class_cards_page": class_cards_page,
        "classes_query_suffix": classes_query_suffix,
        "total_classes": len(class_rows),
        "total_semesters": len(semester_rows),
        "semester_rows": semester_rows,
        "ready_to_publish": ready_to_publish,
        "ready_to_validate": ready_to_validate,
        "in_progress": in_progress,
        "without_notes": without_notes,
        "published": published,
        "can_manage_bulletins": bool(bulletin_scope_class and can_manage_bulletins(request.user, bulletin_scope_class)),
        "can_manage_diplomas": get_user_position(request.user) in {"executive_director", "deputy_executive_director"},
        "diploma_candidate_classes": diploma_candidate_classes,
        "ready_to_validate_count": len(ready_to_validate),
        "published_count": len(published),
        "document_workflows": document_workflows,
        "document_teacher_rows": document_context["document_teacher_rows"],
        "selected_document_teacher": document_context["selected_document_teacher"],
        "teacher_documents": document_context["teacher_documents"],
        "teacher_document_type_choices": document_context["teacher_document_type_choices"],
        "transfer_enrollments": transfer_context["transfer_enrollments"],
        "transfer_rows": transfer_context["transfer_rows"],
        "transfer_target_classes": transfer_context["transfer_target_classes"],
        "transfer_type_choices": transfer_context["transfer_type_choices"],
        "teachers": teacher_rows,
        "teacher_rows_page": teacher_rows_page,
        "teachers_query_suffix": teachers_query_suffix,
        "teachers_total": teacher_context["teachers_total"],
        "teacher_unassigned_count": teacher_context["teacher_unassigned_count"],
        "teacher_assigned_count": teacher_context["teachers_total"] - teacher_context["teacher_unassigned_count"],
        "teacher_q": teacher_q,
        "teacher_form_classes": class_rows,
        "teacher_form_ecs": teacher_form_ecs,
        "selected_semester": selected_semester,
        "selected_semester_row": selected_semester_row,
        "selected_class": selected_class,
        "selected_class_rows": selected_class_rows,
        "selected_class_students": selected_class_students,
        "selected_class_student_count": selected_class_student_count,
        "selected_class_schedule": selected_class_schedule,
        "selected_student_entry": selected_student_entry,
        "result_panel_title": (
            f"{selected_class.display_name} · Semestre {selected_semester.number}"
            if selected_class is not None and selected_semester is not None
            else "Validation semestre"
        ),
        "result_panel_subtitle": (
            f"{selected_semester_row['entered']}/{selected_semester_row['expected']} notes · {selected_semester.get_status_display()}"
            if selected_semester_row is not None and selected_semester is not None
            else "Verifier les anomalies puis decider."
        ),
        "operation_class_rows": classroom_ops_context["operation_class_rows"],
        "selected_operation_class": classroom_ops_context["selected_operation_class"],
        "selected_operation_class_id": (
            classroom_ops_context["selected_operation_class"]["class"].id
            if classroom_ops_context["selected_operation_class"]
            else None
        ),
        "student_q": student_q,
        "assignments_alerts": assignments_alerts,
        "assignments_alerts_count": len(assignments_alerts),
        "upcoming_events": upcoming_events,
        "recent_lesson_logs": recent_lesson_logs,
        "assignment_alert_rows": planning_assignment_context["assignment_alert_rows"],
        "assignment_class_rows": planning_assignment_context["assignment_class_rows"],
        "assignment_teacher_rows": planning_assignment_context["assignment_teacher_rows"],
        "assignment_event_rows": planning_assignment_context["assignment_event_rows"],
        "assignment_log_rows": planning_assignment_context["assignment_log_rows"],
        "assignment_critical_count": planning_assignment_context["assignment_critical_count"],
        "assignment_missing_teacher_count": planning_assignment_context["assignment_missing_teacher_count"],
        "assignment_missing_room_count": planning_assignment_context["assignment_missing_room_count"],
        "assignment_teacher_overload_count": planning_assignment_context["assignment_teacher_overload_count"],
        "semester_rows_by_class": semester_rows_by_class,
        "results_class_queues": results_class_queues,
        "result_table_rows": result_table_rows,
        "result_table_rows_page": result_table_rows_page,
        "result_anomalies": result_anomalies,
        "result_anomalies_page": result_anomalies_page,
        "result_summary": result_summary,
        "results_query_suffix": results_query_suffix,
        "tasks_center": build_director_tasks_center(
            branch=branch,
            semester_rows=semester_rows,
            teacher_unassigned_count=teacher_context["teacher_unassigned_count"],
            result_anomalies=result_anomalies,
        ),
        "student_panel_subtitle": (
            f"{selected_student_entry['matricule']} · {selected_student_entry['email']}"
            if selected_student_entry is not None
            else "Fiche etudiant"
        ),
        "programme_ue_rows": programme_ue_rows,
        "admin_documents": admin_documents,
        "admin_doc_type_choices": admin_doc_type_choices,
        "admin_doc_draft_count": sum(1 for d in admin_documents if d.status == AdministrativeDocument.STATUS_DRAFT),
        "admin_doc_published_count": sum(1 for d in admin_documents if d.status == AdministrativeDocument.STATUS_PUBLISHED),
    }
    if toast:
        context["toast"] = toast
    return context


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_workspace(request):
    context = _build_director_workspace_context(request)
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_drawer(request):
    panel = (request.GET.get("panel") or "").strip().lower()
    section_map = {
        "operation": "operations",
        "result": "results",
        "student": "students",
    }
    template_map = {
        "operation": "portal/staff/director/partials/drawers/operation_drawer.html",
        "result": "portal/staff/director/partials/drawers/result_drawer.html",
        "student": "portal/staff/director/partials/drawers/student_drawer.html",
    }
    section = section_map.get(panel)
    template_name = template_map.get(panel)
    if not section or not template_name:
        return HttpResponseBadRequest("Panneau introuvable.")

    original_get = request.GET
    params = request.GET.copy()
    params["section"] = section
    request.GET = params
    try:
        context = _build_director_workspace_context(request)
    finally:
        request.GET = original_get

    if panel == "operation" and not context.get("selected_operation_class"):
        return HttpResponseBadRequest("Classe introuvable.")
    if panel == "result" and not context.get("selected_semester_row"):
        return HttpResponseBadRequest("Semestre introuvable.")
    if panel == "student" and not context.get("selected_student_entry"):
        return HttpResponseBadRequest("Etudiant introuvable.")

    return render(request, template_name, context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_programme_action(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    if not branch:
        return HttpResponseBadRequest("Aucune annexe.")
    action = (request.POST.get("action") or "").strip()
    class_id = (request.POST.get("class_id") or "").strip()
    toast = None
    try:
        if action == "save_ue":
            ue = save_ue(
                branch=branch,
                ue_id=(request.POST.get("ue_id") or "").strip() or None,
                semester_id=request.POST.get("semester_id"),
                code=request.POST.get("code"),
                title=request.POST.get("title"),
            )
            class_id = str(ue.semester.academic_class_id)
            toast = {"level": "success", "message": "Unite d'enseignement enregistree."}
        elif action == "save_ec":
            ec = save_ec(
                branch=branch,
                ec_id=(request.POST.get("ec_id") or "").strip() or None,
                ue_id=request.POST.get("ue_id"),
                title=request.POST.get("title"),
                coefficient=request.POST.get("coefficient"),
                credit_required=request.POST.get("credit_required"),
            )
            class_id = str(ec.ue.semester.academic_class_id)
            toast = {"level": "success", "message": "Element constitutif enregistre."}
        elif action == "delete_ec":
            delete_ec(branch=branch, ec_id=request.POST.get("ec_id"))
            toast = {"level": "success", "message": "EC supprime."}
        else:
            toast = {"level": "error", "message": "Action inconnue."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages)}

    params = request.GET.copy()
    params["section"] = "programme"
    if class_id:
        params["class_id"] = class_id
    request.GET = params
    context = _build_director_workspace_context(request, toast=toast)
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_correspondance_create(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    if not branch:
        return HttpResponseBadRequest("Aucune annexe.")
    doc_type = request.POST.get("doc_type", "").strip()
    title = request.POST.get("title", "").strip()
    body = request.POST.get("body", "").strip()
    reference = request.POST.get("reference", "").strip()
    recipients = request.POST.get("recipients", "").strip()
    action = request.POST.get("action", "draft")
    if not title or not body:
        return _build_workspace_response(request, toast={"level": "error", "message": "Titre et corps du document sont obligatoires."})
    status = AdministrativeDocument.STATUS_PUBLISHED if action == "publish" else AdministrativeDocument.STATUS_DRAFT
    AdministrativeDocument.objects.create(
        branch=branch,
        doc_type=doc_type,
        title=title,
        body=body,
        reference=reference,
        recipients=recipients,
        status=status,
        created_by=request.user,
    )
    msg = "Document publie avec succes." if status == AdministrativeDocument.STATUS_PUBLISHED else "Brouillon enregistre."
    return _build_workspace_response(request, section_override="correspondances", toast={"level": "success", "message": msg})


def _build_workspace_response(request, section_override=None, toast=None):
    if section_override:
        params = request.GET.copy()
        params["section"] = section_override
        request.GET = params
    context = _build_director_workspace_context(request, toast=toast)
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_correspondance_publish(request, doc_id):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    doc = AdministrativeDocument.objects.filter(id=doc_id, branch=branch).first()
    if not doc:
        return HttpResponseBadRequest("Document introuvable.")
    doc.status = AdministrativeDocument.STATUS_PUBLISHED
    doc.save(update_fields=["status", "updated_at"])
    return _build_workspace_response(request, section_override="correspondances", toast={"level": "success", "message": "Document publie."})


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_correspondance_pdf(request, doc_id):
    branch = _resolve_academic_branch(request)
    doc = AdministrativeDocument.objects.filter(id=doc_id, branch=branch).first()
    if not doc:
        return HttpResponseBadRequest("Document introuvable.")
    html_content = render(request, "portal/admin/correspondances/pdf.html", {
        "doc": doc,
        "branch": branch,
        "today": timezone.localdate(),
    })
    try:
        from weasyprint import HTML as WeasyprintHTML
        pdf_bytes = WeasyprintHTML(string=html_content.content, base_url=request.build_absolute_uri("/")).write_pdf()
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        safe_title = doc.title[:40].replace(" ", "_")
        response["Content-Disposition"] = f'inline; filename="{safe_title}.pdf"'
        return response
    except Exception:
        return html_content


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_results_action(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    action = (request.POST.get("action") or "").strip()

    if action == "publish":
        return _director_request_publish_otp(request, branch=branch)

    toast = {"level": "error", "message": "Action impossible."}
    rejected_semester_ids = {
        int(semester_id)
        for semester_id in (request.session.get("director_rejected_semester_ids") or [])
        if str(semester_id).isdigit()
    }
    try:
        semester = Semester.objects.select_related("academic_class", "academic_class__branch", "academic_class__academic_year").get(pk=request.POST.get("semester_id"))
        if branch and semester.academic_class.branch_id != branch.id:
            raise ValidationError("Action hors annexe refusee.")
        enrollments = list(AcademicEnrollment.objects.filter(academic_class=semester.academic_class, academic_year=semester.academic_class.academic_year, is_active=True))
        previous_status = semester.status
        if action == "validate":
            if not can_publish_semester(semester, enrollments):
                raise ValidationError("Toutes les notes doivent etre renseignees avant validation.")
            semester.status = Semester.STATUS_FINALIZED
            semester.save(update_fields=["status"])
            rejected_semester_ids.discard(semester.id)
            toast = {"level": "success", "message": "Resultats valides. Publication possible."}
            log_action(
                request.user,
                "semester.validated",
                semester,
                old_values={"status": previous_status},
                new_values={"status": semester.status},
                branch=semester.academic_class.branch,
                academic_year=semester.academic_class.academic_year,
                request=request,
            )
        elif action == "reject":
            if semester.status == Semester.STATUS_PUBLISHED:
                raise ValidationError("Un semestre publie ne peut pas etre rejete ici.")
            semester.status = Semester.STATUS_NORMAL_ENTRY
            semester.save(update_fields=["status"])
            rejected_semester_ids.add(semester.id)
            toast = {"level": "success", "message": "Resultats renvoyes en correction."}
            log_action(
                request.user,
                "semester.rejected",
                semester,
                old_values={"status": previous_status},
                new_values={"status": semester.status},
                branch=semester.academic_class.branch,
                academic_year=semester.academic_class.academic_year,
                request=request,
            )
        else:
            raise ValidationError("Action inconnue.")
    except (Semester.DoesNotExist, ValidationError) as exc:
        message = " ".join(getattr(exc, "messages", [])) if hasattr(exc, "messages") else str(exc)
        toast = {"level": "error", "message": message or "Action impossible."}
    request.session["director_rejected_semester_ids"] = sorted(rejected_semester_ids)
    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "results"
    return render(request, "portal/staff/director/partials/workspace.html", context)


def _director_request_publish_otp(request, *, branch):
    """Publication = action irreversible : declenche une demande OTP au DG/DGA
    de l'annexe au lieu de changer le statut directement (cf.
    CAHIER_DES_CHARGES_DIRECTEUR_ETUDES.md, 2.1)."""
    try:
        semester = Semester.objects.select_related(
            "academic_class", "academic_class__branch", "academic_class__academic_year"
        ).get(pk=request.POST.get("semester_id"))
        if branch and semester.academic_class.branch_id != branch.id:
            raise ValidationError("Action hors annexe refusee.")
        if semester.status != Semester.STATUS_FINALIZED:
            raise ValidationError("Le semestre doit etre valide avant publication.")
        otp_request = request_sensitive_action(
            branch=semester.academic_class.branch,
            action_type=SensitiveActionRequest.ACTION_SEMESTER_PUBLISH,
            target_model="Semester",
            target_id=semester.pk,
            previous_state={"status": semester.status},
            requested_state={"status": Semester.STATUS_PUBLISHED},
            requested_by=request.user,
        )
    except (Semester.DoesNotExist, ValidationError, SensitiveActionError) as exc:
        message = " ".join(getattr(exc, "messages", [])) if hasattr(exc, "messages") else str(exc)
        return render(
            request,
            "portal/staff/director/partials/results_otp_modal.html",
            {"otp_error": message or "Publication impossible."},
        )

    return render(
        request,
        "portal/staff/director/partials/results_otp_modal.html",
        {
            "otp_request_id": otp_request.pk,
            "otp_validity_minutes": SensitiveActionRequest.OTP_VALIDITY_MINUTES,
            "class_label": semester.academic_class.display_name,
            "semester_number": semester.number,
        },
    )


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_results_confirm_otp(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    otp_request_id = request.POST.get("otp_request_id")
    otp_code = (request.POST.get("otp_code") or "").strip()

    def _apply(otp_request):
        semester = Semester.objects.select_related(
            "academic_class", "academic_class__branch", "academic_class__academic_year"
        ).get(pk=otp_request.target_id)
        previous_status = semester.status
        semester.status = Semester.STATUS_PUBLISHED
        semester.save(update_fields=["status"])

        rejected_semester_ids = {
            int(semester_id)
            for semester_id in (request.session.get("director_rejected_semester_ids") or [])
            if str(semester_id).isdigit()
        }
        rejected_semester_ids.discard(semester.id)
        request.session["director_rejected_semester_ids"] = sorted(rejected_semester_ids)

        log_action(
            otp_request.requested_by,
            "semester.published",
            semester,
            old_values={"status": previous_status},
            new_values={"status": semester.status},
            branch=semester.academic_class.branch,
            academic_year=semester.academic_class.academic_year,
            reason=otp_request.reason,
            request=request,
        )
        return {"status": semester.status}

    toast = None
    try:
        confirm_sensitive_action(
            request_id=otp_request_id,
            code=otp_code,
            approver=request.user,
            apply_callback=_apply,
            skip_financial_audit=True,
        )
        toast = {"level": "success", "message": "Resultats publies."}
    except (SensitiveActionError, SensitiveActionRequest.DoesNotExist) as exc:
        message = str(exc) if isinstance(exc, SensitiveActionError) else "Demande introuvable."
        toast = {"level": "error", "message": message}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "results"
    response = render(request, "portal/staff/director/partials/workspace.html", context)
    if toast["level"] == "success":
        response["HX-Trigger"] = "director-modal-close"
    return response


@_position_required({"director_of_studies", "super_admin"})
def director_bulletin_action(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    action = (request.POST.get("action") or "").strip()
    toast = {"level": "error", "message": "Action bulletin impossible."}
    try:
        if action == "generate_semester_class":
            semester = Semester.objects.select_related("academic_class", "academic_class__branch").get(pk=request.POST.get("semester_id"))
            if branch and semester.academic_class.branch_id != branch.id:
                raise ValidationError("Action hors annexe refusee.")
            if not can_manage_bulletins(request.user, semester.academic_class):
                raise ValidationError("Seul le Directeur des etudes peut delivrer les bulletins.")
            bulletins = generate_semester_bulletins_for_class(
                academic_class=semester.academic_class,
                semester=semester,
                actor=request.user,
                publish=True,
            )
            toast = {"level": "success", "message": f"{len(bulletins)} bulletin(s) semestriel(s) generes."}
        elif action == "generate_annual_class":
            academic_class = AcademicClass.objects.select_related("branch", "academic_year").get(pk=request.POST.get("class_id"))
            if branch and academic_class.branch_id != branch.id:
                raise ValidationError("Action hors annexe refusee.")
            if not can_manage_bulletins(request.user, academic_class):
                raise ValidationError("Seul le Directeur des etudes peut delivrer les bulletins.")
            bulletins = generate_annual_bulletins_for_class(
                academic_class=academic_class,
                actor=request.user,
                publish=True,
            )
            toast = {"level": "success", "message": f"{len(bulletins)} bulletin(s) annuel(s) generes."}
        else:
            raise ValidationError("Action bulletin inconnue.")
    except (Semester.DoesNotExist, AcademicClass.DoesNotExist):
        toast = {"level": "error", "message": "Classe ou semestre introuvable."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages)}
    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "publications"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required(BULLETIN_MANAGEMENT_POSITIONS)
def director_export_report_xlsx(request):
    branch = _resolve_academic_branch(request)
    wb = build_academic_report_xlsx(branch=branch)
    filename = f"rapport_pedagogique_{branch.slug if branch else 'annexe'}.xlsx"
    return xlsx_response(wb, filename)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_create(request):
    branch = _resolve_academic_branch(request)
    base_context = _build_director_workspace_context(request)

    if request.method == "GET":
        return render(
            request,
            "portal/staff/director/modals/teacher_create_modal.html",
            {
                **base_context,
                "teacher_form_classes": base_context.get("teacher_form_classes", []),
                "teacher_form_ecs": base_context.get("teacher_form_ecs", []),
            },
        )

    if request.method != "POST":
        return _deny_portal_access(request)

    toast = {"level": "error", "message": "Creation impossible."}

    try:
        validate_email((request.POST.get("email") or "").strip())
        creation = create_teacher_with_account(
            {
                "first_name": request.POST.get("first_name"),
                "last_name": request.POST.get("last_name"),
                "email": request.POST.get("email"),
                "phone": request.POST.get("phone"),
                "teacher_hourly_rate": request.POST.get("teacher_hourly_rate"),
                "specialty": request.POST.get("specialty"),
                "class_id": request.POST.get("class_id"),
                "ec_id": request.POST.get("ec_id"),
                "room_label": request.POST.get("room_label"),
                "planned_hours": request.POST.get("planned_hours"),
            },
            request.user,
        )
        toast = {"level": "success", "message": "Enseignant cree, affecte et acces generes."}
        response = render(
            request,
            "portal/staff/director/partials/workspace.html",
            {
                **_build_director_workspace_context(request, toast=toast),
                "section": "teachers",
            },
        )
        response["HX-Trigger"] = "director-modal-close"
        return response
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages) or "Creation impossible."}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "teachers"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_assign(request):
    """Modale d'affectation enseignant -> classe/EC/salle (GET: modal, POST: sauvegarde)."""
    branch = _resolve_academic_branch(request)
    User = get_user_model()

    teacher_id_raw = (request.GET.get("teacher_id") or request.POST.get("teacher_id") or "").strip()
    teacher_id = int(teacher_id_raw) if teacher_id_raw.isdigit() else None
    teacher = None
    if teacher_id:
        teacher = User.objects.select_related("profile", "profile__branch").filter(
            id=teacher_id, is_active=True, profile__position="teacher"
        ).first()

    if request.method == "GET":
        classes_qs = AcademicClass.objects.select_related("programme", "academic_year").filter(is_active=True)
        if branch:
            classes_qs = classes_qs.filter(branch=branch)
        ecs_qs = EC.objects.select_related(
            "ue", "ue__semester", "ue__semester__academic_class"
        )
        if branch:
            ecs_qs = ecs_qs.filter(ue__semester__academic_class__branch=branch)
        existing = []
        if teacher:
            existing = list(
                DirectorTeacherAssignment.objects.select_related("academic_class", "ec").filter(
                    teacher=teacher, is_active=True
                )
            )
        selected_class_id = ""
        selected_ec_id = ""
        selected_room_label = ""
        selected_planned_hours = ""
        if existing:
            selected_class_id = str(existing[0].academic_class_id or "")
            selected_ec_id = str(existing[0].ec_id or "")
            selected_room_label = existing[0].room_label or ""
            selected_planned_hours = str(existing[0].planned_hours or "")
        return render(request, "portal/staff/director/modals/teacher_assign_modal.html", {
            "assign_teacher": teacher,
            "assign_teacher_name": teacher.get_full_name() or teacher.username if teacher else "",
            "assign_classes": list(classes_qs.order_by("level", "programme__title")[:80]),
            "assign_ecs": list(ecs_qs.order_by("title")[:120]),
            "assign_existing": existing,
            "assign_selected_class_id": selected_class_id,
            "assign_selected_ec_id": selected_ec_id,
            "assign_selected_room_label": selected_room_label,
            "assign_selected_planned_hours": selected_planned_hours,
        })

    # POST
    action = (request.POST.get("action") or "add").strip()
    toast = {"level": "error", "message": "Action impossible."}
    try:
        if teacher is None:
            raise ValidationError("Enseignant introuvable.")
        if branch and getattr(getattr(teacher, "profile", None), "branch_id", None) != branch.id:
            raise ValidationError("Affectation hors perimetre refusee.")

        if action == "remove":
            asg_id_raw = (request.POST.get("assignment_id") or "").strip()
            asg_id = int(asg_id_raw) if asg_id_raw.isdigit() else None
            if not asg_id:
                raise ValidationError("Affectation introuvable.")
            deleted, _ = DirectorTeacherAssignment.objects.filter(id=asg_id, teacher=teacher).delete()
            if not deleted:
                raise ValidationError("Affectation introuvable ou deja supprimee.")
            toast = {"level": "success", "message": "Affectation retiree."}
        else:
            class_id_raw = (request.POST.get("class_id") or "").strip()
            class_id = int(class_id_raw) if class_id_raw.isdigit() else None
            ec_id_raw = (request.POST.get("ec_id") or "").strip()
            ec_id = int(ec_id_raw) if ec_id_raw.isdigit() else None
            room_label = (request.POST.get("room_label") or "").strip()
            planned_hours_raw = (request.POST.get("planned_hours") or "").strip()

            if not class_id:
                raise ValidationError("La classe est obligatoire.")
            if not room_label:
                raise ValidationError("La salle est obligatoire pour cette affectation.")
            if not planned_hours_raw:
                raise ValidationError("Le volume horaire est obligatoire pour cette affectation.")
            academic_class = AcademicClass.objects.filter(id=class_id, is_active=True).first()
            if academic_class is None:
                raise ValidationError("Classe introuvable.")
            if branch and academic_class.branch_id != branch.id:
                raise ValidationError("Classe hors perimetre refusee.")
            ec_obj = EC.objects.filter(id=ec_id).first() if ec_id else None
            if ec_obj is not None and ec_obj.ue.semester.academic_class_id != academic_class.id:
                raise ValidationError("La matiere selectionnee n'appartient pas a la classe choisie.")

            assignment, created = DirectorTeacherAssignment.objects.get_or_create(
                teacher=teacher,
                academic_class=academic_class,
                ec=ec_obj,
                defaults={
                    "branch": branch or academic_class.branch,
                    "created_by": request.user,
                    "is_active": True,
                    "room_label": room_label,
                    "planned_hours": planned_hours_raw.replace(",", "."),
                },
            )
            if not created:
                assignment.branch = branch or academic_class.branch
                assignment.is_active = True
                assignment.room_label = room_label
                assignment.planned_hours = planned_hours_raw.replace(",", ".")
                if assignment.created_by_id is None:
                    assignment.created_by = request.user
                assignment.save(update_fields=["branch", "is_active", "room_label", "planned_hours", "created_by", "updated_at"])
            toast = {
                "level": "success",
                "message": (
                    f"Affectation mise a jour : {academic_class.display_name} -> {room_label}, {planned_hours_raw} h."
                    if not created
                    else f"Affectation enregistree : {academic_class.display_name} -> {room_label}, {planned_hours_raw} h."
                ),
            }

        response = render(
            request,
            "portal/staff/director/partials/workspace.html",
            {**_build_director_workspace_context(request, toast=toast), "section": "teachers"},
        )
        response["HX-Trigger"] = "director-modal-close"
        return response

    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(getattr(exc, "messages", [])) or str(exc)}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "teachers"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_contract_download(request, teacher_id: int):
    branch = _resolve_academic_branch(request)
    User = get_user_model()
    teacher = User.objects.select_related("profile", "profile__branch").filter(
        id=teacher_id,
        profile__position="teacher",
    ).first()
    if teacher is None:
        return HttpResponseBadRequest("Enseignant introuvable.")
    if branch and getattr(getattr(teacher, "profile", None), "branch_id", None) != branch.id:
        return HttpResponseForbidden("Acces inter-annexes refuse.")

    try:
        pdf_bytes = generate_teacher_contract_pdf(teacher)
    except Exception as exc:
        return HttpResponseBadRequest(str(exc) or "Generation du contrat impossible.")

    filename = f"contrat-enseignant-{teacher.username}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_document_upload(request):
    base_context = _build_director_workspace_context(request)
    teacher_id_raw = (request.GET.get("teacher_id") or request.POST.get("teacher_id") or "").strip()
    teacher_id = int(teacher_id_raw) if teacher_id_raw.isdigit() else None

    if request.method == "GET":
        return render(
            request,
            "portal/staff/director/modals/teacher_document_upload_modal.html",
            {
                **base_context,
                "selected_document_teacher": next(
                    (teacher for teacher in base_context.get("document_teacher_rows", []) if teacher.id == teacher_id),
                    base_context.get("selected_document_teacher"),
                ),
                "teacher_document_type_choices": base_context.get("teacher_document_type_choices", []),
            },
        )
    if request.method != "POST":
        return _deny_portal_access(request)

    toast = {"level": "error", "message": "Upload impossible."}

    try:
        if teacher_id is None:
            raise ValidationError("Selectionnez un enseignant.")
        upload_teacher_document(
            user=request.user,
            teacher_id=teacher_id,
            document_type=(request.POST.get("document_type") or "").strip(),
            file=request.FILES.get("file"),
            note=request.POST.get("note"),
        )
        toast = {"level": "success", "message": "Piece enseignant televersee."}
        response = render(
            request,
            "portal/staff/director/partials/workspace.html",
            {
                **_build_director_workspace_context(request, toast=toast),
                "section": "documents",
            },
        )
        response["HX-Trigger"] = "director-modal-close"
        return response
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages) or "Upload impossible."}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "documents"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_documents_modal(request):
    base_context = _build_director_workspace_context(request)
    teacher_id_raw = (request.GET.get("teacher_id") or "").strip()
    teacher_id = int(teacher_id_raw) if teacher_id_raw.isdigit() else None

    selected_teacher = next(
        (teacher for teacher in base_context.get("document_teacher_rows", []) if teacher.id == teacher_id),
        base_context.get("selected_document_teacher"),
    )

    return render(
        request,
        "portal/staff/director/modals/teacher_documents_modal.html",
        {
            **base_context,
            "selected_document_teacher": selected_teacher,
            "teacher_documents": base_context.get("teacher_documents", []),
        },
    )


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_transfer_create(request):
    base_context = _build_director_workspace_context(request)
    if request.method == "GET":
        return render(
            request,
            "portal/staff/director/modals/transfer_create_modal.html",
            {
                **base_context,
                "transfer_enrollments": base_context.get("transfer_enrollments", []),
                "transfer_target_classes": base_context.get("transfer_target_classes", []),
                "transfer_type_choices": base_context.get("transfer_type_choices", []),
            },
        )
    if request.method != "POST":
        return _deny_portal_access(request)

    toast = {"level": "error", "message": "Creation du transfert impossible."}
    try:
        enrollment_raw = (request.POST.get("enrollment_id") or "").strip()
        enrollment_id = int(enrollment_raw) if enrollment_raw.isdigit() else None
        target_class_raw = (request.POST.get("target_class_id") or "").strip()
        target_class_id = int(target_class_raw) if target_class_raw.isdigit() else None
        create_transfer_request(
            user=request.user,
            enrollment_id=enrollment_id,
            transfer_type=(request.POST.get("transfer_type") or "").strip(),
            target_class_id=target_class_id,
            target_school_name=request.POST.get("target_school_name"),
            reason=request.POST.get("reason"),
            attachment=request.FILES.get("attachment"),
        )
        toast = {"level": "success", "message": "Demande de transfert enregistree."}
        response = render(
            request,
            "portal/staff/director/partials/workspace.html",
            {
                **_build_director_workspace_context(request, toast=toast),
                "section": "documents",
            },
        )
        response["HX-Trigger"] = "director-modal-close"
        return response
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages) or "Creation du transfert impossible."}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "documents"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_teacher_document_review(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    toast = {"level": "error", "message": "Decision sur la piece impossible."}
    try:
        document_raw = (request.POST.get("document_id") or "").strip()
        document_id = int(document_raw) if document_raw.isdigit() else None
        action = (request.POST.get("action") or "").strip().lower()
        if document_id is None:
            raise ValidationError("Document introuvable.")
        review_teacher_document(
            user=request.user,
            document_id=document_id,
            verify=action == "verify",
        )
        toast = {"level": "success", "message": "Etat du document mis a jour."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages) or "Decision sur la piece impossible."}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "documents"
    return render(request, "portal/staff/director/partials/workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_transfer_review(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    toast = {"level": "error", "message": "Decision sur le transfert impossible."}
    try:
        transfer_raw = (request.POST.get("transfer_id") or "").strip()
        transfer_id = int(transfer_raw) if transfer_raw.isdigit() else None
        if transfer_id is None:
            raise ValidationError("Demande de transfert introuvable.")
        review_transfer_request(
            user=request.user,
            transfer_id=transfer_id,
            action=request.POST.get("action"),
        )
        toast = {"level": "success", "message": "Demande de transfert mise a jour."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages) or "Decision sur le transfert impossible."}

    context = _build_director_workspace_context(request, toast=toast)
    context["section"] = "documents"
    return render(request, "portal/staff/director/partials/workspace.html", context)


def _user_initials(display_name: str) -> str:
    parts = [p for p in (display_name or "").strip().split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return (parts[0][:2] if len(parts[0]) > 1 else parts[0] + "?").upper()
    return (parts[0][0] + parts[-1][0]).upper()


def _render_it_dashboard(request, *, initial_module=None, initial_workspace_html=None):
    context = _build_it_dashboard_context(request)
    context.update(
        build_it_grade_selection_context(
            request.user,
            class_id=request.GET.get("class_id"),
            semester_id=request.GET.get("semester_id"),
        )
    )
    if initial_module:
        context["initial_module"] = initial_module
    if initial_workspace_html is not None:
        context["initial_workspace_html"] = mark_safe(initial_workspace_html)
    return render(
        request,
        "portal/informaticien/dashboard.html",
        context,
    )


def _build_it_dashboard_context_for_request(request, overrides=None):
    original_get = request.GET
    if overrides:
        merged = request.GET.copy()
        for key, value in overrides.items():
            if value:
                merged[key] = value
        request.GET = merged
    try:
        return _build_it_dashboard_context(request)
    finally:
        request.GET = original_get


def _get_manageable_target_user(request):
    branch = _resolve_academic_branch(request)
    target_user = request.user.__class__.objects.filter(pk=request.POST.get("target_user_id")).first()
    if target_user is None or not can_manage_user_in_branch(branch=branch, target_user=target_user):
        return branch, None
    return branch, target_user


def _render_it_action_panel(request, *, default_panel="diagnostics"):
    panel = request.POST.get("panel") or default_panel
    context = _build_it_dashboard_context_for_request(
        request,
        overrides={
            "q": request.POST.get("q"),
            "kind": request.POST.get("kind"),
            "id": request.POST.get("id"),
        },
    )
    template_name = (
        "portal/informaticien/partials/accounts_panel.html"
        if panel == "accounts"
        else "portal/informaticien/partials/diagnostics_panel.html"
    )
    return render(request, template_name, context)


@login_required
def portal_home(request):
    return redirect(get_post_login_portal_url(request.user))


@login_required
def portal_dashboard(request):
    position = get_user_position(request.user)

    if position in {"finance_manager", "payment_agent"}:
        return redirect("accounts:finance_dashboard")
    if position == "secretary":
        return redirect("secretary:secretary_dashboard")
    if position == "admissions":
        return redirect("accounts:admissions_dashboard")
    if position == "super_admin" or request.user.is_superuser:
        return redirect("superadmin:dashboard")
    if position in {"executive_director", "deputy_executive_director"}:
        return dg_portal(request)
    if position == "director_of_studies":
        return _render_director_dashboard(request)
    if position == "academic_supervisor":
        from portal.views.supervisor import _render_supervisor_dashboard

        return _render_supervisor_dashboard(request)
    if position == "it_support":
        return _render_it_dashboard(request)
    if position == "marketing_manager":
        return redirect("marketing:dashboard")

    if can_access(request.user, "view_portal", "student"):
        return redirect("portal_student:dashboard")
    if can_access(request.user, "view_portal", "teacher"):
        return teacher_portal(request)
    if can_access(request.user, "view_portal", "staff"):
        return staff_portal(request)
    if request.user.is_superuser:
        return redirect("admin:index")
    return _deny_portal_access(request)


@login_required
def student_portal(request):
    if not can_access(request.user, "view_portal", "student"):
        return _deny_portal_access(request)
    return redirect("accounts_portal:portal_dashboard")


@login_required
def staff_portal(request):
    if not can_access(request.user, "view_portal", "staff"):
        return _deny_portal_access(request)

    context = _build_portal_context(
        request,
        page_title="Portail staff",
        module_cards=[
            "Gestion etudiants",
            "Documents",
            "Admissions",
            "Finance",
            "Secretariat",
            "Supervision academique",
        ],
    )
    return render(request, "portal/staff.html", context)


TEACHER_PORTAL_SECTIONS = (
    "overview",
    "classes",
    "supports",
    "schedule",
    "logs",
    "salary",
    "notifications",
    "settings",
)

TEACHER_SECTION_CONTEXT_BUILDERS = {
    "overview": build_teacher_overview_context,
    "classes": build_teacher_classes_context,
    "supports": build_teacher_supports_context,
    "schedule": build_teacher_schedule_context,
    "logs": build_teacher_logs_context,
    "salary": build_teacher_salary_context,
    "notifications": build_teacher_notifications_context,
    "settings": build_teacher_settings_context,
}


def _teacher_nav_items(active_section):
    definitions = (
        ("overview", "Accueil", "layout-dashboard"),
        ("classes", "Mes classes", "graduation-cap"),
        ("supports", "Supports", "folder-open"),
        ("schedule", "Planning", "calendar-days"),
        ("logs", "Cahier texte", "book-open"),
        ("salary", "Honoraires", "wallet"),
        ("notifications", "Notifications", "bell"),
        ("settings", "Paramètres", "settings"),
    )
    return [
        {
            "id": section,
            "label": label,
            "icon": icon,
            "url": f"?section={section}",
            "hx_get": f"?section={section}",
            "hx_target": "#teacher-workspace",
            "hx_swap": "outerHTML",
            "active": active_section == section,
        }
        for section, label, icon in definitions
    ]


@login_required
def teacher_portal(request):
    if not can_access(request.user, "view_portal", "teacher"):
        return _deny_portal_access(request)

    branch = _resolve_academic_branch(request)
    section = request.GET.get("section", "overview")
    active_section = section if section in TEACHER_PORTAL_SECTIONS else "overview"

    if section == "search":
        q = (request.GET.get("q") or "").strip()
        if branch and len(q) >= 1:
            from students.models import Student

            classes = AcademicClass.objects.filter(
                branch=branch,
            ).filter(
                Q(name__icontains=q) | Q(level__icontains=q) | Q(programme__title__icontains=q)
            )[:5]

            ecs = EC.objects.filter(
                ue__semester__academic_class__branch=branch,
                teachingassignment__teacher__user=request.user,
            ).filter(
                Q(title__icontains=q) | Q(ue__code__icontains=q)
            ).distinct()[:5]

            students = Student.objects.filter(
                inscription__candidature__branch=branch,
                is_active=True,
            ).filter(
                Q(matricule__icontains=q)
                | Q(inscription__candidature__first_name__icontains=q)
                | Q(inscription__candidature__last_name__icontains=q)
            ).distinct()[:5]

            return render(request, "portal/partials/teacher_search_results.html", {
                "q": q,
                "classes": classes,
                "ecs": ecs,
                "students": students,
            })
        return render(request, "portal/partials/teacher_search_results.html", {"q": q})

    if branch is None:
        context = {
            **_build_portal_context(
                request,
                page_title="Dashboard enseignant",
                module_cards=[],
            ),
            "dashboard_kind": "Enseignant",
            "branch": None,
            "branch_missing": True,
            "active_section": active_section,
            "teacher_nav_items": _teacher_nav_items(active_section),
            "status_summary": {"branch_name": "Annexe non définie"},
            "pending_lesson_logs_count": 0,
            "notifications_count": 0,
            "kpi_cards": [],
        }
        return render(request, "portal/teacher/v2/dashboard.html", context)

    context_builder = TEACHER_SECTION_CONTEXT_BUILDERS[active_section]
    context = context_builder(
        request,
        branch=branch,
        base_context_builder=_build_portal_context,
    )
    context["active_section"] = active_section
    context["teacher_nav_items"] = _teacher_nav_items(active_section)
    context.setdefault("pending_lesson_logs_count", 0)
    context["notifications_count"] = get_user_unread_count(
        request.user,
        exclude_sources=TEACHER_EXCLUDED_NOTIFICATION_SOURCES,
    )
    context.setdefault("teacher_notifications", {"unread_count": 0, "items": []})
    context["kpi_cards"] = [
        {"label": "Mes classes", "value": len(context.get("class_focus_rows") or []), "icon": "graduation-cap"},
        {"label": "Cours aujourd'hui", "value": context.get("teacher_kpis", {}).get("today_courses", 0), "icon": "calendar-check-2"},
        {"label": "Cahiers en attente", "value": context.get("pending_lesson_logs_count", 0), "icon": "book-open"},
        {"label": "Cahiers du mois", "value": context.get("teacher_kpis", {}).get("month_done_logs", 0), "icon": "clipboard-check"},
    ]
    teacher_kpis = context.get("teacher_kpis", {})
    support_stats = context.get("teacher_support_stats", {})
    section_kpis = {
        "classes": [
            {"label": "Classes", "value": len(context.get("class_focus_rows") or []), "icon": "graduation-cap", "tone": "primary"},
            {"label": "Matières", "value": teacher_kpis.get("subjects_count", 0), "icon": "book-open", "tone": "info"},
            {"label": "Étudiants", "value": teacher_kpis.get("visible_students", 0), "icon": "users", "tone": "success"},
        ],
        "supports": [
            {"label": "Classes", "value": len(context.get("class_focus_rows") or []), "icon": "graduation-cap", "tone": "primary"},
            {"label": "Chapitres", "value": support_stats.get("chapters", 0), "icon": "book-open", "tone": "info"},
            {"label": "Supports", "value": support_stats.get("contents", 0), "icon": "folder-open", "tone": "success"},
            {"label": "Fichiers", "value": support_stats.get("files", 0), "icon": "file", "tone": "neutral"},
        ],
        "schedule": [
            {"label": "Aujourd'hui", "value": teacher_kpis.get("today_courses", 0), "icon": "calendar-check", "tone": "primary"},
            {"label": "Cette semaine", "value": teacher_kpis.get("week_courses", 0), "icon": "calendar-days", "tone": "success"},
            {"label": "Classes", "value": len(context.get("class_focus_rows") or []), "icon": "graduation-cap", "tone": "info"},
            {"label": "Salles", "value": teacher_kpis.get("weekly_rooms", 0), "icon": "door-open", "tone": "warning"},
        ],
        "logs": [
            {"label": "Récents", "value": len(context.get("recent_lesson_logs") or []), "icon": "book-open", "tone": "success"},
            {"label": "Cette semaine", "value": context.get("week_lesson_logs_count", 0), "icon": "calendar-days", "tone": "primary"},
            {"label": "En attente", "value": context.get("pending_lesson_logs_count", 0), "icon": "alert-triangle", "tone": "warning"},
            {"label": "Ce mois", "value": teacher_kpis.get("month_done_logs", 0), "icon": "clipboard-check", "tone": "info"},
        ],
        "salary": [
            {"label": "Heures ce mois", "value": context.get("teacher_hours", {}).get("month_hours", 0), "icon": "clock", "tone": "primary"},
            {"label": "Total année", "value": context.get("teacher_hours", {}).get("total_hours", 0), "icon": "calendar", "tone": "success"},
            {"label": "En attente", "value": context.get("teacher_payments", {}).get("pending_count", 0), "icon": "hourglass", "tone": "warning"},
            {"label": "Payés", "value": context.get("teacher_payments", {}).get("paid_count", 0), "icon": "circle-check", "tone": "success"},
        ],
    }
    context["section_kpi_cards"] = section_kpis.get(active_section, [])
    if request.headers.get("HX-Request") == "true":
        return render(request, "portal/teacher/v2/workspace.html", context)
    return render(request, "portal/teacher/v2/dashboard.html", context)


@_position_required({"teacher"})
def teacher_class_detail(request, class_id: int):
    branch = _resolve_academic_branch(request)
    try:
        raw_week_start = (request.GET.get("week_start") or "").strip()
        week_start = timezone.localdate()
        if raw_week_start:
            week_start = datetime.strptime(raw_week_start, "%Y-%m-%d").date()
        context = build_teacher_class_detail_context(
            request,
            branch=branch,
            class_id=class_id,
            week_start=week_start,
        )
    except (ValidationError, ValueError) as exc:
        error_message = exc.messages[0] if isinstance(exc, ValidationError) and exc.messages else str(exc)
        context = {
            "toast": {"level": "error", "message": error_message},
            "academic_class": None,
        }
    return render(request, "portal/partials/teacher_class_detail.html", context)


@_position_required({"teacher"})
def teacher_support_workspace(request):
    branch = _resolve_academic_branch(request)

    def _parse_id(raw_value):
        raw_value = (raw_value or "").strip()
        if not raw_value:
            return None
        return int(raw_value)

    class_id = request.POST.get("class_id") if request.method == "POST" else request.GET.get("class_id")
    ec_id = request.POST.get("ec_id") if request.method == "POST" else request.GET.get("ec_id")
    chapter_id = request.POST.get("chapter_id") if request.method == "POST" else request.GET.get("chapter_id")
    support_target = (
        request.POST.get("support_target")
        or request.GET.get("support_target")
        or "#sg-drawer-content"
    ).strip()
    if not support_target.startswith("#"):
        support_target = "#sg-drawer-content"

    try:
        parsed_class_id = _parse_id(class_id)
        parsed_ec_id = _parse_id(ec_id)
        parsed_chapter_id = _parse_id(chapter_id)

        if request.method == "POST":
            context = build_teacher_support_workspace_context(
                request,
                branch=branch,
                class_id=parsed_class_id,
                ec_id=parsed_ec_id,
                chapter_id=parsed_chapter_id,
            )
            context["teacher_support_target"] = support_target
            selected_class = context["selected_class"]
            selected_ec = context["selected_ec"]
            selected_chapter = context["selected_chapter"]
            action = (request.POST.get("action") or "").strip()

            if action in ("edit_content", "delete_content", "fetch_content"):
                content_id = _parse_id(request.POST.get("content_id"))
                if content_id is None:
                    raise ValidationError("Identifiant du contenu manquant.")

                if action == "fetch_content":
                    editing_content = get_teacher_content_for_edit(
                        teacher=request.user,
                        branch=branch,
                        content_id=content_id,
                    )
                    context = build_teacher_support_workspace_context(
                        request, branch=branch,
                        class_id=editing_content["class_id"],
                        ec_id=editing_content.get("ec_id"),
                        chapter_id=editing_content.get("chapter_id"),
                    )
                    context["teacher_support_target"] = support_target
                    context["editing_content"] = editing_content
                    return render(request, "portal/partials/teacher_support_workspace.html", context)

                elif action == "edit_content":
                    edit_content_id = content_id
                    chapter_id_edit = _parse_id(request.POST.get("chapter_id"))
                    updated = update_teacher_content(
                        teacher=request.user,
                        branch=branch,
                        content_id=edit_content_id,
                        title=request.POST.get("content_title"),
                        content_type=request.POST.get("content_type"),
                        chapter_id=chapter_id_edit,
                        file=request.FILES.get("file"),
                        video_url=request.POST.get("video_url"),
                        text_content=request.POST.get("text_content"),
                    )
                    toast_message = f"Support '{updated.title}' modifie avec succes."
                    context = build_teacher_support_workspace_context(
                        request, branch=branch,
                        class_id=parsed_class_id,
                        ec_id=parsed_ec_id,
                        chapter_id=updated.chapter_id,
                        toast={"level": "success", "message": toast_message},
                    )
                    context["teacher_support_target"] = support_target
                    return render(request, "portal/partials/teacher_support_workspace.html", context)

                elif action == "delete_content":
                    deleted = delete_teacher_content(
                        teacher=request.user,
                        branch=branch,
                        content_id=content_id,
                    )
                    toast_message = f"Support '{deleted.title}' supprime avec succes."
                    context = build_teacher_support_workspace_context(
                        request, branch=branch,
                        class_id=parsed_class_id,
                        ec_id=parsed_ec_id,
                        chapter_id=parsed_chapter_id,
                        toast={"level": "success", "message": toast_message},
                    )
                    context["teacher_support_target"] = support_target
                    return render(request, "portal/partials/teacher_support_workspace.html", context)

            if selected_class is None:
                raise ValidationError("Aucune classe disponible pour ce compte.")
            if selected_ec is None:
                raise ValidationError("Selectionnez d'abord une matiere.")

            toast_message = "Support enregistre avec succes."
            if action == "create_chapter":
                chapter_title = (request.POST.get("chapter_title") or "").strip()
                if not chapter_title:
                    raise ValidationError("Le titre du chapitre est obligatoire.")
                chapter_order = selected_ec.chapters.count() + 1
                created_chapter = ECChapter.objects.create(
                    ec=selected_ec,
                    title=chapter_title,
                    order=chapter_order,
                )
                parsed_chapter_id = created_chapter.id
                toast_message = f"Chapitre '{chapter_title}' cree avec succes."
            elif action == "create_content":
                new_chapter_title = (request.POST.get("new_chapter_title") or "").strip()
                chapter_selection_blank = not (request.POST.get("chapter_id") or "").strip()
                if new_chapter_title:
                    selected_chapter = ECChapter.objects.create(
                        ec=selected_ec,
                        title=new_chapter_title,
                        order=selected_ec.chapters.count() + 1,
                    )
                    parsed_chapter_id = selected_chapter.id
                elif chapter_selection_blank:
                    selected_chapter = selected_ec.chapters.filter(title="Supports de cours").first()
                    if selected_chapter is None:
                        selected_chapter = ECChapter.objects.create(
                            ec=selected_ec,
                            title="Supports de cours",
                            order=selected_ec.chapters.count() + 1,
                        )
                    parsed_chapter_id = selected_chapter.id
                elif selected_chapter is None:
                    selected_chapter = ECChapter.objects.create(
                        ec=selected_ec,
                        title="Supports de cours",
                        order=selected_ec.chapters.count() + 1,
                    )
                    parsed_chapter_id = selected_chapter.id
                content_title = (request.POST.get("content_title") or "").strip()
                if not content_title:
                    raise ValidationError("Le titre du support est obligatoire.")
                content_type = (request.POST.get("content_type") or "").strip()
                allowed_types = {choice[0] for choice in ECContent.CONTENT_TYPE_CHOICES}
                if content_type not in allowed_types:
                    raise ValidationError("Type de contenu invalide.")

                uploaded_file = request.FILES.get("file")
                video_url = (request.POST.get("video_url") or "").strip() or None
                text_content = (request.POST.get("text_content") or "").strip()

                if content_type == ECContent.CONTENT_TYPE_TEXT:
                    if uploaded_file is not None:
                        raise ValidationError("Le type Texte n'accepte pas de fichier.")
                    if video_url:
                        raise ValidationError("Le type Texte n'accepte pas d'URL video.")
                    if not text_content:
                        raise ValidationError("Le texte du support est obligatoire pour ce type de contenu.")
                elif content_type == ECContent.CONTENT_TYPE_VIDEO:
                    if text_content:
                        raise ValidationError("Le type Video n'accepte pas de texte direct.")
                    if not uploaded_file and not video_url:
                        raise ValidationError("Ajoutez un fichier video ou une URL YouTube.")
                else:
                    if video_url:
                        raise ValidationError("L'URL video est reservee au type Video.")
                    if text_content:
                        raise ValidationError("Le texte direct est reserve au type Texte.")
                    if uploaded_file is None:
                        raise ValidationError("Ajoutez le fichier du support.")

                _validate_content_file_extension(content_type=content_type, uploaded_file=uploaded_file)

                content_order = selected_chapter.contents.count() + 1
                content = ECContent(
                    chapter=selected_chapter,
                    title=content_title,
                    content_type=content_type,
                    file=uploaded_file,
                    video_url=video_url,
                    text_content=text_content,
                    order=content_order,
                )
                content.full_clean()
                content.save()
                toast_message = f"Support '{content_title}' ajoute avec succes."
            else:
                raise ValidationError("Action support inconnue.")

            context = build_teacher_support_workspace_context(
                request,
                branch=branch,
                class_id=selected_class.id,
                ec_id=selected_ec.id,
                chapter_id=parsed_chapter_id or getattr(selected_chapter, "id", None),
                toast={"level": "success", "message": toast_message},
            )
            context["teacher_support_target"] = support_target
            return render(request, "portal/partials/teacher_support_workspace.html", context)

        context = build_teacher_support_workspace_context(
            request,
            branch=branch,
            class_id=parsed_class_id,
            ec_id=parsed_ec_id,
            chapter_id=parsed_chapter_id,
        )
        context["teacher_support_target"] = support_target
    except (ValidationError, ValueError, TypeError) as exc:
        error_message = exc.messages[0] if isinstance(exc, ValidationError) and exc.messages else str(exc)
        try:
            context = build_teacher_support_workspace_context(
                request,
                branch=branch,
                class_id=None,
                ec_id=None,
                chapter_id=None,
                toast={"level": "error", "message": error_message},
            )
            context["teacher_support_target"] = support_target
        except ValidationError:
            context = {
                "branch": branch,
                "toast": {"level": "error", "message": error_message},
                "teacher_support_target": support_target,
                "teacher_support_classes": [],
                "selected_class": None,
                "selected_ec": None,
                "selected_ec_summary": None,
                "selected_chapter": None,
                "teacher_support_ecs": [],
                "teacher_support_chapters": [],
                "teacher_support_contents": [],
                "content_type_choices": ECContent.CONTENT_TYPE_CHOICES,
                "file_content_type_choices": [
                    choice for choice in ECContent.CONTENT_TYPE_CHOICES
                    if choice[0] != ECContent.CONTENT_TYPE_TEXT
                ],
            }
    return render(request, "portal/partials/teacher_support_workspace.html", context)


@_position_required({"teacher"})
def teacher_settings_workspace(request):
    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(
            request,
            "portal/teacher/partials/settings_workspace.html",
            {
                "toast": {"level": "error", "message": "Aucune annexe rattachee a ce compte enseignant."},
                "teacher_dashboard_preference": {
                    "dark_mode": False,
                    "sidebar_collapsed": False,
                    "compact_mode": False,
                    "default_section": "overview",
                    "notify_lesson_reminders": True,
                    "notify_schedule_changes": True,
                    "notify_support_messages": True,
                },
                "teacher_preferences_choices": [],
                "status_summary": {
                    "employment_status": "Inconnu",
                    "employee_code": "Non renseigne",
                    "branch_name": "Non rattache",
                    "hire_date": None,
                },
                "teacher_settings_stats": {"active_classes": 0, "display_mode": "Standard"},
            },
        )

    toast = None
    if request.method == "POST":
        try:
            update_teacher_dashboard_preference(
                actor=request.user,
                teacher=request.user,
                branch=branch,
                dark_mode=request.POST.get("dark_mode") == "on",
                sidebar_collapsed=request.POST.get("sidebar_collapsed") == "on",
                compact_mode=request.POST.get("compact_mode") == "on",
                default_section=request.POST.get("default_section"),
                notify_lesson_reminders=request.POST.get("notify_lesson_reminders") == "on",
                notify_schedule_changes=request.POST.get("notify_schedule_changes") == "on",
                notify_support_messages=request.POST.get("notify_support_messages") == "on",
            )
            toast = {"level": "success", "message": "Parametres enregistres."}
        except ValidationError as exc:
            toast = {"level": "error", "message": exc.messages[0] if getattr(exc, "messages", None) else str(exc)}

    context = build_teacher_settings_context(
        request,
        branch=branch,
        base_context_builder=_build_portal_context,
    )
    context["toast"] = toast
    return render(request, "portal/teacher/partials/settings_workspace.html", context)


@_position_required({"teacher"})
def teacher_content_viewer(request, content_id: int):
    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(
            request,
            "portal/partials/teacher_content_viewer.html",
            {"content": None, "error": "Aucune annexe rattachée à ce compte enseignant."},
        )
    try:
        get_teacher_content_for_edit(
            teacher=request.user,
            branch=branch,
            content_id=content_id,
        )
        content = ECContent.objects.select_related("chapter__ec").get(
            pk=content_id,
            is_active=True,
        )
        if content.chapter is None:
            raise ValidationError("Ce contenu n'est pas rattache a un chapitre.")
        ec = content.chapter.ec
        branch_match = ec.ue.semester.academic_class.branch_id == branch.id
        if not branch_match:
            raise ValidationError("Contenu non accessible pour cette annexe.")

        if content.content_type == ECContent.CONTENT_TYPE_TEXT:
            rendered_text = content.text_content
        else:
            rendered_text = ""

        context = {
            "content": content,
            "rendered_text": rendered_text,
            "branch": branch,
        }
    except (ValidationError, ECContent.DoesNotExist) as exc:
        error_message = exc.messages[0] if isinstance(exc, ValidationError) and exc.messages else str(exc)
        context = {
            "content": None,
            "error": error_message,
        }
    return render(request, "portal/partials/teacher_content_viewer.html", context)


@_position_required({"teacher"})
def teacher_lesson_log_panel(request, event_id: int):
    branch = _resolve_academic_branch(request)

    if request.method == "POST":
        allowed_statuses = {
            LessonLog.STATUS_DONE,
            LessonLog.STATUS_CANCELLED,
            LessonLog.STATUS_PLANNED,
        }
        try:
            context = build_teacher_lesson_log_context(
                request,
                branch=branch,
                event_id=event_id,
            )
            schedule_event = context["schedule_event"]
            status = (request.POST.get("status") or "").strip()
            if status not in allowed_statuses:
                raise ValidationError("Statut de cahier invalide.")
            payload = {
                "status": status,
                "content": request.POST.get("content", ""),
                "homework": request.POST.get("homework", ""),
                "observations": request.POST.get("observations", ""),
            }
            lesson_log = context["lesson_log"]
            if lesson_log is None:
                create_lesson_log(
                    academic_class=schedule_event.academic_class,
                    ec=schedule_event.ec,
                    teacher=schedule_event.teacher,
                    date=timezone.localdate(schedule_event.start_datetime),
                    start_time=timezone.localtime(schedule_event.start_datetime).time(),
                    end_time=timezone.localtime(schedule_event.end_datetime).time(),
                    branch=schedule_event.branch,
                    created_by=request.user,
                    schedule_event=schedule_event,
                    **payload,
                )
            else:
                update_lesson_log(
                    lesson_log,
                    updated_by=request.user,
                    **payload,
                )

            mark_present = request.POST.get("mark_present") == "1"
            if mark_present:
                TeacherAttendance.objects.update_or_create(
                    teacher=request.user,
                    schedule_event=schedule_event,
                    branch=schedule_event.branch,
                    defaults={
                        "date": timezone.localdate(schedule_event.start_datetime),
                        "status": TeacherAttendance.STATUS_PRESENT,
                        "recorded_by": request.user,
                    },
                )
            else:
                TeacherAttendance.objects.filter(
                    teacher=request.user,
                    schedule_event=schedule_event,
                ).delete()

            context = build_teacher_lesson_log_context(
                request,
                branch=branch,
                event_id=event_id,
                toast={"level": "success", "message": "Cahier enregistre avec succes."},
            )
        except ValidationError as exc:
            error_message = exc.messages[0] if exc.messages else str(exc)
            try:
                context = build_teacher_lesson_log_context(
                    request,
                    branch=branch,
                    event_id=event_id,
                    toast={"level": "error", "message": error_message},
                )
            except ValidationError:
                context = {
                    "toast": {"level": "error", "message": error_message},
                    "schedule_event": None,
                    "lesson_log": None,
                    "lesson_log_status_choices": [],
                }
            return render(request, "portal/partials/teacher_lesson_log_panel.html", context)

        return render(request, "portal/partials/teacher_lesson_log_panel.html", context)

    try:
        context = build_teacher_lesson_log_context(
            request,
            branch=branch,
            event_id=event_id,
        )
    except ValidationError as exc:
        error_message = exc.messages[0] if exc.messages else str(exc)
        context = {
            "toast": {"level": "error", "message": error_message},
            "schedule_event": None,
            "lesson_log": None,
            "lesson_log_status_choices": [],
        }
    return render(request, "portal/partials/teacher_lesson_log_panel.html", context)


@_position_required({"finance_manager", "payment_agent"})
def finance_portal(request):
    return redirect("accounts_portal:portal_dashboard")


@_position_required({"secretary"})
def secretary_portal(request):
    return redirect("secretary:secretary_dashboard")


@_position_required({"admissions"})
def admissions_portal(request):
    return redirect("accounts_portal:portal_dashboard")


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_portal(request):
    return redirect("accounts_portal:portal_dashboard")


@_position_required({"it_support"})
def it_portal(request):
    return redirect("accounts_portal:portal_dashboard")


@_position_required({"it_support"})
def it_toggle_account(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch = _resolve_academic_branch(request)
    target_user = get_scoped_staff_queryset(branch=branch).filter(pk=request.POST.get("target_user_id")).first()
    if target_user is None:
        target_user = request.user.__class__.objects.filter(pk=request.POST.get("target_user_id")).first()

    if target_user is None or not can_manage_user_in_branch(branch=branch, target_user=target_user):
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)

    if target_user == request.user:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte informaticien courant ne peut pas etre desactive depuis ce dashboard.",
        )
        return _redirect_it_dashboard(request)

    target_user.is_active = not target_user.is_active
    target_user.save(update_fields=["is_active"])
    action_type = (
        "account_activated"
        if target_user.is_active
        else "account_deactivated"
    )
    log_support_action(
        actor=request.user,
        branch=branch,
        action_type=action_type,
        target_user=target_user,
        target_label=target_user.get_full_name() or target_user.username,
        details=f"Activation modifiee via dashboard IT ({branch.name if branch else 'global'}).",
    )
    _store_it_support_feedback(
        request,
        level="success",
        title="Compte mis a jour",
        message=f"Le compte {target_user.get_full_name() or target_user.username} est maintenant {'actif' if target_user.is_active else 'inactif'}.",
    )
    if request.headers.get("HX-Request") == "true":
        panel = request.POST.get("panel") or "diagnostics"
        context = _build_it_dashboard_context_for_request(
            request,
            overrides={
                "q": request.POST.get("q"),
                "kind": request.POST.get("kind"),
                "id": request.POST.get("id"),
            },
        )
        template_name = (
            "portal/informaticien/partials/accounts_panel.html"
            if panel == "accounts"
            else "portal/informaticien/partials/diagnostics_panel.html"
        )
        return render(request, template_name, context)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_reset_password(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch = _resolve_academic_branch(request)
    target_user = request.user.__class__.objects.filter(pk=request.POST.get("target_user_id")).first()
    if target_user is None or not can_manage_user_in_branch(branch=branch, target_user=target_user):
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)

    temp_password = create_temp_password()
    target_user.set_password(temp_password)
    target_user.save(update_fields=["password"])
    log_support_action(
        actor=request.user,
        branch=branch,
        action_type="password_reset",
        target_user=target_user,
        target_label=target_user.get_full_name() or target_user.username,
        details=f"Mot de passe reinitialise via dashboard IT ({branch.name if branch else 'global'}).",
    )
    _store_it_support_feedback(
        request,
        level="success",
        title="Mot de passe reinitialise",
        message=f"Un mot de passe temporaire a ete genere pour {target_user.get_full_name() or target_user.username}.",
        password=temp_password,
    )
    if request.headers.get("HX-Request") == "true":
        panel = request.POST.get("panel") or "diagnostics"
        context = _build_it_dashboard_context_for_request(
            request,
            overrides={
                "q": request.POST.get("q"),
                "kind": request.POST.get("kind"),
                "id": request.POST.get("id"),
            },
        )
        template_name = (
            "portal/informaticien/partials/accounts_panel.html"
            if panel == "accounts"
            else "portal/informaticien/partials/diagnostics_panel.html"
        )
        return render(request, template_name, context)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_suspend_account(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch, target_user = _get_manageable_target_user(request)
    if target_user is None:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)
    if target_user == request.user:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte informaticien courant ne peut pas etre suspendu depuis ce dashboard.",
        )
        return _redirect_it_dashboard(request)

    suspend_account(
        actor=request.user,
        branch=branch,
        target_user=target_user,
        reason=request.POST.get("reason"),
    )
    _store_it_support_feedback(
        request,
        level="success",
        title="Compte suspendu",
        message=f"Le compte {target_user.get_full_name() or target_user.username} est suspendu et ne peut plus se connecter.",
    )
    if request.headers.get("HX-Request") == "true":
        return _render_it_action_panel(request)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_reactivate_account(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch, target_user = _get_manageable_target_user(request)
    if target_user is None:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)

    reactivate_account(actor=request.user, branch=branch, target_user=target_user)
    _store_it_support_feedback(
        request,
        level="success",
        title="Compte reactive",
        message=f"Le compte {target_user.get_full_name() or target_user.username} est reactive.",
    )
    if request.headers.get("HX-Request") == "true":
        return _render_it_action_panel(request)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_unblock_account(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch, target_user = _get_manageable_target_user(request)
    if target_user is None:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)

    unblock_account(actor=request.user, branch=branch, target_user=target_user)
    _store_it_support_feedback(
        request,
        level="success",
        title="Compte debloque",
        message=f"Les blocages du compte {target_user.get_full_name() or target_user.username} ont ete leves.",
    )
    if request.headers.get("HX-Request") == "true":
        return _render_it_action_panel(request)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_update_account_email(request):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch, target_user = _get_manageable_target_user(request)
    if target_user is None:
        _store_it_support_feedback(
            request,
            level="error",
            title="Action refusee",
            message="Le compte cible est introuvable ou hors du perimetre de cette annexe.",
        )
        return _redirect_it_dashboard(request)
    email = (request.POST.get("email") or "").strip()
    try:
        validate_email(email)
    except ValidationError:
        _store_it_support_feedback(
            request,
            level="error",
            title="Email invalide",
            message="Adresse email non valide.",
        )
        if request.headers.get("HX-Request") == "true":
            return _render_it_action_panel(request)
        return _redirect_it_dashboard(request)

    update_account_email(actor=request.user, branch=branch, target_user=target_user, email=email)
    _store_it_support_feedback(
        request,
        level="success",
        title="Email corrige",
        message=f"Le nouvel email du compte est {target_user.email}.",
    )
    if request.headers.get("HX-Request") == "true":
        return _render_it_action_panel(request)
    return _redirect_it_dashboard(request)


@_position_required({"it_support"})
def it_diagnostics_panel(request):
    context = _build_it_dashboard_context(request)
    return render(request, "portal/informaticien/partials/diagnostics_panel.html", context)


@_position_required({"it_support"})
def it_accounts_panel(request):
    context = _build_it_dashboard_context(request)
    return render(request, "portal/informaticien/partials/accounts_panel.html", context)


def _build_it_support_panel_context(request):
    branch = _resolve_academic_branch(request)
    status = (request.GET.get("status") or "").strip()
    context = _build_it_dashboard_context(request)
    context.update(
        {
            "ticket_status_filter": status,
            "ticket_status_choices": SupportTicket.STATUS_CHOICES,
            "ticket_category_choices": SupportTicket.CATEGORY_CHOICES,
            "ticket_priority_choices": SupportTicket.PRIORITY_CHOICES,
            "support_ticket_metrics": get_support_ticket_metrics(branch=branch),
            "support_tickets": list(get_support_ticket_queryset(branch=branch, status=status)[:12]),
        }
    )
    return context


@_position_required({"it_support"})
def it_support_panel(request):
    return render(
        request,
        "portal/informaticien/partials/support_panel.html",
        _build_it_support_panel_context(request),
    )


@_position_required({"it_support"})
def it_create_ticket(request):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    requester_user = None
    student = None
    inscription = None
    kind = (request.POST.get("kind") or "").strip()
    object_id = (request.POST.get("object_id") or "").strip()
    if object_id.isdigit():
        diagnostic = build_diagnostic_payload(branch=branch, kind=kind, object_id=int(object_id))
        if diagnostic:
            if diagnostic.get("target_user_id"):
                requester_user = request.user.__class__.objects.filter(pk=diagnostic["target_user_id"]).first()
            if kind == "student":
                student = Student.objects.filter(pk=int(object_id)).first()
                inscription = student.inscription if student else None
            elif kind == "inscription":
                inscription = Inscription.objects.filter(pk=int(object_id)).first()
                student = getattr(inscription, "student", None) if inscription else None
    try:
        create_support_ticket(
            actor=request.user,
            branch=branch,
            title=request.POST.get("title"),
            description=request.POST.get("description"),
            category=request.POST.get("category"),
            priority=request.POST.get("priority"),
            requester_user=requester_user,
            student=student,
            inscription=inscription,
        )
    except ValueError as exc:
        context = _build_it_support_panel_context(request)
        context["ticket_feedback"] = {"level": "error", "message": str(exc)}
        return render(request, "portal/informaticien/partials/support_panel.html", context)
    context = _build_it_support_panel_context(request)
    context["ticket_feedback"] = {"level": "success", "message": "Ticket cree et ajoute a la file support."}
    return render(request, "portal/informaticien/partials/support_panel.html", context)


def _get_scoped_ticket_or_none(*, request, ticket_id):
    branch = _resolve_academic_branch(request)
    ticket = get_support_ticket_queryset(branch=branch).filter(pk=ticket_id).first()
    return branch, ticket


@_position_required({"it_support"})
def it_assign_ticket(request, ticket_id):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch, ticket = _get_scoped_ticket_or_none(request=request, ticket_id=ticket_id)
    if not ticket:
        return _deny_portal_access(request)
    assign_support_ticket(actor=request.user, branch=branch, ticket=ticket)
    return render(request, "portal/informaticien/partials/support_panel.html", _build_it_support_panel_context(request))


@_position_required({"it_support"})
def it_update_ticket_status(request, ticket_id):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch, ticket = _get_scoped_ticket_or_none(request=request, ticket_id=ticket_id)
    if not ticket:
        return _deny_portal_access(request)
    try:
        update_support_ticket_status(
            actor=request.user,
            branch=branch,
            ticket=ticket,
            status=request.POST.get("status"),
            resolution=request.POST.get("resolution"),
        )
    except ValueError as exc:
        context = _build_it_support_panel_context(request)
        context["ticket_feedback"] = {"level": "error", "message": str(exc)}
        return render(request, "portal/informaticien/partials/support_panel.html", context)
    return render(request, "portal/informaticien/partials/support_panel.html", _build_it_support_panel_context(request))


@_position_required({"it_support"})
def it_comment_ticket(request, ticket_id):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch, ticket = _get_scoped_ticket_or_none(request=request, ticket_id=ticket_id)
    if not ticket:
        return _deny_portal_access(request)
    try:
        add_support_ticket_comment(
            actor=request.user,
            branch=branch,
            ticket=ticket,
            body=request.POST.get("body"),
        )
    except ValueError as exc:
        context = _build_it_support_panel_context(request)
        context["ticket_feedback"] = {"level": "error", "message": str(exc)}
        return render(request, "portal/informaticien/partials/support_panel.html", context)
    return render(request, "portal/informaticien/partials/support_panel.html", _build_it_support_panel_context(request))


def _parse_slot_time_hhmm(raw: str):
    value = (raw or "").strip()
    if not value:
        raise ValidationError("Heure obligatoire.")
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValidationError("Format d'heure attendu : HH:MM.")


def _build_weekly_slots_workspace_context(
    request,
    *,
    branch,
    class_id: int,
    week_start,
    editing_slot_id=None,
    toast=None,
):
    ctx = build_class_detail_context(
        request,
        branch=branch,
        class_id=class_id,
        week_start=week_start,
    )
    academic_class = ctx["academic_class"]
    slots = list_weekly_slots_for_class(academic_class)
    ctx["weekly_slot_rows"] = [serialize_weekly_slot_for_ui(s) for s in slots]
    ctx["weekday_choices"] = WeeklyScheduleSlot.WEEKDAY_CHOICES
    ctx["editing_slot"] = None
    if editing_slot_id:
        slot = (
            WeeklyScheduleSlot.objects.select_related("ec", "teacher")
            .filter(
                pk=editing_slot_id,
                academic_class=academic_class,
                branch=branch,
                is_active=True,
            )
            .first()
        )
        if slot:
            ctx["editing_slot"] = serialize_weekly_slot_for_ui(slot)
    if toast is not None:
        ctx["toast"] = toast
    return ctx


def _inject_planner_route_context(context, *, role_prefix: str, workspace_target_id: str):
    context.update(
        {
            "workspace_target_id": workspace_target_id,
            "planner_hub_url": f"accounts_portal:{role_prefix}_planner_hub",
            "planner_view_url": f"accounts_portal:{role_prefix}_planner_view_workspace",
            "planner_workspace_url": f"accounts_portal:{role_prefix}_planner_workspace",
            "weekly_slots_workspace_url": f"accounts_portal:{role_prefix}_weekly_slots_workspace",
            "weekly_slot_save_url": f"accounts_portal:{role_prefix}_weekly_slot_save",
            "week_materialize_url": f"accounts_portal:{role_prefix}_week_materialize",
            "month_materialize_url": f"accounts_portal:{role_prefix}_month_materialize",
            "create_schedule_event_url": f"accounts_portal:{role_prefix}_create_schedule_event",
            "class_print_url": "accounts_portal:schedule_class_print",
            "teacher_print_url": "accounts_portal:schedule_teacher_print",
        }
    )
    return context


def _materialize_period_from_weekly_slots(*, user, academic_class, week_start, weeks_count: int):
    from academics.services.schedule_service import materialize_week_events_from_weekly_slots

    created = 0
    skipped = 0
    normalized_week = week_start
    for offset in range(max(1, weeks_count)):
        cursor = normalized_week + timedelta(days=7 * offset)
        result = materialize_week_events_from_weekly_slots(
            user=user,
            academic_class=academic_class,
            week_start=cursor,
        )
        created += result["created"]
        skipped += result["skipped_existing"]
    return {"created": created, "skipped_existing": skipped, "weeks_count": max(1, weeks_count)}


def _parse_planning_period_request(request):
    week_start_raw = (request.GET.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    period = (request.GET.get("period") or "week").strip().lower()
    if period not in {"week", "month"}:
        period = "week"
    normalized_week = week_start - timedelta(days=week_start.weekday())
    weeks = 4 if period == "month" else 1
    return normalized_week, period, weeks


def _build_print_week_blocks(schedule_builder, target, week_start, weeks):
    blocks = []
    for offset in range(weeks):
        cursor = week_start + timedelta(days=7 * offset)
        blocks.append(schedule_builder(target, cursor))
    return blocks


def _parse_director_planner_request(request):
    branch = _resolve_academic_branch(request)
    class_id_raw = (request.GET.get("class_id") or request.POST.get("class_id") or "").strip()
    week_start_raw = (request.GET.get("week_start") or request.POST.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    class_id = int(class_id_raw) if class_id_raw.isdigit() else None
    return branch, class_id, week_start


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_planner_hub(request):
    branch, class_id, week_start = _parse_director_planner_request(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/planner_class_hub.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    if class_id is None:
        return render(request, "portal/staff/supervisor/partials/planner_class_hub.html", {"branch": branch, "academic_class": None})
    context = build_class_detail_context(request, branch=branch, class_id=class_id, week_start=week_start)
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/planner_class_hub.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_planner_view_workspace(request):
    branch, class_id, week_start = _parse_director_planner_request(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/planner_view_workspace.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    if class_id is None:
        return render(request, "portal/staff/supervisor/partials/planner_view_workspace.html", {"branch": branch, "academic_class": None})
    context = build_class_detail_context(request, branch=branch, class_id=class_id, week_start=week_start)
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/planner_view_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_planner_workspace(request):
    branch, class_id, week_start = _parse_director_planner_request(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/planner_workspace.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    if class_id is None:
        return render(request, "portal/staff/supervisor/partials/planner_workspace.html", {"branch": branch, "academic_class": None})
    context = build_class_detail_context(request, branch=branch, class_id=class_id, week_start=week_start)
    context["planner_intent"] = (request.GET.get("intent") or "create").strip() or "create"
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/planner_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_weekly_slots_workspace(request, class_id: int):
    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    week_start_raw = (request.GET.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    edit_raw = (request.GET.get("edit") or "").strip()
    editing_slot_id = int(edit_raw) if edit_raw.isdigit() else None
    try:
        context = _build_weekly_slots_workspace_context(request, branch=branch, class_id=class_id, week_start=week_start, editing_slot_id=editing_slot_id)
    except AcademicClass.DoesNotExist:
        return render(
            request,
            "portal/staff/supervisor/partials/weekly_slots_workspace.html",
            {"academic_class": None, "toast": {"level": "error", "message": "Classe introuvable pour cette annexe."}},
        )
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_weekly_slot_save(request, class_id: int):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(
            request,
            "portal/staff/supervisor/partials/weekly_slots_workspace.html",
            {"toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}, "academic_class": None},
        )

    week_start_raw = (request.POST.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()

    action = (request.POST.get("action") or "create").strip().lower()
    toast = None

    from django.contrib.auth import get_user_model

    User = get_user_model()

    try:
        academic_class = AcademicClass.objects.select_related("academic_year", "branch").get(pk=class_id, branch=branch, is_active=True)

        if action == "delete":
            slot_raw = (request.POST.get("slot_id") or "").strip()
            if not slot_raw.isdigit():
                raise ValidationError("Creneau invalide.")
            slot = WeeklyScheduleSlot.objects.get(pk=int(slot_raw), academic_class=academic_class, branch=branch)
            deactivate_weekly_schedule_slot(slot)
            toast = {"level": "success", "message": "Creneau retire de la grille hebdomadaire."}
        elif action in {"update", "create"}:
            ec = EC.objects.select_related("ue", "ue__semester").get(pk=request.POST.get("ec_id"), ue__semester__academic_class=academic_class)
            teacher = User.objects.get(pk=request.POST.get("teacher_id"), is_active=True)
            weekday_raw = (request.POST.get("weekday") or "").strip()
            if not weekday_raw.isdigit():
                raise ValidationError("Jour de semaine invalide.")
            weekday = int(weekday_raw)
            if weekday < 0 or weekday > 6:
                raise ValidationError("Jour de semaine hors plage (0-6).")
            start_t = _parse_slot_time_hhmm(request.POST.get("start_time"))
            end_t = _parse_slot_time_hhmm(request.POST.get("end_time"))
            if end_t <= start_t:
                raise ValidationError("L'heure de fin doit etre apres l'heure de debut.")
            room = (request.POST.get("room") or "").strip()

            if action == "update":
                slot_raw = (request.POST.get("slot_id") or "").strip()
                if not slot_raw.isdigit():
                    raise ValidationError("Creneau invalide.")
                slot = WeeklyScheduleSlot.objects.get(pk=int(slot_raw), academic_class=academic_class, branch=branch, is_active=True)
                update_weekly_schedule_slot(slot, weekday=weekday, ec=ec, teacher=teacher, start_time=start_t, end_time=end_t, room=room, is_active=True)
                toast = {"level": "success", "message": "Creneau hebdomadaire mis a jour."}
            else:
                create_weekly_schedule_slot(
                    user=request.user,
                    academic_class=academic_class,
                    ec=ec,
                    teacher=teacher,
                    branch=branch,
                    academic_year=academic_class.academic_year,
                    weekday=weekday,
                    start_time=start_t,
                    end_time=end_t,
                    room=room,
                    is_active=True,
                )
                toast = {"level": "success", "message": "Creneau hebdomadaire cree."}
        else:
            raise ValidationError("Action non reconnue.")
    except (AcademicClass.DoesNotExist, WeeklyScheduleSlot.DoesNotExist, EC.DoesNotExist, User.DoesNotExist):
        toast = {"level": "error", "message": "Donnee introuvable ou hors du perimetre de cette classe."}
    except ValidationError as exc:
        toast = {"level": "error", "message": " ".join(exc.messages)}

    context = _build_weekly_slots_workspace_context(
        request,
        branch=branch,
        class_id=class_id,
        week_start=week_start,
        editing_slot_id=None,
        toast=toast,
    )
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_week_materialize(request, class_id: int):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    week_start_raw = (request.POST.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    try:
        academic_class = AcademicClass.objects.select_related("academic_year", "branch").get(pk=class_id, branch=branch, is_active=True)
        result = _materialize_period_from_weekly_slots(user=request.user, academic_class=academic_class, week_start=week_start, weeks_count=1)
        toast = {"level": "success", "message": f"Semaine generee: {result['created']} cours crees ({result['skipped_existing']} deja presents)."}
    except AcademicClass.DoesNotExist:
        toast = {"level": "error", "message": "Classe introuvable pour cette annexe."}
    context = _build_weekly_slots_workspace_context(request, branch=branch, class_id=class_id, week_start=week_start, editing_slot_id=None, toast=toast)
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_month_materialize(request, class_id: int):
    if request.method != "POST":
        return _deny_portal_access(request)
    branch = _resolve_academic_branch(request)
    if branch is None:
        return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", {"academic_class": None, "toast": {"level": "error", "message": "Aucune annexe n'est rattachee a ce compte."}})
    week_start_raw = (request.POST.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    try:
        academic_class = AcademicClass.objects.select_related("academic_year", "branch").get(pk=class_id, branch=branch, is_active=True)
        result = _materialize_period_from_weekly_slots(user=request.user, academic_class=academic_class, week_start=week_start, weeks_count=4)
        toast = {"level": "success", "message": f"Mois pedagogique genere: {result['created']} cours crees ({result['skipped_existing']} deja presents)."}
    except AcademicClass.DoesNotExist:
        toast = {"level": "error", "message": "Classe introuvable pour cette annexe."}
    context = _build_weekly_slots_workspace_context(request, branch=branch, class_id=class_id, week_start=week_start, editing_slot_id=None, toast=toast)
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/weekly_slots_workspace.html", context)


@_position_required({"director_of_studies", "executive_director", "deputy_executive_director", "super_admin"})
def director_create_schedule_event(request, class_id: int):
    if request.method != "POST":
        return _deny_portal_access(request)

    branch = _resolve_academic_branch(request)
    if branch is None:
        return _deny_portal_access(request)

    from django.contrib.auth import get_user_model
    from django.core.exceptions import ValidationError

    from academics.models import AcademicClass, AcademicScheduleEvent, EC
    from academics.services.schedule_service import create_schedule_event

    User = get_user_model()
    academic_class = AcademicClass.objects.select_related("academic_year", "branch").get(pk=class_id, branch=branch, is_active=True)
    ec = EC.objects.select_related("ue", "ue__semester").get(pk=request.POST.get("ec_id"), ue__semester__academic_class=academic_class)
    teacher = User.objects.get(pk=request.POST.get("teacher_id"), is_active=True, profile__branch=branch, profile__position="teacher")
    date_value = (request.POST.get("date") or "").strip()
    start_time = (request.POST.get("start_time") or "").strip()
    end_time = (request.POST.get("end_time") or "").strip()
    location = (request.POST.get("location") or "").strip()
    is_online = bool((request.POST.get("is_online") or "").strip())

    try:
        if not (date_value and start_time and end_time):
            raise ValidationError("Date, heure debut et heure fin sont obligatoires.")
        start_dt = timezone.make_aware(datetime.strptime(f"{date_value} {start_time}", "%Y-%m-%d %H:%M"))
        end_dt = timezone.make_aware(datetime.strptime(f"{date_value} {end_time}", "%Y-%m-%d %H:%M"))
        if end_dt <= start_dt:
            raise ValidationError("L'heure de fin doit etre apres l'heure de debut.")

        create_schedule_event(
            user=request.user,
            event_type=AcademicScheduleEvent.EVENT_TYPE_COURSE,
            status=AcademicScheduleEvent.STATUS_PLANNED,
            academic_class=academic_class,
            academic_year=academic_class.academic_year,
            branch=branch,
            ec=ec,
            teacher=teacher,
            title=f"{ec.title} - {academic_class.display_name}",
            description="",
            start_datetime=start_dt,
            end_datetime=end_dt,
            location=location,
            is_online=is_online,
            is_active=True,
        )
        toast = {"level": "success", "message": "Cours programme sur la semaine."}
    except (ValidationError, EC.DoesNotExist, User.DoesNotExist) as exc:
        msg = " ".join(getattr(exc, "messages", [])) if hasattr(exc, "messages") else str(exc)
        toast = {"level": "error", "message": msg or "Impossible de programmer ce cours."}

    week_start_raw = (request.POST.get("week_start") or "").strip()
    week_start = timezone.localdate()
    if week_start_raw:
        try:
            week_start = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
        except ValueError:
            week_start = timezone.localdate()
    context = build_class_detail_context(request, branch=branch, class_id=class_id, week_start=week_start)
    context["toast"] = toast
    context["planner_intent"] = (request.POST.get("planner_intent") or "create").strip() or "create"
    _inject_planner_route_context(context, role_prefix="director", workspace_target_id="#director-workspace")
    return render(request, "portal/staff/supervisor/partials/planner_workspace.html", context)


@_position_required({"director_of_studies", "academic_supervisor", "executive_director", "deputy_executive_director", "super_admin"})
def schedule_class_print(request, class_id: int):
    branch = _resolve_academic_branch(request)
    if branch is None:
        return HttpResponseForbidden("Aucune annexe n'est rattachee a ce compte.")

    academic_class = AcademicClass.objects.select_related("programme", "academic_year", "branch").filter(
        pk=class_id,
        branch=branch,
        is_active=True,
    ).first()
    if academic_class is None:
        return HttpResponseForbidden("Classe introuvable pour cette annexe.")

    week_start, period, weeks = _parse_planning_period_request(request)
    week_blocks = _build_print_week_blocks(get_class_week_schedule, academic_class, week_start, weeks)
    return render(
        request,
        "portal/staff/shared/prints/class_schedule_print.html",
        {
            "branch": branch,
            "academic_class": academic_class,
            "week_blocks": week_blocks,
            "period": period,
            "generated_at": timezone.now(),
        },
    )


@_position_required({"director_of_studies", "academic_supervisor", "executive_director", "deputy_executive_director", "super_admin"})
def schedule_teacher_print(request, teacher_id: int):
    branch = _resolve_academic_branch(request)
    if branch is None:
        return HttpResponseForbidden("Aucune annexe n'est rattachee a ce compte.")

    teacher = get_user_model().objects.select_related("profile").filter(
        pk=teacher_id,
        is_active=True,
        profile__branch=branch,
        profile__position="teacher",
    ).first()
    if teacher is None:
        return HttpResponseForbidden("Enseignant introuvable pour cette annexe.")

    week_start, period, weeks = _parse_planning_period_request(request)
    week_blocks = _build_print_week_blocks(get_teacher_week_schedule, teacher, week_start, weeks)
    return render(
        request,
        "portal/staff/shared/prints/teacher_schedule_print.html",
        {
            "branch": branch,
            "teacher": teacher,
            "week_blocks": week_blocks,
            "period": period,
            "generated_at": timezone.now(),
        },
    )

@login_required
def dg_portal(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    context = build_dg_dashboard_context(request, _build_portal_context)
    return render(request, "portal/dg/dashboard.html", context)


@login_required
def dg_section(request, section: str):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    context = build_dg_section_context(request, section, _build_portal_context)
    if section == "workflows":
        terminal_classes = AcademicClass.objects.select_related("branch", "academic_year", "programme").filter(
            is_active=True,
            level__in=["L3", "M2"],
        )
        diploma_rows = []
        for academic_class in terminal_classes[:30]:
            semesters = list(academic_class.semesters.all())
            if not semesters:
                continue
            published_count = sum(1 for semester in semesters if semester.status == Semester.STATUS_PUBLISHED)
            if published_count < len(semesters):
                continue
            student_count = academic_class.enrollments.filter(is_active=True).count()
            awards_count = AcademicDiplomaAward.objects.filter(academic_class=academic_class).exclude(
                status=AcademicDiplomaAward.STATUS_CANCELLED
            ).count()
            diploma_rows.append(
                {
                    "class": academic_class,
                    "student_count": student_count,
                    "published_count": published_count,
                    "semester_count": len(semesters),
                    "awards_count": awards_count,
                }
            )
        context["diploma_candidate_classes"] = diploma_rows
    template_map = {
        "kpis": "portal/dg/partials/kpis/overview.html",
        "alerts": "portal/dg/partials/alerts/priority_table.html",
        "workflows": "portal/dg/partials/workflows/reenrollment.html",
        "finance": "portal/dg/partials/finance/summary.html",
        "annexes": "portal/dg/partials/annexes/performance_table.html",
        "analytics": "portal/dg/partials/analytics/charts.html",
        "schedule": "portal/dg/partials/schedule/overview.html",
        "realtime": "portal/dg/partials/realtime/monitoring.html",
        "rh": "portal/dg/partials/rh/staff.html",
    }
    template = template_map.get(section)
    if not template:
        return HttpResponseBadRequest("Section DG inconnue.")
    return render(request, template, context)


@login_required
def dg_drawer(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    return render(request, "portal/dg/drawers/detail.html", build_dg_drawer_context(request))


@login_required
def dg_modal(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    modal = (request.GET.get("modal") or "recruitment").strip().lower()
    if modal == "recruitment":
        return render(request, "portal/dg/modals/recruitment.html", {"form": DgRecruitmentForm()})
    if modal in {"branch", "alert", "case", "workflow", "finance", "analytics", "realtime", "rh"}:
        context = build_dg_drawer_context(request)
        context.update(
            {
                "modal_kind": modal,
                "modal_title": {
                    "branch": "Fiche annexe detaillee",
                    "alert": "Alerte detaillee",
                    "case": "Dossier etudiant detaille",
                    "workflow": "Workflow detaille",
                    "finance": "Synthese financiere detaillee",
                    "analytics": "Lecture analytique detaillee",
                    "realtime": "Vue live detaillee",
                    "rh": "Lecture RH detaillee",
                }[modal],
                "modal_finance": context.get("drawer_branch_finance") or context.get("drawer_finance") or context.get("finance"),
            }
        )
        return render(request, "portal/dg/modals/detail.html", context)
    return HttpResponseBadRequest("Modal DG inconnue.")


@login_required
def dg_recruit_staff(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    if request.method != "POST":
        return HttpResponseBadRequest("Methode invalide.")
    form = DgRecruitmentForm(request.POST)
    if not form.is_valid():
        return render(request, "portal/dg/modals/recruitment.html", {"form": form}, status=400)
    result = create_staff_from_recruitment(actor=request.user, form=form)
    return render(request, "portal/dg/modals/recruitment_success.html", result)


@login_required
def dg_action(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Acces reserve au Directeur General.")
    if request.method != "POST":
        return HttpResponseBadRequest("Methode invalide.")

    action = (request.POST.get("action") or "").strip()
    object_id = (request.POST.get("object_id") or "").strip()
    if action not in {"resolve_alert", "resolve_case", "escalate_case", "followup_finance"}:
        return JsonResponse({"ok": False, "message": "Action DG inconnue."}, status=400)
    if not object_id.isdigit():
        return JsonResponse({"ok": False, "message": "Reference invalide."}, status=400)

    try:
        if action == "resolve_alert":
            result = resolve_attendance_alert(actor=request.user, alert_id=int(object_id))
        elif action == "resolve_case":
            result = resolve_student_case(actor=request.user, case_id=int(object_id))
        elif action == "escalate_case":
            result = escalate_student_case(actor=request.user, case_id=int(object_id))
        else:
            branch = Branch.objects.get(id=int(object_id), is_active=True)
            result = create_finance_followup(actor=request.user, branch=branch)
    except (AttendanceAlert.DoesNotExist, StudentCase.DoesNotExist, Branch.DoesNotExist):
        return JsonResponse({"ok": False, "message": "Element introuvable."}, status=404)
    return JsonResponse({"ok": True, **result})


@login_required
def dg_diploma_action(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Acces reserve a la Direction Generale.")
    if request.method != "POST":
        return HttpResponseBadRequest("Methode invalide.")
    class_id = (request.POST.get("class_id") or "").strip()
    publish = (request.POST.get("publish") or "").strip().lower() in {"1", "true", "yes", "on"}
    if not class_id.isdigit():
        return JsonResponse({"ok": False, "message": "Classe invalide."}, status=400)
    try:
        academic_class = AcademicClass.objects.select_related("branch", "academic_year", "programme").get(pk=int(class_id), is_active=True)
        if not can_manage_diplomas(request.user, academic_class):
            return JsonResponse({"ok": False, "message": "Action non autorisee."}, status=403)
        result = prepare_diploma_awards_for_class(academic_class=academic_class, actor=request.user, publish=publish)
    except AcademicClass.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Classe introuvable."}, status=404)
    return JsonResponse(
        {
            "ok": True,
            "message": f"{len(result['awards'])} diplome(s) prepare(s). {len(result['skipped'])} dossier(s) ignore(s).",
            "prepared": len(result["awards"]),
            "skipped": result["skipped"],
        }
    )


@login_required
def dg_exec_action(request):
    """Point d'entrée unique pour les actions DG réelles (nommer, diplômer, valider, arbitrer, cycle)."""
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    if request.method != "POST":
        return HttpResponseBadRequest("Méthode invalide.")

    action = (request.POST.get("action") or "").strip()
    try:
        if action == "nominate_manager":
            branch_id = int(request.POST["branch_id"])
            user_id = int(request.POST["user_id"])
            result = nominate_branch_manager(actor=request.user, branch_id=branch_id, user_id=user_id)
        elif action == "deliver_diploma":
            award_id = int(request.POST["award_id"])
            result = deliver_diploma(actor=request.user, award_id=award_id)
        elif action == "validate_closure":
            closure_id = int(request.POST["closure_id"])
            result = validate_closure(actor=request.user, closure_id=closure_id)
        elif action == "arbitrate_decision":
            decision_id = int(request.POST["decision_id"])
            approve = (request.POST.get("approve") or "").strip().lower() in {"1", "true", "oui"}
            reason = (request.POST.get("reason") or "").strip()
            result = arbitrate_decision(actor=request.user, decision_id=decision_id, approve=approve, reason=reason)
        elif action == "publish_class_diplomas":
            class_id = int(request.POST["class_id"])
            result = publish_class_diplomas(actor=request.user, class_id=class_id)
        elif action == "transition_cycle":
            cycle_id = int(request.POST["cycle_id"])
            target_status = (request.POST.get("target_status") or "").strip()
            result = transition_branch_cycle(actor=request.user, cycle_id=cycle_id, target_status=target_status)
        else:
            return JsonResponse({"ok": False, "message": "Action DG inconnue."}, status=400)
    except (Branch.DoesNotExist, AcademicDiplomaAward.DoesNotExist,
            BranchMonthlyClosure.DoesNotExist, StudentYearDecision.DoesNotExist,
            KeyError, ValueError) as exc:
        return JsonResponse({"ok": False, "message": f"Erreur : {exc}"}, status=404)
    return JsonResponse(result)


@login_required
def dg_assign_manager_modal(request):
    """Popup pour choisir un utilisateur à nommer gestionnaire d'une annexe."""
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    branch_id = (request.GET.get("branch_id") or "").strip()
    branch = None
    if branch_id.isdigit():
        branch = Branch.objects.filter(pk=int(branch_id), is_active=True).first()
    candidates = User.objects.filter(is_active=True).exclude(profile__position__in=["student"]).select_related("profile").order_by("last_name", "first_name")[:50]
    return render(request, "portal/dg/modals/assign_manager.html", {
        "branch": branch,
        "candidates": candidates,
    })


@login_required
def dg_closure_detail(request):
    """Drawer détaillant une clôture mensuelle."""
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Accès réservé au Directeur Général.")
    closure_id = (request.GET.get("closure_id") or "").strip()
    closure = None
    if closure_id.isdigit():
        closure = BranchMonthlyClosure.objects.select_related("branch", "created_by", "validated_by").filter(pk=int(closure_id)).first()
    return render(request, "portal/dg/modals/closure_detail.html", {
        "closure": closure,
    })


@login_required
def dg_export(request):
    position = get_user_position(request.user)
    if position not in {"executive_director", "deputy_executive_director"}:
        return HttpResponseForbidden("Acces reserve au Directeur General.")
    kind = (request.GET.get("kind") or "branches").strip().lower()
    export_format = (request.GET.get("format") or "csv").strip().lower()
    context = build_dg_dashboard_context(request, _build_portal_context)
    rows = []

    class _RowWriter:
        def writerow(self, row):
            rows.append([str(value) if value is not None else "" for value in row])

    writer = _RowWriter()

    if kind == "finance":
        writer.writerow(["Annexe", "Revenus", "Depenses", "Solde", "Paiements valides", "Alertes ouvertes"])
        for item in context["branch_summaries"]:
            writer.writerow([
                item["branch"].name,
                item["revenue_total"],
                item["expense_total"],
                item["balance_total"],
                len(item["latest_payments"]),
                item["open_alert_count"],
            ])
    elif kind == "alerts":
        writer.writerow(["Gravite", "Type", "Annexe", "Description", "Responsable", "Statut", "Age"])
        for alert in context["priority_alerts"]:
            writer.writerow([
                alert.severity,
                alert.type,
                alert.branch_name,
                alert.description,
                alert.owner,
                alert.status,
                alert.age,
            ])
    elif kind == "students":
        writer.writerow(["Matricule", "Nom", "Email", "Annexe", "Formation", "Classe", "Inscription", "Solde"])
        branch_ids = [item["branch"].id for item in context["branch_summaries"]]
        students_qs = (
            Student.objects.filter(is_active=True, inscription__candidature__branch_id__in=branch_ids)
            .select_related(
                "inscription",
                "inscription__candidature",
                "inscription__candidature__branch",
                "inscription__candidature__programme",
                "current_academic_enrollment__academic_class",
            )
            .order_by("inscription__candidature__last_name", "inscription__candidature__first_name")
        )
        for student in students_qs:
            candidature = student.inscription.candidature
            writer.writerow([
                student.matricule,
                student.full_name,
                candidature.email,
                candidature.branch.name if candidature.branch else "",
                candidature.programme.title if candidature.programme else "",
                student.current_academic_enrollment.academic_class if student.current_academic_enrollment else "",
                student.inscription.reference,
                student.inscription.balance,
            ])
    elif kind == "payments":
        writer.writerow(["Date", "Annexe", "Candidat", "Reference", "Methode", "Statut", "Montant", "Agent"])
        branch_ids = [item["branch"].id for item in context["branch_summaries"]]
        payments_qs = (
            Payment.objects.filter(inscription__candidature__branch_id__in=branch_ids)
            .select_related("agent__user", "inscription__candidature", "inscription__candidature__branch")
            .order_by("-paid_at", "-id")
        )
        for payment in payments_qs:
            candidature = payment.inscription.candidature
            writer.writerow([
                payment.paid_at.strftime("%Y-%m-%d %H:%M") if payment.paid_at else "",
                candidature.branch.name if candidature.branch else "",
                str(candidature),
                payment.reference,
                payment.get_method_display(),
                payment.get_status_display(),
                payment.amount,
                payment.agent.user.get_full_name() or payment.agent.user.username if payment.agent else "",
            ])
    elif kind == "staff":
        writer.writerow(["Nom", "Email", "Poste", "Role", "Annexe", "Statut", "Salaire", "Derniere activite"])
        branch_ids = [item["branch"].id for item in context["branch_summaries"]]
        staff_qs = (
            Profile.objects.filter(user_type="staff")
            .filter(Q(branch_id__in=branch_ids) | Q(branch__isnull=True))
            .select_related("user", "branch")
            .order_by("branch__name", "position", "user__last_name")
        )
        for profile in staff_qs:
            writer.writerow([
                profile.user.get_full_name() or profile.user.username,
                profile.user.email,
                profile.get_position_display() or profile.position,
                profile.get_role_display() or profile.role,
                profile.branch.name if profile.branch else "Global",
                profile.get_employment_status_display(),
                profile.salary_base,
                profile.last_seen.strftime("%Y-%m-%d %H:%M") if profile.last_seen else "",
            ])
    elif kind == "audit":
        writer.writerow(["Date", "Action", "Acteur", "Annexe", "Cible", "Details"])
        branch_ids = [item["branch"].id for item in context["branch_summaries"]]
        audit_qs = (
            SupportAuditLog.objects.select_related("actor", "branch")
            .filter(Q(branch_id__in=branch_ids) | Q(branch__isnull=True))
            .order_by("-created_at")[:1000]
        )
        for row in audit_qs:
            writer.writerow([
                row.created_at.strftime("%Y-%m-%d %H:%M"),
                row.get_action_type_display(),
                row.actor.get_full_name() or row.actor.username if row.actor else "",
                row.branch.name if row.branch else "",
                row.target_label,
                row.details,
            ])
    else:
        writer.writerow(["Annexe", "Manager", "Etudiants", "Classes", "Inscriptions", "Candidatures", "Revenus", "Depenses", "Solde", "Performance"])
        for item in context["branch_summaries"]:
            writer.writerow([
                item["branch"].name,
                item["manager_name"],
                item["student_count"],
                item["class_count"],
                item["active_inscription_count"],
                item["candidature_count"],
                item["revenue_total"],
                item["expense_total"],
                item["balance_total"],
                item["performance_label"],
            ])
    if export_format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"DG {kind}"[:31]
        for row in rows:
            sheet.append(row)
        if rows:
            header_fill = PatternFill("solid", fgColor="0F172A")
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            sheet.freeze_panes = "A2"
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="dg-{kind}.xlsx"'
        workbook.save(response)
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="dg-{kind}.csv"'
    csv_writer = csv.writer(response)
    for row in rows:
        csv_writer.writerow(row)
    return response
