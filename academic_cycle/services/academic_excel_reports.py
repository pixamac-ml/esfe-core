"""Export Excel des rapports pedagogiques pour le Directeur des Etudes
(cf. CAHIER_DES_CHARGES_DIRECTEUR_ETUDES.md, 2.6). Reutilise le meme moteur
de mise en forme que les rapports financiers du gestionnaire plutot que de
le redupliquer.
"""
from decimal import Decimal

from openpyxl import Workbook

from academics.models import AcademicEnrollment, Semester
from academics.services.semester import compute_semester_result
from accounts.services.excel_reports import (
    BODY_FONT,
    BOLD_FONT,
    TITLE_FONT,
    _auto_width,
    _style_header_row,
    _write_cell,
    xlsx_response,
)

__all__ = ["build_academic_report_xlsx", "xlsx_response"]


def _build_class_semester_rows(branch):
    """Une ligne par (classe, semestre) : notes saisies, taux de reussite,
    moyenne de classe, desinscriptions - reutilise compute_semester_result()
    deja utilise par la section resultats du dashboard plutot que de
    recalculer une logique d'agregation distincte."""
    semesters = (
        Semester.objects.select_related("academic_class", "academic_class__programme", "academic_class__branch")
        .filter(academic_class__is_active=True)
    )
    if branch:
        semesters = semesters.filter(academic_class__branch=branch)

    rows = []
    for semester in semesters.order_by("academic_class__level", "academic_class__programme__title", "number"):
        academic_class = semester.academic_class
        enrollments = list(
            AcademicEnrollment.objects.filter(
                academic_class=academic_class,
                academic_year=academic_class.academic_year,
                is_active=True,
            )
        )
        dropouts = AcademicEnrollment.objects.filter(
            academic_class=academic_class,
            academic_year=academic_class.academic_year,
            is_active=False,
        ).count()

        total_average = Decimal("0.00")
        average_count = 0
        validated_count = 0
        for enrollment in enrollments:
            result = compute_semester_result(semester, enrollment)
            if result["average"] is not None:
                total_average += Decimal(str(result["average"]))
                average_count += 1
            if result["is_validated"]:
                validated_count += 1

        class_average = (total_average / average_count) if average_count else None
        success_rate = (validated_count / len(enrollments) * 100) if enrollments else None

        rows.append({
            "class_label": academic_class.display_name,
            "programme": academic_class.programme.title if academic_class.programme_id else "-",
            "semester_number": semester.number,
            "status": semester.get_status_display(),
            "student_count": len(enrollments),
            "class_average": class_average,
            "success_rate": success_rate,
            "dropouts": dropouts,
        })
    return rows


def build_academic_report_xlsx(*, branch):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport pedagogique"
    ws.sheet_properties.tabColor = "1e4f6f"

    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    title_cell = ws.cell(row=row, column=1, value=f"Rapport pedagogique - {branch.name if branch else 'Toutes annexes'}")
    title_cell.font = TITLE_FONT
    row += 2

    headers = [
        "Classe", "Programme", "Semestre", "Statut", "Etudiants actifs",
        "Moyenne de classe", "Taux de reussite", "Desinscriptions",
    ]
    for col_i, header in enumerate(headers, 1):
        _write_cell(ws, row, col_i, header, font=BOLD_FONT)
    _style_header_row(ws, row, len(headers))
    row += 1

    for entry in _build_class_semester_rows(branch):
        _write_cell(ws, row, 1, entry["class_label"], font=BODY_FONT)
        _write_cell(ws, row, 2, entry["programme"], font=BODY_FONT)
        _write_cell(ws, row, 3, entry["semester_number"], font=BODY_FONT, align="center")
        _write_cell(ws, row, 4, entry["status"], font=BODY_FONT)
        _write_cell(ws, row, 5, entry["student_count"], font=BODY_FONT, align="center")
        _write_cell(
            ws, row, 6,
            float(entry["class_average"]) if entry["class_average"] is not None else "-",
            font=BODY_FONT, align="center",
        )
        _write_cell(
            ws, row, 7,
            f"{entry['success_rate']:.0f}%" if entry["success_rate"] is not None else "-",
            font=BODY_FONT, align="center",
        )
        _write_cell(ws, row, 8, entry["dropouts"], font=BODY_FONT, align="center")
        row += 1

    _auto_width(ws)
    return wb
