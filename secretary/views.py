import json

from django.contrib import messages
from django.contrib.messages import get_messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Count

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.http import urlencode
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from accounts.dashboards.helpers import get_user_branch
from accounts.forms import ProfileForm, UserPreferenceForm
from accounts.models import PayrollEntry, Profile, UserPreference
from communication.models import CommunicationNotification
from communication.selectors import get_user_notifications, get_user_unread_count
from communication.services import NotificationService
from .forms import (
    AppointmentForm,
    DocumentReceiptForm,
    RegistryEntryForm,
    SecretaryTaskForm,
    VisitorLogForm,
)
from .models import DocumentReceipt, RegistryEntry, SecretaryTask
from .permissions import ensure_secretary_access
from .selectors import (
    get_active_students,
    get_active_visits_queryset,
    get_appointments_queryset,
    get_classes_queryset,
    get_documents_queryset,
    get_registry_queryset,
    get_student_snapshot_queryset,
    get_tasks_queryset,
    get_today_appointments_queryset,
    get_today_visits_queryset,
    get_visits_queryset,
)
from .services import (
    archive_appointment,
    archive_document,
    archive_registry_entry,
    complete_appointment,
    complete_task,
    close_visit,
    create_appointment,
    create_registry_entry,
    create_task,
    start_appointment_processing,
    start_document_processing,
    start_registry_entry_processing,
    start_task_processing,
    get_secretary_dashboard_data,
    get_secretary_recent_messages,
    get_registry_routing_rules_for_ui,
    get_secretary_class_students,
    get_student_snapshot,
    mark_registry_processed,
    move_registry_entry_status,
    register_document,
    register_visitor,
    search_academic_classes,
    search_students,
    update_appointment,
    update_document,
    update_registry_entry,
    update_task,
    update_visitor,
)


def _paginate(request, queryset, per_page=20, page_param="page"):
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get(page_param)
    return paginator.get_page(page_number)


def _common_filters(request):
    return {
        "q": request.GET.get("q", "").strip(),
        "status": request.GET.get("status", "").strip(),
        "archived": False,
        "active_only": True,
    }


def _form_kwargs(request):
    return {"user": request.user, "branch": get_user_branch(request.user)}


def _is_htmx(request):
    return request.headers.get("HX-Request") == "true"


def _secretary_profile_form(instance):
    form = ProfileForm(instance=instance)
    for field_name, field in form.fields.items():
        if field_name == "avatar":
            field.widget.attrs.update({
                "class": "sg-form-input",
                "accept": "image/jpeg,image/png,image/webp",
            })
        elif field.widget.__class__.__name__ == "Textarea":
            field.widget.attrs.update({"class": "sg-form-input", "rows": 3})
        else:
            field.widget.attrs.update({"class": "sg-form-input"})
    return form


def _secretary_preference_form(instance):
    form = UserPreferenceForm(instance=instance)
    for field in form.fields.values():
        field.widget.attrs.update({"style": "width:18px;height:18px;flex-shrink:0;"})
    return form


SECRETARY_DASHBOARD_SECTIONS = {
    "overview",
    "registry",
    "visits",
    "classes",
    "students",
    "appointments",
    "deposits",
    "reports",
    "salary",
    "notifications",
    "settings",
}

SECRETARY_REFRESH_EVENTS = {
    "kpis": "secretary:kpis-updated",
    "sidebar": "secretary:sidebar-updated",
    "registry": "secretary:registry-updated",
    "visits": "secretary:visits-updated",
    "appointments": "secretary:appointments-updated",
    "documents": "secretary:documents-updated",
    "tasks": "secretary:tasks-updated",
    "notifications": "secretary:notifications-updated",
}


def _dashboard_section(request):
    section = request.GET.get("section", "overview").strip()
    return section if section in SECRETARY_DASHBOARD_SECTIONS else "overview"


SECRETARY_TOAST_TONES = {"success": "success", "error": "error", "warning": "notification", "info": "notification", "debug": "notification"}
SECRETARY_TOAST_TITLES = {"success": "Succes", "error": "Erreur", "warning": "Attention", "info": "Information", "debug": "Information"}


def _pending_messages_toasts(request):
    toasts = []
    for message in get_messages(request):
        toasts.append({
            "title": SECRETARY_TOAST_TITLES.get(message.tags, "Information"),
            "body": str(message),
            "tone": SECRETARY_TOAST_TONES.get(message.tags, "notification"),
        })
    return toasts


def _trigger_payload(*groups, close_drawer=False, toasts=None):
    payload = {}
    if close_drawer:
        payload["secretary-drawer-close"] = True
    for group in groups:
        event_name = SECRETARY_REFRESH_EVENTS.get(group)
        if event_name:
            payload[event_name] = True
    if groups:
        payload[SECRETARY_REFRESH_EVENTS["sidebar"]] = True
    if toasts:
        payload["secretary-toast"] = toasts
    return payload


def _drawer_close_response(request, *groups):
    response = HttpResponse("", content_type="text/html")
    response["HX-Trigger"] = json.dumps(_trigger_payload(*groups, close_drawer=True, toasts=_pending_messages_toasts(request)))
    return response


def _refresh_response(request, *groups):
    response = HttpResponse("", content_type="text/html")
    response["HX-Trigger"] = json.dumps(_trigger_payload(*groups, toasts=_pending_messages_toasts(request)))
    return response


def _dashboard_base_context(request):
    ensure_secretary_access(request.user)
    return get_secretary_dashboard_data(request.user)


def _validation_error_message(error):
    messages_list = getattr(error, "messages", None)
    if messages_list:
        return " ".join(str(message) for message in messages_list)
    return str(error)


def _render_create_drawer_or_page(request, drawer_template, page_template, context):
    if _is_htmx(request):
        return render(request, drawer_template, context)
    return render(request, page_template, context)


@login_required
def secretary_dashboard(request):
    ensure_secretary_access(request.user)
    context = get_secretary_dashboard_data(request.user)
    branch = context.get("branch")
    context["active_section"] = _dashboard_section(request)
    context["student_query"] = request.GET.get("student_q", "").strip()
    if context["student_query"]:
        context["active_section"] = "students"
    selected_notification = None
    notification_id = request.GET.get("notification_id")
    if notification_id:
        selected_notification = get_object_or_404(
            CommunicationNotification.objects.select_related("actor", "event"),
            pk=notification_id,
            recipient=request.user,
            channel=CommunicationNotification.CHANNEL_IN_APP,
        )
        NotificationService.mark_as_read(selected_notification)
        context["active_section"] = "notifications"
    if context["student_query"]:
        students_queryset = search_students(context["student_query"], user=request.user)
    else:
        students_queryset = get_active_students(user=request.user, branch=branch)
    context["student_results_page"] = _paginate(request, students_queryset, per_page=10, page_param="students_page")
    context["student_results"] = context["student_results_page"].object_list

    context["dashboard_pending_registry_page"] = _paginate(
        request,
        get_registry_queryset(
            {"status": RegistryEntry.STATUS_PENDING, "archived": False, "active_only": True},
            user=request.user,
            branch=branch,
        ),
        per_page=6,
        page_param="overview_registry_page",
    )
    context["pending_registry_rows"] = context["dashboard_pending_registry_page"].object_list

    context["registry_page"] = _paginate(
        request,
        get_registry_queryset({"archived": False, "active_only": True}, user=request.user, branch=branch),
        per_page=10,
        page_param="registry_page",
    )
    context["recent_registry"] = context["registry_page"].object_list

    registry_qs = get_registry_queryset({"archived": False, "active_only": True}, user=request.user, branch=branch)
    type_counts = registry_qs.values("entry_type").annotate(count=Count("id")).order_by("-count")
    total = sum(tc["count"] for tc in type_counts) or 1
    context["registry_type_distribution"] = [
        {
            "entry_type": tc["entry_type"],
            "label": dict(RegistryEntry.ENTRY_TYPE_CHOICES).get(tc["entry_type"], tc["entry_type"]),
            "count": tc["count"],
            "pct": round(tc["count"] / total * 100),
        }
        for tc in type_counts
    ]

    documents_for_funnel = get_documents_queryset({}, user=request.user, branch=branch)
    document_status_counts = {
        row["status"]: row["count"]
        for row in documents_for_funnel.values("status").annotate(count=Count("id"))
    }
    documents_total = sum(document_status_counts.values()) or 1
    notified_count = sum(
        document_status_counts.get(status, 0)
        for status in (
            DocumentReceipt.STATUS_IN_PROGRESS,
            DocumentReceipt.STATUS_TRANSFERRED,
            DocumentReceipt.STATUS_COMPLETED,
            DocumentReceipt.STATUS_ARCHIVED,
        )
    )
    delivered_count = sum(
        document_status_counts.get(status, 0)
        for status in (DocumentReceipt.STATUS_COMPLETED, DocumentReceipt.STATUS_ARCHIVED)
    )
    context["document_process_funnel"] = [
        {
            "label": "1. Depot identifie",
            "sub": "Recu et numerote",
            "pct": 100 if document_status_counts else 0,
            "color": "#2563eb",
        },
        {
            "label": "2. Destinataire notifie",
            "sub": "En attente de remise",
            "pct": round(notified_count / documents_total * 100),
            "color": "#7c3aed",
        },
        {
            "label": "3. Remise confirmee",
            "sub": "Traite",
            "pct": round(delivered_count / documents_total * 100),
            "color": "#16a34a",
        },
    ]

    context["registry_routing_summary"] = get_registry_routing_rules_for_ui()

    context["open_visits_page"] = _paginate(
        request,
        get_active_visits_queryset(user=request.user, branch=branch),
        per_page=10,
        page_param="open_visits_page",
    )
    context["open_visits_rows"] = context["open_visits_page"].object_list
    context["today_visits_page"] = _paginate(
        request,
        get_today_visits_queryset(user=request.user, branch=branch),
        per_page=10,
        page_param="today_visits_page",
    )
    context["today_visits_rows"] = context["today_visits_page"].object_list

    context["appointments_page"] = _paginate(
        request,
        get_today_appointments_queryset(user=request.user, branch=branch),
        per_page=10,
        page_param="appointments_page",
    )
    context["today_appointments_rows"] = context["appointments_page"].object_list

    context["documents_page"] = _paginate(
        request,
        get_documents_queryset(
            {"status": DocumentReceipt.STATUS_PENDING, "archived": False, "active_only": True},
            user=request.user,
            branch=branch,
        ),
        per_page=10,
        page_param="documents_page",
    )
    context["pending_documents_rows"] = context["documents_page"].object_list

    context["classes_page"] = _paginate(
        request,
        get_classes_queryset(user=request.user, branch=branch),
        per_page=12,
        page_param="classes_page",
    )
    context["classes_rows"] = context["classes_page"].object_list

    context["tasks_page"] = _paginate(
        request,
        get_tasks_queryset(
            {"status": SecretaryTask.STATUS_PENDING, "archived": False, "active_only": True},
            user=request.user,
            branch=branch,
        ),
        per_page=8,
        page_param="tasks_page",
    )
    context["pending_tasks_rows"] = context["tasks_page"].object_list

    profile, _profile_created = Profile.objects.get_or_create(user=request.user)
    preference, _preference_created = UserPreference.objects.get_or_create(user=request.user)
    context["secretary_profile"] = profile
    context["secretary_preference"] = preference
    context["profile_form"] = _secretary_profile_form(profile)
    context["preference_form"] = _secretary_preference_form(preference)

    payroll_queryset = PayrollEntry.objects.filter(employee=request.user).select_related("branch", "created_by", "updated_by")
    context["salary_entries_page"] = _paginate(request, payroll_queryset, per_page=8, page_param="salary_page")
    context["salary_entries"] = context["salary_entries_page"].object_list
    context["latest_salary_entry"] = payroll_queryset.first()

    notifications_queryset = get_user_notifications(
        request.user,
        channel=CommunicationNotification.CHANNEL_IN_APP,
    )
    context["notifications_page"] = _paginate(request, notifications_queryset, per_page=10, page_param="notifications_page")
    context["notifications_rows"] = context["notifications_page"].object_list
    context["selected_notification"] = selected_notification
    context["messages_count"] = get_user_unread_count(request.user)

    context["quick_registry_form"] = RegistryEntryForm(**_form_kwargs(request))
    context["quick_appointment_form"] = AppointmentForm(**_form_kwargs(request))
    context["quick_visitor_form"] = VisitorLogForm(**_form_kwargs(request))
    context["quick_document_form"] = DocumentReceiptForm(**_form_kwargs(request))
    context["quick_task_form"] = SecretaryTaskForm(**_form_kwargs(request))
    return render(request, "secretary/dashboard.html", context)


@login_required
@require_POST
def secretary_profile_update(request):
    ensure_secretary_access(request.user)
    profile, _created = Profile.objects.get_or_create(user=request.user)
    form = ProfileForm(request.POST, request.FILES, instance=profile)
    if form.is_valid():
        form.save()
        messages.success(request, "Profil mis a jour.")
    else:
        messages.error(request, "Le profil contient des informations a corriger.")
    return redirect(f"{reverse('secretary:secretary_dashboard')}?section=settings")


@login_required
@require_POST
def secretary_preferences_update(request):
    ensure_secretary_access(request.user)
    preference, _created = UserPreference.objects.get_or_create(user=request.user)
    form = UserPreferenceForm(request.POST, instance=preference)
    if form.is_valid():
        form.save()
        messages.success(request, "Preferences mises a jour.")
    else:
        messages.error(request, "Les preferences contiennent des informations a corriger.")
    return redirect(f"{reverse('secretary:secretary_dashboard')}?section=settings")


KANBAN_STATUSES = [
    (RegistryEntry.STATUS_PENDING, "En attente"),
    (RegistryEntry.STATUS_IN_PROGRESS, "En cours"),
    (RegistryEntry.STATUS_TRANSFERRED, "Transfere"),
    (RegistryEntry.STATUS_COMPLETED, "Traite"),
]
KANBAN_STATUS_ORDER = [status for status, _ in KANBAN_STATUSES]
KANBAN_STATUS_LABELS = dict(KANBAN_STATUSES)


def _kanban_filters(request):
    return {
        "q": request.GET.get("q", "").strip(),
        "priority": request.GET.get("priority", "").strip(),
        "target_service": request.GET.get("target_service", "").strip(),
        "archived": False,
        "active_only": True,
    }


def _kanban_querystring(filters):
    params = {key: filters[key] for key in ("q", "priority", "target_service") if filters.get(key)}
    return f"?{urlencode(params)}" if params else ""


def _kanban_columns(request, filters):
    queryset = get_registry_queryset(filters, user=request.user)
    entries_by_status = {status: [] for status in KANBAN_STATUS_ORDER}
    for entry in queryset:
        if entry.status not in entries_by_status:
            continue
        allowed = entry.get_allowed_status_transitions()
        entry.kanban_moves = [
            (status, KANBAN_STATUS_LABELS[status])
            for status in KANBAN_STATUS_ORDER
            if status != entry.status and status in allowed
        ]
        entries_by_status[entry.status].append(entry)
    return [
        {"status": status, "label": label, "entries": entries_by_status[status], "count": len(entries_by_status[status])}
        for status, label in KANBAN_STATUSES
    ]


def _registry_target_services(request):
    return (
        get_registry_queryset({"archived": False, "active_only": True}, user=request.user)
        .exclude(target_service="")
        .order_by("target_service")
        .values_list("target_service", flat=True)
        .distinct()
    )


@login_required
def registry_kanban(request):
    ensure_secretary_access(request.user)
    filters = _kanban_filters(request)
    return render(
        request,
        "secretary/registry_kanban.html",
        {
            "columns": _kanban_columns(request, filters),
            "filters": filters,
            "querystring": _kanban_querystring(filters),
            "priority_choices": RegistryEntry.PRIORITY_CHOICES,
            "target_service_choices": _registry_target_services(request),
        },
    )


@login_required
@require_POST
def registry_kanban_move(request, pk, new_status):
    ensure_secretary_access(request.user)
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    try:
        move_registry_entry_status(entry, new_status)
        messages.success(request, "Statut de l'entree mis a jour.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    filters = _kanban_filters(request)
    response = render(
        request,
        "secretary/partials/registry_kanban_board.html",
        {
            "columns": _kanban_columns(request, filters),
            "querystring": _kanban_querystring(filters),
        },
    )
    response["HX-Trigger"] = json.dumps(_trigger_payload("registry", "kpis", toasts=_pending_messages_toasts(request)))
    return response


@login_required
def registry_list(request):
    ensure_secretary_access(request.user)
    queryset = get_registry_queryset(_common_filters(request), user=request.user)
    return render(
        request,
        "secretary/registry_list.html",
        {
            "page_obj": _paginate(request, queryset),
            "filters": _common_filters(request),
        },
    )


@login_required
def registry_create(request):
    ensure_secretary_access(request.user)
    initial = {}
    student_id = request.GET.get("student")
    if student_id:
        initial["related_student"] = student_id
    form = RegistryEntryForm(
        request.POST or None,
        request.FILES or None,
        initial=initial,
        **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            create_registry_entry(created_by=request.user, **form.cleaned_data)
            messages.success(request, "Entree de registre enregistree.")
            if _is_htmx(request):
                return _drawer_close_response(request, "registry", "kpis")
            return redirect("secretary:registry_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/registry_form.html",
        {
            "form": form,
            "page_title": "Nouvelle entree de registre",
            "modal_kicker": "Registre",
            "modal_description": "Ajouter une entree sans quitter le tableau de bord.",
            "submit_label": "Enregistrer l'entree",
            "return_url": "secretary:registry_list",
            "form_action_url": "secretary:registry_create",
            "registry_routing_rules": get_registry_routing_rules_for_ui(),
        },
    )


@login_required
def registry_update(request, pk):
    ensure_secretary_access(request.user)
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    form = RegistryEntryForm(
        request.POST or None, request.FILES or None,
        instance=entry, **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            update_registry_entry(entry, **form.cleaned_data)
            messages.success(request, "Entree du registre modifiee.")
            if _is_htmx(request):
                return _drawer_close_response(request, "registry", "kpis")
            return redirect("secretary:registry_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/registry_form.html",
        {
            "form": form,
            "page_title": "Modifier l'entree du registre",
            "modal_kicker": "Registre",
            "modal_description": "Modifier les informations de l'entree.",
            "submit_label": "Enregistrer les modifications",
            "return_url": "secretary:registry_list",
            "form_action_url": "secretary:registry_update",
            "form_action_pk": entry.pk,
            "registry_routing_rules": get_registry_routing_rules_for_ui(),
        },
    )


@login_required
@require_POST
def registry_mark_processed(request, pk):
    ensure_secretary_access(request.user)
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    try:
        mark_registry_processed(entry)
        messages.success(request, "Entree marquee comme traitee.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "registry", "kpis")
    return redirect("secretary:registry_list")


@login_required
def registry_start(request, pk):
    ensure_secretary_access(request.user)
    if request.method != "POST":
        return redirect("secretary:registry_list")
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    try:
        start_registry_entry_processing(entry)
        messages.success(request, "Entree prise en charge.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "registry", "kpis")
    return redirect("secretary:registry_list")


@login_required
def registry_detail(request, pk):
    ensure_secretary_access(request.user)
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    history_events = []
    for event in reversed(entry.history or []):
        at = event.get("at")
        parsed_at = parse_datetime(at) if at else None
        if parsed_at and timezone.is_aware(parsed_at):
            parsed_at = timezone.localtime(parsed_at)
        history_events.append({
            "at": parsed_at,
            "by": event.get("by", ""),
            "action": event.get("action", ""),
            "details": event.get("details") or {},
        })
    return render(
        request,
        "secretary/partials/registry_detail_drawer.html",
        {"entry": entry, "history_events": history_events},
    )


@login_required
@require_POST
def registry_archive(request, pk):
    ensure_secretary_access(request.user)
    entry = get_object_or_404(get_registry_queryset({}, user=request.user), pk=pk)
    try:
        archive_registry_entry(entry)
        messages.success(request, "Entree archivee.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "registry", "kpis")
    return redirect("secretary:registry_list")


@login_required
def appointment_list(request):
    ensure_secretary_access(request.user)
    filters = _common_filters(request)
    filters["date_from"] = request.GET.get("date_from") or None
    filters["date_to"] = request.GET.get("date_to") or None
    queryset = get_appointments_queryset(filters, user=request.user)
    return render(
        request,
        "secretary/appointment_list.html",
        {
            "page_obj": _paginate(request, queryset),
            "filters": filters,
        },
    )


@login_required
def appointment_create(request):
    ensure_secretary_access(request.user)
    form = AppointmentForm(request.POST or None, **_form_kwargs(request))
    if request.method == "POST" and form.is_valid():
        try:
            create_appointment(created_by=request.user, **form.cleaned_data)
            messages.success(request, "Rendez-vous cree.")
            if _is_htmx(request):
                return _drawer_close_response(request, "appointments", "kpis")
            return redirect("secretary:appointment_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/appointment_form.html",
        {
            "form": form,
            "page_title": "Nouveau rendez-vous",
            "modal_kicker": "Agenda",
            "modal_description": "Planifier un rendez-vous sans sortir du pilotage.",
            "submit_label": "Enregistrer le rendez-vous",
            "return_url": "secretary:appointment_list",
            "form_action_url": "secretary:appointment_create",
        },
    )


@login_required
def appointment_update(request, pk):
    ensure_secretary_access(request.user)
    appointment = get_object_or_404(get_appointments_queryset({}, user=request.user), pk=pk)
    form = AppointmentForm(
        request.POST or None, instance=appointment, **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            update_appointment(appointment, actor=request.user, **form.cleaned_data)
            messages.success(request, "Rendez-vous modifie.")
            if _is_htmx(request):
                return _drawer_close_response(request, "appointments", "kpis")
            return redirect("secretary:appointment_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/appointment_form.html",
        {
            "form": form,
            "page_title": "Modifier le rendez-vous",
            "modal_kicker": "Agenda",
            "modal_description": "Modifier les informations du rendez-vous.",
            "submit_label": "Enregistrer les modifications",
            "return_url": "secretary:appointment_list",
            "form_action_url": "secretary:appointment_update",
            "form_action_pk": appointment.pk,
        },
    )


@login_required
@require_POST
def appointment_complete(request, pk):
    ensure_secretary_access(request.user)
    appointment = get_object_or_404(get_appointments_queryset({}, user=request.user), pk=pk)
    try:
        complete_appointment(appointment)
        messages.success(request, "Rendez-vous marque comme termine.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "appointments", "kpis")
    return redirect("secretary:appointment_list")


@login_required
@require_POST
def appointment_archive(request, pk):
    ensure_secretary_access(request.user)
    appointment = get_object_or_404(get_appointments_queryset({}, user=request.user), pk=pk)
    try:
        archive_appointment(appointment)
        messages.success(request, "Rendez-vous archive.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "appointments", "kpis")
    return redirect("secretary:appointment_list")


@login_required
def visitor_list(request):
    ensure_secretary_access(request.user)
    queryset = get_visits_queryset(_common_filters(request), user=request.user)
    return render(
        request,
        "secretary/visitor_list.html",
        {
            "page_obj": _paginate(request, queryset),
            "filters": _common_filters(request),
        },
    )


@login_required
def visitor_update(request, pk):
    ensure_secretary_access(request.user)
    visitor = get_object_or_404(get_visits_queryset({}, user=request.user), pk=pk)
    form = VisitorLogForm(
        request.POST or None, instance=visitor, **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            update_visitor(visitor, **form.cleaned_data)
            messages.success(request, "Visite modifiee.")
            if _is_htmx(request):
                return _drawer_close_response(request, "visits", "kpis")
            return redirect("secretary:visitor_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/visitor_form.html",
        {
            "form": form,
            "page_title": "Modifier la visite",
            "modal_kicker": "Accueil",
            "modal_description": "Modifier les informations de la visite.",
            "submit_label": "Enregistrer les modifications",
            "return_url": "secretary:visitor_list",
            "form_action_url": "secretary:visitor_update",
            "form_action_pk": visitor.pk,
        },
    )


@login_required
@require_POST
def visitor_complete(request, pk):
    ensure_secretary_access(request.user)
    visitor = get_object_or_404(get_visits_queryset({}, user=request.user), pk=pk)
    try:
        close_visit(visitor)
        messages.success(request, "Visite cloturee.")
    except ValidationError:
        messages.info(request, "Cette visite etait deja cloturee.")
    if _is_htmx(request):
        return _refresh_response(request, "visits", "kpis")
    return redirect("secretary:visitor_list")


@login_required
def visitor_create(request):
    ensure_secretary_access(request.user)
    form = VisitorLogForm(request.POST or None, **_form_kwargs(request))
    if request.method == "POST" and form.is_valid():
        try:
            register_visitor(created_by=request.user, **form.cleaned_data)
            messages.success(request, "Visite enregistree.")
            if _is_htmx(request):
                return _drawer_close_response(request, "visits", "kpis")
            return redirect("secretary:visitor_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/visitor_form.html",
        {
            "form": form,
            "page_title": "Nouvelle visite",
            "modal_kicker": "Accueil",
            "modal_description": "Enregistrer un visiteur et son lien avec le service.",
            "submit_label": "Enregistrer la visite",
            "return_url": "secretary:visitor_list",
            "form_action_url": "secretary:visitor_create",
        },
    )


@login_required
def document_receipt_list(request):
    ensure_secretary_access(request.user)
    queryset = get_documents_queryset(_common_filters(request), user=request.user)
    return render(
        request,
        "secretary/document_receipt_list.html",
        {
            "page_obj": _paginate(request, queryset),
            "filters": _common_filters(request),
        },
    )


@login_required
def document_receipt_create(request):
    ensure_secretary_access(request.user)
    form = DocumentReceiptForm(request.POST or None, request.FILES or None, **_form_kwargs(request))
    if request.method == "POST" and form.is_valid():
        try:
            register_document(received_by=request.user, **form.cleaned_data)
            messages.success(request, "Document enregistre.")
            if _is_htmx(request):
                return _drawer_close_response(request, "documents", "kpis")
            return redirect("secretary:document_receipt_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/document_receipt_form.html",
        {
            "form": form,
            "page_title": "Nouveau document",
            "modal_kicker": "Pieces",
            "modal_description": "Recevoir un document avec sa trace administrative.",
            "submit_label": "Enregistrer le document",
            "return_url": "secretary:document_receipt_list",
            "form_action_url": "secretary:document_receipt_create",
        },
    )


@login_required
def document_receipt_update(request, pk):
    ensure_secretary_access(request.user)
    document = get_object_or_404(get_documents_queryset({}, user=request.user), pk=pk)
    form = DocumentReceiptForm(
        request.POST or None, request.FILES or None,
        instance=document, **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            update_document(document, **form.cleaned_data)
            messages.success(request, "Document modifie.")
            if _is_htmx(request):
                return _drawer_close_response(request, "documents", "kpis")
            return redirect("secretary:document_receipt_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/document_receipt_form.html",
        {
            "form": form,
            "page_title": "Modifier le document",
            "modal_kicker": "Pieces",
            "modal_description": "Modifier les informations du document.",
            "submit_label": "Enregistrer les modifications",
            "return_url": "secretary:document_receipt_list",
            "form_action_url": "secretary:document_receipt_update",
            "form_action_pk": document.pk,
        },
    )


@login_required
@require_POST
def document_receipt_archive(request, pk):
    ensure_secretary_access(request.user)
    document = get_object_or_404(get_documents_queryset({}, user=request.user), pk=pk)
    try:
        archive_document(document)
        messages.success(request, "Document archive.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "documents", "kpis")
    return redirect("secretary:document_receipt_list")


@login_required
def document_receipt_start(request, pk):
    ensure_secretary_access(request.user)
    if request.method != "POST":
        return redirect("secretary:document_receipt_list")
    document = get_object_or_404(get_documents_queryset({}, user=request.user), pk=pk)
    try:
        start_document_processing(document)
        messages.success(request, "Document pris en charge.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "documents", "kpis")
    return redirect("secretary:document_receipt_list")


@login_required
def task_list(request):
    ensure_secretary_access(request.user)
    queryset = get_tasks_queryset(_common_filters(request), user=request.user)
    return render(
        request,
        "secretary/task_list.html",
        {
            "page_obj": _paginate(request, queryset),
            "filters": _common_filters(request),
        },
    )


@login_required
def task_create(request):
    ensure_secretary_access(request.user)
    form = SecretaryTaskForm(request.POST or None, **_form_kwargs(request))
    if request.method == "POST" and form.is_valid():
        try:
            create_task(created_by=request.user, **form.cleaned_data)
            messages.success(request, "Tache creee.")
            if _is_htmx(request):
                return _drawer_close_response(request, "tasks", "kpis")
            return redirect("secretary:task_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/task_form.html",
        {
            "form": form,
            "page_title": "Nouvelle tache",
            "modal_kicker": "Suivi",
            "modal_description": "Créer une tâche de pilotage sans quitter la page.",
            "submit_label": "Enregistrer la tache",
            "return_url": "secretary:task_list",
            "form_action_url": "secretary:task_create",
        },
    )


@login_required
def task_update(request, pk):
    ensure_secretary_access(request.user)
    task = get_object_or_404(get_tasks_queryset({}, user=request.user), pk=pk)
    form = SecretaryTaskForm(
        request.POST or None, instance=task, **_form_kwargs(request),
    )
    if request.method == "POST" and form.is_valid():
        try:
            update_task(task, actor=request.user, **form.cleaned_data)
            messages.success(request, "Tache modifiee.")
            if _is_htmx(request):
                return _drawer_close_response(request, "tasks", "kpis")
            return redirect("secretary:task_list")
        except ValidationError as error:
            form.add_error(None, _validation_error_message(error))
    return _render_create_drawer_or_page(
        request,
        "secretary/drawers/form_drawer.html",
        "secretary/task_form.html",
        {
            "form": form,
            "page_title": "Modifier la tache",
            "modal_kicker": "Suivi",
            "modal_description": "Modifier les informations de la tache.",
            "submit_label": "Enregistrer les modifications",
            "return_url": "secretary:task_list",
            "form_action_url": "secretary:task_update",
            "form_action_pk": task.pk,
        },
    )


@login_required
@require_POST
def task_complete(request, pk):
    ensure_secretary_access(request.user)
    task = get_object_or_404(get_tasks_queryset({}, user=request.user), pk=pk)
    try:
        complete_task(task)
        messages.success(request, "Tache marquee comme terminee.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "tasks", "kpis")
    return redirect("secretary:task_list")


@login_required
def task_start(request, pk):
    ensure_secretary_access(request.user)
    if request.method != "POST":
        return redirect("secretary:task_list")
    task = get_object_or_404(get_tasks_queryset({}, user=request.user), pk=pk)
    try:
        start_task_processing(task, request.user)
        messages.success(request, "Tache prise en charge.")
    except ValidationError as error:
        messages.info(request, _validation_error_message(error))
    if _is_htmx(request):
        return _refresh_response(request, "tasks", "kpis")
    return redirect("secretary:task_list")


@login_required
def student_snapshot_view(request, student_id):
    ensure_secretary_access(request.user)
    snapshot = get_student_snapshot(student_id, user=request.user)
    # Get student object for schedule lookup
    student = get_student_snapshot_queryset(student_id, user=request.user).first()
    # Load weekly schedule (current week) so secretary can see if student is in class
    student_schedule = None
    if student:
        from django.utils import timezone
        from academics.services.schedule_service import get_student_week_schedule
        student_schedule = get_student_week_schedule(student, timezone.localdate())
    if _is_htmx(request):
        return render(request, "secretary/partials/student_drawer.html", {
            "snapshot": snapshot,
            "student_schedule": student_schedule,
        })
    return JsonResponse(snapshot)


COMMAND_PALETTE_ACTIONS = [
    {"label": "Nouvelle entree registre", "icon": "fa-solid fa-book-medical", "keywords": "registre entree accueil parent paiement colis", "url_name": "registry_create", "drawer": True},
    {"label": "Nouveau rendez-vous", "icon": "fa-regular fa-calendar-plus", "keywords": "rdv rendez-vous agenda", "url_name": "appointment_create", "drawer": True},
    {"label": "Nouvelle visite", "icon": "fa-solid fa-user-plus", "keywords": "visite visiteur arrivee sortie", "url_name": "visitor_create", "drawer": True},
    {"label": "Nouveau depot / document", "icon": "fa-solid fa-box", "keywords": "depot document colis livraison diplome", "url_name": "document_receipt_create", "drawer": True},
    {"label": "Nouvelle tache", "icon": "fa-solid fa-list-check", "keywords": "tache todo action", "url_name": "task_create", "drawer": True},
    {"label": "Vue Kanban du registre", "icon": "fa-solid fa-table-columns", "keywords": "kanban registre tableau colonnes", "url_name": "registry_kanban", "drawer": False},
    {"label": "Registre administratif", "icon": "fa-solid fa-book-open", "keywords": "registre liste journal", "url_name": "registry_list", "drawer": False},
    {"label": "Rendez-vous", "icon": "fa-regular fa-calendar", "keywords": "rendez-vous agenda liste", "url_name": "appointment_list", "drawer": False},
    {"label": "Visites", "icon": "fa-solid fa-door-open", "keywords": "visites liste entrees sorties", "url_name": "visitor_list", "drawer": False},
    {"label": "Documents et depots", "icon": "fa-solid fa-box-archive", "keywords": "documents depots liste", "url_name": "document_receipt_list", "drawer": False},
    {"label": "Taches du secretariat", "icon": "fa-solid fa-list-check", "keywords": "taches liste todo", "url_name": "task_list", "drawer": False},
    {"label": "Ouvrir une classe", "icon": "fa-solid fa-chalkboard-user", "keywords": "classe eleves etudiants effectif", "url_name": "secretary_dashboard", "url_query": "?section=classes", "drawer": False},
]


@login_required
def htmx_command_palette(request):
    ensure_secretary_access(request.user)
    query = request.GET.get("q", "").strip()
    students = search_students(query, user=request.user)[:5] if len(query) >= 2 else []
    query_lower = query.lower()
    actions = []
    for action in COMMAND_PALETTE_ACTIONS:
        if query and query_lower not in action["label"].lower() and query_lower not in action["keywords"]:
            continue
        actions.append({
            **action,
            "url": reverse(f"secretary:{action['url_name']}") + action.get("url_query", ""),
        })
    return render(
        request,
        "secretary/partials/command_palette_results.html",
        {"query": query, "students": students, "actions": actions},
    )


@login_required
def htmx_student_results(request):
    ensure_secretary_access(request.user)
    query = request.GET.get("q", "").strip()
    students_queryset = search_students(query, user=request.user) if query else get_active_students(user=request.user)
    page_obj = _paginate(request, students_queryset, per_page=10)
    return render(
        request,
        "secretary/partials/student_results.html",
        {"students": page_obj.object_list, "page_obj": page_obj, "query": query},
    )


@login_required
def htmx_class_results(request):
    ensure_secretary_access(request.user)
    query = request.GET.get("q", "").strip()
    classes_queryset = search_academic_classes(query, user=request.user)
    page_obj = _paginate(request, classes_queryset, per_page=12)
    return render(
        request,
        "secretary/partials/class_results.html",
        {"classes": page_obj.object_list, "page_obj": page_obj, "query": query},
    )


@login_required
def htmx_class_students(request, class_id):
    ensure_secretary_access(request.user)
    academic_class, students = get_secretary_class_students(class_id, user=request.user)
    page_obj = _paginate(request, students, per_page=15)
    return render(
        request,
        "secretary/partials/class_students.html",
        {"academic_class": academic_class, "students": page_obj.object_list, "page_obj": page_obj},
    )


@login_required
def htmx_registry_results(request):
    ensure_secretary_access(request.user)
    filters = _common_filters(request)
    page_obj = _paginate(request, get_registry_queryset(filters, user=request.user), per_page=12)
    return render(
        request,
        "secretary/partials/registry_results.html",
        {"entries": page_obj.object_list, "page_obj": page_obj, "query": filters.get("q", "")},
    )


@login_required
def htmx_appointment_results(request):
    ensure_secretary_access(request.user)
    filters = _common_filters(request)
    page_obj = _paginate(request, get_appointments_queryset(filters, user=request.user), per_page=12)
    return render(
        request,
        "secretary/partials/appointment_results.html",
        {"appointments": page_obj.object_list, "page_obj": page_obj, "query": filters.get("q", "")},
    )


@login_required
def htmx_document_results(request):
    ensure_secretary_access(request.user)
    filters = _common_filters(request)
    page_obj = _paginate(request, get_documents_queryset(filters, user=request.user), per_page=12)
    return render(
        request,
        "secretary/partials/document_results.html",
        {"receipts": page_obj.object_list, "page_obj": page_obj, "query": filters.get("q", "")},
    )


@login_required
def htmx_task_results(request):
    ensure_secretary_access(request.user)
    filters = _common_filters(request)
    page_obj = _paginate(request, get_tasks_queryset(filters, user=request.user), per_page=12)
    return render(
        request,
        "secretary/partials/task_results.html",
        {"tasks": page_obj.object_list, "page_obj": page_obj, "query": filters.get("q", "")},
    )


@login_required
def htmx_messages_panel(request):
    ensure_secretary_access(request.user)
    messages_list = get_user_notifications(
        request.user,
        limit=8,
        channel=CommunicationNotification.CHANNEL_IN_APP,
    )
    return render(
        request,
        "secretary/partials/messages_panel.html",
        {
            "recent_messages": messages_list,
            "unread_count": get_user_unread_count(request.user),
        },
    )


@login_required
def htmx_dashboard_scope(request):
    context = _dashboard_base_context(request)
    return render(request, "secretary/partials/dashboard_scope.html", context)


@login_required
def htmx_dashboard_kpis(request):
    context = _dashboard_base_context(request)
    return render(request, "secretary/partials/dashboard_kpis.html", context)


@login_required
def htmx_sidebar_counters(request):
    context = _dashboard_base_context(request)
    context["messages_count"] = get_user_unread_count(request.user)
    context["active_section"] = _dashboard_section(request)
    return render(request, "secretary/partials/sidebar_counters.html", context)


@login_required
def htmx_overview_registry(request):
    context = _dashboard_base_context(request)
    branch = context.get("branch")
    context["dashboard_pending_registry_page"] = _paginate(
        request,
        get_registry_queryset(
            {"status": RegistryEntry.STATUS_PENDING, "archived": False, "active_only": True},
            user=request.user,
            branch=branch,
        ),
        per_page=6,
        page_param="overview_registry_page",
    )
    context["pending_registry_rows"] = context["dashboard_pending_registry_page"].object_list
    return render(request, "secretary/partials/dashboard_overview_registry.html", context)


@login_required
def htmx_overview_appointments(request):
    context = _dashboard_base_context(request)
    return render(request, "secretary/partials/dashboard_overview_appointments.html", context)


@login_required
def htmx_overview_visits(request):
    context = _dashboard_base_context(request)
    return render(request, "secretary/partials/dashboard_overview_visits.html", context)


@login_required
def htmx_overview_tasks(request):
    context = _dashboard_base_context(request)
    return render(request, "secretary/partials/dashboard_overview_tasks.html", context)


@login_required
def htmx_visits_open(request):
    context = _dashboard_base_context(request)
    branch = context.get("branch")
    context["open_visits_page"] = _paginate(
        request,
        get_active_visits_queryset(user=request.user, branch=branch),
        per_page=10,
        page_param="open_visits_page",
    )
    context["open_visits_rows"] = context["open_visits_page"].object_list
    return render(request, "secretary/partials/dashboard_visits_open.html", context)


@login_required
def htmx_appointments_today(request):
    context = _dashboard_base_context(request)
    branch = context.get("branch")
    context["appointments_page"] = _paginate(
        request,
        get_today_appointments_queryset(user=request.user, branch=branch),
        per_page=10,
        page_param="appointments_page",
    )
    context["today_appointments_rows"] = context["appointments_page"].object_list
    return render(request, "secretary/partials/dashboard_appointments_today.html", context)


@login_required
def htmx_documents_pending(request):
    context = _dashboard_base_context(request)
    branch = context.get("branch")
    context["documents_page"] = _paginate(
        request,
        get_documents_queryset(
            {"status": DocumentReceipt.STATUS_PENDING, "archived": False, "active_only": True},
            user=request.user,
            branch=branch,
        ),
        per_page=10,
        page_param="documents_page",
    )
    context["pending_documents_rows"] = context["documents_page"].object_list
    return render(request, "secretary/partials/dashboard_documents_pending.html", context)


@login_required
def daily_registry_report(request):
    ensure_secretary_access(request.user)
    branch = get_user_branch(request.user)
    report_date = request.GET.get("date") or timezone.localdate()
    entries = get_registry_queryset(
        {"archived": False, "active_only": True},
        user=request.user,
        branch=branch,
    ).filter(created_at__date=report_date)

    rows = [
        [
            entry.registry_number or entry.pk,
            entry.created_at.strftime("%H:%M"),
            entry.get_entry_type_display(),
            entry.visitor_name,
            entry.related_student.full_name if entry.related_student_id else "",
            entry.target_service,
            entry.motive,
            ", ".join(entry.linked_actions or []),
            entry.get_status_display(),
            entry.exited_at.strftime("%H:%M") if entry.exited_at else "",
        ]
        for entry in entries
    ]
    headers = ["N", "Heure", "Type", "Visiteur", "Etudiant", "Service", "Motif", "Action", "Statut", "Heure sortie"]

    if request.GET.get("format") == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registre journalier"
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for row in rows:
            sheet.append(row)
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 3, 38)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="registre_journalier.xlsx"'
        workbook.save(response)
        return response

    if request.GET.get("format") == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="registre_journalier.pdf"'
        document = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Rapport journalier du registre administratif", styles["Title"]),
            Paragraph(f"Annexe : {getattr(branch, 'name', 'Toutes annexes')} - Date : {report_date}", styles["Normal"]),
            Spacer(1, 12),
        ]
        table = Table([headers] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        document.build(story)
        return response

    return render(
        request,
        "secretary/daily_registry_report.html",
        {"entries": entries, "branch": branch, "report_date": report_date},
    )
