from __future__ import annotations

import io

from django.db.models import QuerySet

from academics.models import AcademicClass, AcademicEnrollment, EC, Semester


def _get_class_enrollments(academic_class: AcademicClass) -> QuerySet[AcademicEnrollment]:
    return (
        AcademicEnrollment.objects.filter(
            academic_class=academic_class,
            academic_year=academic_class.academic_year,
            is_active=True,
        )
        .select_related(
            "student__student_profile__inscription__candidature",
            "student",
            "academic_class",
            "academic_year",
        )
        .order_by(
            "student__student_profile__inscription__candidature__last_name",
            "student__student_profile__inscription__candidature__first_name",
        )
    )


def _get_semester_ecs(semester: Semester) -> QuerySet[EC]:
    return (
        EC.objects.filter(ue__semester=semester)
        .select_related("ue")
        .order_by("ue__id", "id")
    )


def _get_student_last_name(enrollment: AcademicEnrollment) -> str:
    profile = getattr(enrollment.student, "student_profile", None)
    candidature = getattr(getattr(profile, "inscription", None), "candidature", None) if profile else None
    return (getattr(candidature, "last_name", "") or "").strip()


def _get_student_first_name(enrollment: AcademicEnrollment) -> str:
    profile = getattr(enrollment.student, "student_profile", None)
    candidature = getattr(getattr(profile, "inscription", None), "candidature", None) if profile else None
    return (getattr(candidature, "first_name", "") or "").strip()


def _get_student_matricule(enrollment: AcademicEnrollment) -> str:
    profile = getattr(enrollment.student, "student_profile", None)
    return (getattr(profile, "matricule", "") or "").strip()


def _get_ec_human_label(ec: EC) -> str:
    return f"{ec.ue.code} - {ec.title}"


def _get_ec_note_label(ec: EC) -> str:
    return f"NOTE /20 - {_get_ec_human_label(ec)}"


def generate_import_template(academic_class: AcademicClass, semester: Semester) -> io.BytesIO:
    """Genere un template Excel lisible pour l'import des notes."""

    if semester.academic_class_id != academic_class.id:
        raise ValueError("Le semestre ne correspond pas a la classe academique.")

    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = f"Import S{semester.number}"

    enrollments = list(_get_class_enrollments(academic_class))
    ecs = list(_get_semester_ecs(semester))

    ws.append(["Ecole", "ESFE"])
    ws.append(["Classe", academic_class.display_name])
    ws.append(["Annee academique", str(academic_class.academic_year)])
    ws.append(["Semestre", f"Semestre {semester.number} - saisir les notes dans les colonnes jaunes NOTE /20"])

    headers = ["ENROLLMENT_ID", "MATRICULE", "NOM", "PRENOM", *[_get_ec_note_label(ec) for ec in ecs]]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    note_header_fill = PatternFill("solid", fgColor="D97706")
    note_fill = PatternFill("solid", fgColor="FEF3C7")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    header_row = 5
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = note_header_fill if col_idx >= 5 else header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        if col_idx >= 5:
            cell.comment = Comment("Saisir ici la note de cette matiere sur 20. Exemple: 14,5", "ESFE")

    for enrollment in enrollments:
        ws.append([
            enrollment.id,
            _get_student_matricule(enrollment),
            _get_student_last_name(enrollment),
            _get_student_first_name(enrollment),
            *([""] * len(ecs)),
        ])

    ws.freeze_panes = "E6"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.row_dimensions[5].height = 42
    for col_idx in range(5, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = 32

    if ecs and enrollments:
        validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="20",
            allow_blank=True,
        )
        validation.error = "La note doit etre comprise entre 0 et 20."
        validation.errorTitle = "Note invalide"
        validation.prompt = "Entrer la note de la matiere sur 20."
        validation.promptTitle = "Note /20"
        ws.add_data_validation(validation)
        first_note_cell = ws.cell(row=6, column=5).coordinate
        last_note_cell = ws.cell(row=ws.max_row, column=len(headers)).coordinate
        validation.add(f"{first_note_cell}:{last_note_cell}")

    for row_idx in range(6, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).alignment = align_left
        ws.cell(row=row_idx, column=2).alignment = align_left
        ws.cell(row=row_idx, column=3).alignment = align_left
        ws.cell(row=row_idx, column=4).alignment = align_left
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).border = thin_border
        for col_idx in range(5, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = align_center
            ws.cell(row=row_idx, column=col_idx).number_format = "0.00"
            ws.cell(row=row_idx, column=col_idx).fill = note_fill
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
